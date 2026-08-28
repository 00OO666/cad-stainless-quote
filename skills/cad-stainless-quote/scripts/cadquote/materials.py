"""Material-table parsing and MT code normalization.

The source workbooks used by interior-design teams are rarely tidy databases.  A
material may occupy one row in a conventional table, or a small key/value block
(``材料编号`` on one row, ``材料名称`` on the next).  This module accepts either
shape and keeps every parsed value traceable to its source row.

Parsing is deliberately conservative.  A missing value stays missing and a
conflict is recorded; neither condition is silently repaired from another
project.
"""

from __future__ import annotations

import hashlib
import json
import math
import re
import stat
import unicodedata
import zipfile
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any, Literal
from xml.etree import ElementTree

from .models import CadEntity, MaterialSpec, ReviewStatus

_CODE_SEPARATOR = r"[-—–－_:/／\\]"
_MATERIAL_CODE_RE = re.compile(
    rf"(?<![A-Z0-9])(?P<raw>(?P<prefix>[A-Z]{{1,8}}(?:\s*{_CODE_SEPARATOR}\s*"
    rf"[A-Z]{{1,8}})*)\s*{_CODE_SEPARATOR}?\s*(?P<number>\d{{1,4}}))(?!\d)",
    re.I,
)
_NUMBER_ONLY_RE = re.compile(r"^\s*0*(\d{1,3})\s*$")
_THICKNESS_RE = re.compile(
    r"(?<!\d)(\d+(?:\.\d+)?)\s*"
    r"(?:(?:mm|毫米)\s*(?:厚|thk|t)?|(?:厚|thk))(?![A-Z0-9])",
    re.I,
)
_GRADE_RE = re.compile(r"(?<!\d)(201|304L?|316L?|430)(?!\d)", re.I)
_FINISH_TERMS = (
    "黑色镜面",
    "镜面黑色",
    "青古铜",
    "红古铜",
    "古铜色",
    "苹果砂",
    "玫瑰金",
    "镜面",
    "拉丝",
    "喷砂",
    "烤漆",
    "镀色",
)
_PROCESS_TERMS = ("折弯", "刨槽", "雕花", "蚀刻", "满焊", "点焊", "焊接", "激光")
_CAD_MATERIAL_RE = re.compile(
    r"不锈钢|钢板|金属|铝板|铝合金|铜板|铁板|玻璃|钛金|玫瑰金|镜面|拉丝|喷砂|"
    r"苹果砂|古铜|烤漆|镀色|雕花|蚀刻"
)
_DOCX_NON_DESCRIPTION_RE = re.compile(
    r"^(?:[￥¥$]\s*)?\d+(?:\.\d+)?\s*(?:元|元/(?:m|㎡|m2)|mm|cm|m|kg|%|号)?$",
    re.I,
)
_DOCX_USAGE_RE = re.compile(
    r"\s+(?=[^\s]*(?:踢脚|脚线|顶线|线条|门套|窗套|包板|收口|嵌条|墙面|顶面|"
    r"屏风|柜|台|栏杆|压条|层架|挂衣杆))"
)
_OOXML_WORD_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
_OOXML_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_DOCX_DOCUMENT_MEMBER = "word/document.xml"
_DOCX_MAX_DOCUMENT_XML_BYTES = 64 * 1024 * 1024
_DOCX_MAX_XML_COMPRESSION_RATIO = 1_000.0

DEFAULT_STAINLESS_CODE_FAMILIES = frozenset({"MT", "GC-SS"})
DEFAULT_REVIEW_CODE_FAMILIES = frozenset({"GC-MT"})
# Auxiliary families are parsed from material schedules so they can be bound to
# a composite assembly.  They remain absent from the default occurrence policy,
# therefore a glass code never creates an independent stainless quotation row.
DEFAULT_AUXILIARY_CODE_FAMILIES = frozenset({"GC-GL"})


@dataclass(frozen=True)
class MaterialCodeMatch:
    """A material code plus its source spelling and review disposition."""

    raw_code: str
    normalized_code: str
    family: str
    disposition: Literal["stainless", "review"]


_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "mt_code": ("MT编号", "材料编号", "物料编号", "色号", "编号", "NUMBER", "MATERIALNO"),
    "name": ("材料名称", "物料名称", "名称", "品名", "NAME", "DESCRIPTION"),
    "grade": ("材质", "牌号", "GRADE", "MATERIAL"),
    "thickness_mm": ("厚度", "材料规格", "物料规格", "规格", "SIZE", "THICKNESS", "THK"),
    "finish": ("表面处理", "饰面", "表面", "颜色", "FINISH", "COLOR", "COLOUR"),
    "process": ("加工方式", "加工工艺", "工艺", "PROCESS", "CRAFT"),
    "brand": ("材料品牌", "物料品牌", "品牌", "BRAND"),
    "model": ("材料型号", "物料型号", "型号", "MODEL"),
}


def normalize_text(value: Any) -> str:
    """Return stable human text for heterogeneous spreadsheet cells."""

    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        if value.is_integer():
            value = int(value)
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\u00a0", " ").replace("\u3000", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_material_code_family(value: Any) -> str | None:
    """Normalize a code family (``ＧＣ＿ＳＳ`` → ``GC-SS``), without a number."""

    text = normalize_text(value).upper()
    text = re.sub(rf"\s*{_CODE_SEPARATOR}\s*", "-", text).strip("-")
    if not re.fullmatch(r"[A-Z]{1,8}(?:-[A-Z]{1,8})*", text):
        return None
    return text


def _family_sets(
    stainless_families: Iterable[str] | None,
    review_families: Iterable[str] | None,
) -> tuple[frozenset[str], frozenset[str]]:
    stainless_values = (
        DEFAULT_STAINLESS_CODE_FAMILIES
        if stainless_families is None
        else frozenset(
            family
            for value in stainless_families
            if (family := normalize_material_code_family(value)) is not None
        )
    )
    review_values = (
        DEFAULT_REVIEW_CODE_FAMILIES
        if review_families is None
        else frozenset(
            family
            for value in review_families
            if (family := normalize_material_code_family(value)) is not None
        )
    )
    # Explicit stainless configuration wins over the conservative review set.
    return frozenset(stainless_values), frozenset(review_values - stainless_values)


def material_code_disposition(
    family: Any,
    *,
    stainless_families: Iterable[str] | None = None,
    review_families: Iterable[str] | None = None,
) -> Literal["stainless", "review"] | None:
    """Classify a family according to the active stainless/review policy."""

    normalized = normalize_material_code_family(family)
    stainless, review = _family_sets(stainless_families, review_families)
    if normalized in stainless:
        return "stainless"
    if normalized in review:
        return "review"
    return None


def find_material_codes(
    value: Any,
    *,
    stainless_families: Iterable[str] | None = None,
    review_families: Iterable[str] | None = None,
) -> list[MaterialCodeMatch]:
    """Find configured material-code families without treating every code as steel.

    ``MT`` and ``GC-SS`` are stainless candidates by default. ``GC-MT`` is
    retained as a lower-confidence review family. Unconfigured families such as
    ``GC-GL`` and ``GC-MR`` are deliberately ignored.
    """

    stainless, review = _family_sets(stainless_families, review_families)
    text = normalize_text(value).upper()
    source_spelling = str(value).strip() if value is not None else ""
    result: list[MaterialCodeMatch] = []
    seen: set[str] = set()
    for match in _MATERIAL_CODE_RE.finditer(text):
        family = normalize_material_code_family(match.group("prefix"))
        if family not in stainless and family not in review:
            continue
        number_text = match.group("number")
        width = max(2, len(number_text))
        code = f"{family}-{int(number_text):0{width}d}"
        if code in seen:
            continue
        seen.add(code)
        raw_code = match.group("raw").strip()
        if match.span("raw") == (0, len(text)) and normalize_text(source_spelling).upper() == text:
            raw_code = source_spelling
        result.append(
            MaterialCodeMatch(
                raw_code=raw_code,
                normalized_code=code,
                family=family,
                disposition="stainless" if family in stainless else "review",
            )
        )
    return result


def normalize_material_code(
    value: Any,
    *,
    stainless_families: Iterable[str] | None = None,
    review_families: Iterable[str] | None = None,
) -> str | None:
    """Return the first configured normalized material code in ``value``."""

    matches = find_material_codes(
        value,
        stainless_families=stainless_families,
        review_families=review_families,
    )
    return matches[0].normalized_code if matches else None


def normalize_mt_code(value: Any) -> str | None:
    """Normalize legacy ``MT01``/``ＭＴ－０１``/``MT 1`` to ``MT-01`` only."""

    return normalize_material_code(value, stainless_families={"MT"}, review_families=())


def find_mt_codes(value: Any) -> list[str]:
    """Find only legacy MT codes, preserving the pre-generalization API."""

    return [
        match.normalized_code
        for match in find_material_codes(
            value,
            stainless_families={"MT"},
            review_families=(),
        )
    ]


def find_material_codes_in_cells(
    cells: Sequence[Any],
    *,
    stainless_families: Iterable[str] | None = None,
    review_families: Iterable[str] | None = None,
) -> list[tuple[MaterialCodeMatch, tuple[int, ...]]]:
    """Find explicit codes and immediately adjacent family/number cell pairs."""

    stainless, review = _family_sets(stainless_families, review_families)
    texts = [normalize_text(value) for value in cells]
    found: list[tuple[MaterialCodeMatch, tuple[int, ...]]] = []
    for index, cell_text in enumerate(texts):
        for code in find_material_codes(
            cell_text,
            stainless_families=stainless,
            review_families=review,
        ):
            found.append((code, (index,)))
        family = normalize_material_code_family(cell_text)
        if family not in stainless and family not in review:
            continue
        for neighbour in (index + 1, index - 1):
            if neighbour < 0 or neighbour >= len(texts):
                continue
            match = _NUMBER_ONLY_RE.fullmatch(texts[neighbour])
            if match:
                number_text = match.group(1)
                width = max(2, len(number_text))
                code = f"{family}-{int(number_text):0{width}d}"
                found.append(
                    (
                        MaterialCodeMatch(
                            raw_code=f"{cell_text} {texts[neighbour]}".strip(),
                            normalized_code=code,
                            family=family,
                            disposition="stainless" if family in stainless else "review",
                        ),
                        (index, neighbour),
                    )
                )
                break

    unique: list[tuple[MaterialCodeMatch, tuple[int, ...]]] = []
    seen: set[tuple[str, tuple[int, ...]]] = set()
    for item in found:
        identity = (item[0].normalized_code, item[1])
        if identity not in seen:
            seen.add(identity)
            unique.append(item)
    return unique


def find_mt_codes_in_cells(cells: Sequence[Any]) -> list[tuple[str, tuple[int, ...]]]:
    """Find legacy MT codes in cells, preserving the pre-generalization API."""

    return [
        (match.normalized_code, indexes)
        for match, indexes in find_material_codes_in_cells(
            cells,
            stainless_families={"MT"},
            review_families=(),
        )
    ]


def _compact_label(value: Any) -> str:
    text = normalize_text(value).upper()
    text = re.sub(r"[\s:：()（）\[\]【】._/\\-]+", "", text)
    return text


def _field_for_label(value: Any) -> str | None:
    label = _compact_label(value)
    if not label:
        return None
    for field, aliases in _FIELD_ALIASES.items():
        for alias in aliases:
            compact = _compact_label(alias)
            ascii_alias = compact.isascii()
            if (
                label == compact
                or label.startswith(compact)
                or (ascii_alias and label.endswith(compact))
                or (ascii_alias and len(compact) >= 4 and compact in label)
            ):
                return field
    return None


def _clean_value(value: Any) -> str | None:
    text = normalize_text(value)
    # Some legacy BIFF files advertise UTF-16 while containing text that xlrd
    # can only decode as U+FFFD.  Treat replacement-heavy output as missing
    # evidence instead of exporting mojibake as a material name.
    if text.count("\ufffd") >= 2 and text.count("\ufffd") / max(1, len(text)) >= 0.15:
        return None
    return text or None


def _parse_thickness(*values: Any) -> float | None:
    for value in values:
        text = normalize_text(value)
        match = _THICKNESS_RE.search(text)
        if match:
            thickness = float(match.group(1))
            if 0 < thickness <= 20:
                return thickness
        # In a column explicitly named 厚度, bare numeric cells are common.
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            thickness = float(value)
            if math.isfinite(thickness) and 0 < thickness <= 20:
                return thickness
    return None


def _parse_grade(*values: Any) -> str | None:
    for value in values:
        match = _GRADE_RE.search(normalize_text(value).upper())
        if match:
            return match.group(1).upper()
    return None


def _first_term(values: Sequence[Any], terms: Sequence[str]) -> str | None:
    for value in values:
        text = normalize_text(value)
        for term in terms:
            if term in text:
                return term
    return None


def _stable_id(prefix: str, payload: Mapping[str, Any]) -> str:
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    digest = hashlib.sha256(encoded.encode("utf-8")).hexdigest()[:24]
    return f"{prefix}:{digest}"


def _row_values(row: Mapping[str, Any] | Sequence[Any]) -> list[Any]:
    if isinstance(row, Mapping):
        return list(row.values())
    if isinstance(row, (str, bytes)):
        return [row]
    return list(row)


def _mapping_record(row: Mapping[str, Any]) -> dict[str, Any]:
    record: dict[str, Any] = {}
    for key, value in row.items():
        field = _field_for_label(key)
        if field and _clean_value(value) is not None:
            record[field] = value
    return record


def _header_map(cells: Sequence[Any]) -> dict[int, str]:
    result: dict[int, str] = {}
    for index, value in enumerate(cells):
        field = _field_for_label(value)
        if field:
            result[index] = field
    return result


def _extract_pairs(rows: Sequence[Sequence[Any]]) -> dict[str, Any]:
    """Extract repeated label/value pairs from a material-card row block."""

    record: dict[str, Any] = {}
    for row in rows:
        cells = list(row)
        for index, cell in enumerate(cells):
            field = _field_for_label(cell)
            if not field:
                continue
            for candidate in cells[index + 1 :]:
                if _field_for_label(candidate):
                    break
                if _clean_value(candidate) is not None:
                    record.setdefault(field, candidate)
                    break
    return record


def _make_spec(
    raw: Mapping[str, Any],
    *,
    code: str,
    raw_code: str | None = None,
    code_family: str | None = None,
    source_file_id: str | None,
    source_location: str,
    source_type: str | None = None,
    source_sha256: str | None = None,
    source_evidence: Sequence[str] = (),
) -> MaterialSpec:
    name = _clean_value(raw.get("name"))
    raw_grade = _clean_value(raw.get("grade"))
    grade = _parse_grade(raw_grade, raw.get("name"), raw.get("thickness_mm")) or raw_grade
    thickness = _parse_thickness(raw.get("thickness_mm"), raw.get("name"))
    finish = _clean_value(raw.get("finish")) or _first_term((raw.get("name"),), _FINISH_TERMS)
    process = _clean_value(raw.get("process")) or _first_term((raw.get("name"),), _PROCESS_TERMS)
    brand = _clean_value(raw.get("brand"))
    model = _clean_value(raw.get("model"))
    payload = {
        "mt_code": code,
        "raw_material_code": raw_code or code,
        "material_code_family": code_family or code.rsplit("-", 1)[0],
        "name": name,
        "grade": grade,
        "thickness_mm": thickness,
        "finish": finish,
        "process": process,
        "brand": brand,
        "model": model,
        "source_file_id": source_file_id,
        "source_type": source_type,
        "source_sha256": source_sha256,
        "source_location": source_location,
        "source_evidence": sorted(set(source_evidence)),
    }
    return MaterialSpec(
        id=_stable_id("material", payload),
        **payload,
        status=ReviewStatus.REVIEW,
    )


def _choose_primary_code(
    cells: Sequence[Any],
    codes: Sequence[tuple[MaterialCodeMatch, tuple[int, ...]]],
) -> MaterialCodeMatch:
    """Prefer the code adjacent to a NUMBER/编号 label over cross references."""

    for code, indexes in codes:
        first = min(indexes)
        for label_index in range(max(0, first - 2), first):
            if _field_for_label(cells[label_index]) == "mt_code":
                return code
    return codes[0][0]


def _annotate_conflicts(specs: list[MaterialSpec]) -> list[MaterialSpec]:
    by_code: dict[str, list[MaterialSpec]] = {}
    for spec in specs:
        by_code.setdefault(spec.mt_code, []).append(spec)

    fields = ("name", "grade", "thickness_mm", "finish", "process", "brand", "model")
    output: list[MaterialSpec] = []
    for spec in specs:
        conflicts: list[str] = list(spec.conflicts)
        peers = by_code[spec.mt_code]
        for field in fields:
            values = {
                normalize_text(getattr(peer, field)).casefold()
                for peer in peers
                if getattr(peer, field) is not None
            }
            if len(values) > 1:
                rendered = sorted(
                    {
                        normalize_text(getattr(peer, field))
                        for peer in peers
                        if getattr(peer, field) is not None
                    }
                )
                conflicts.append(f"{field}: {' | '.join(rendered)}")
        output.append(spec.model_copy(update={"conflicts": sorted(set(conflicts))}))
    return output


def annotate_material_conflicts(specs: Iterable[MaterialSpec]) -> list[MaterialSpec]:
    """Annotate conflicts across workbook and CAD-derived material records."""

    unique = {spec.id: spec for spec in specs}
    return _annotate_conflicts(list(unique.values()))


def _entity_center(entity: CadEntity) -> tuple[float, float] | None:
    if entity.insert is not None:
        return float(entity.insert[0]), float(entity.insert[1])
    if entity.bbox is not None:
        return (
            float(entity.bbox[0] + entity.bbox[2]) / 2.0,
            float(entity.bbox[1] + entity.bbox[3]) / 2.0,
        )
    return None


def parse_cad_material_specs(entities: Iterable[CadEntity]) -> list[MaterialSpec]:
    """Recover conservative MT/material rows from native CAD text tables.

    A material table usually places a material code and its description on the same
    horizontal baseline. This parser deliberately requires that row evidence;
    nearest-text distance alone is not enough because adjacent table rows often
    contain different MT definitions.
    """

    material_review_families = (
        DEFAULT_REVIEW_CODE_FAMILIES | DEFAULT_AUXILIARY_CODE_FAMILIES
    )
    grouped: dict[tuple[str, str | None, str], list[CadEntity]] = {}
    for entity in entities:
        if not entity.text or _entity_center(entity) is None:
            continue
        key = (entity.source_file_id, entity.sheet_id, entity.space)
        grouped.setdefault(key, []).append(entity)

    specs: list[MaterialSpec] = []
    seen_original_seeds: set[tuple[str, str]] = set()
    for (source_id, sheet_id, _space), values in sorted(grouped.items()):
        heights = [
            abs(entity.bbox[3] - entity.bbox[1])
            for entity in values
            if entity.bbox and abs(entity.bbox[3] - entity.bbox[1]) > 0
        ]
        typical_height = sorted(heights)[len(heights) // 2] if heights else 1.0
        for seed in sorted(values, key=lambda value: value.id):
            codes = find_material_codes(
                seed.text,
                review_families=material_review_families,
            )
            if len(codes) != 1:
                continue
            code_match = codes[0]
            seed_identity = str(seed.geometry.get("original_entity_id") or seed.id)
            identity = (source_id, seed_identity)
            if identity in seen_original_seeds:
                continue
            seed_text = normalize_text(seed.text)
            seed_point = _entity_center(seed)
            assert seed_point is not None
            seed_height = (
                abs(seed.bbox[3] - seed.bbox[1])
                if seed.bbox and abs(seed.bbox[3] - seed.bbox[1]) > 0
                else typical_height
            )
            row_tolerance = max(0.5, min(typical_height * 2.0, seed_height * 2.5))
            row_values: list[tuple[float, float, str, CadEntity]] = []
            for candidate in values:
                if candidate.id == seed.id or not candidate.text:
                    continue
                text = normalize_text(candidate.text)
                if not text or find_material_codes(
                    text,
                    review_families=material_review_families,
                ):
                    continue
                point = _entity_center(candidate)
                if point is None:
                    continue
                dy = abs(point[1] - seed_point[1])
                if dy > row_tolerance:
                    continue
                if not _CAD_MATERIAL_RE.search(text):
                    continue
                dx = point[0] - seed_point[0]
                # Table descriptions normally sit to the right. A left-side
                # value is accepted only as a weaker fallback.
                direction_penalty = 0.0 if dx >= 0 else abs(dx) * 0.35
                row_values.append((abs(dx) + direction_penalty, dy, candidate.id, candidate))

            direct_description = seed_text if _CAD_MATERIAL_RE.search(seed_text) else None
            if not row_values and direct_description is None:
                continue
            descriptor = min(row_values)[3] if row_values else seed
            name = direct_description or normalize_text(descriptor.text)
            # Do not borrow finish/process terms from another material phrase
            # merely because two independent tables happen to share a baseline.
            combined = " | ".join(dict.fromkeys([seed_text, name]))
            raw = {
                "name": name,
                "grade": _parse_grade(combined),
                "thickness_mm": _parse_thickness(combined),
                "finish": _first_term((combined,), _FINISH_TERMS),
                "process": _first_term((combined,), _PROCESS_TERMS),
            }
            entity_ids = [seed.id]
            if descriptor.id != seed.id:
                entity_ids.append(descriptor.id)
            specs.append(
                _make_spec(
                    raw,
                    code=code_match.normalized_code,
                    raw_code=code_match.raw_code,
                    code_family=code_match.family,
                    source_file_id=source_id,
                    source_location=(
                        f"cad:{sheet_id or '<unassigned>'}:entities=" + ",".join(sorted(entity_ids))
                    ),
                    source_type="cad_material_table",
                    source_evidence=sorted(entity_ids),
                )
            )
            seen_original_seeds.add(identity)
    return annotate_material_conflicts(specs)


def parse_material_rows(
    rows: Iterable[Mapping[str, Any] | Sequence[Any]],
    *,
    source_file_id: str | None = None,
    sheet_name: str = "Sheet1",
    start_row: int = 1,
) -> list[MaterialSpec]:
    """Parse a conventional table or material-card blocks into records.

    ``rows`` may be dictionaries, tuples from ``openpyxl``/``xlrd``, or text
    cells copied from a PDF table.  Row numbers in ``source_location`` remain
    one-based and deterministic.
    """

    materialized = list(rows)
    if not materialized:
        return []

    specs: list[MaterialSpec] = []
    material_review_families = (
        DEFAULT_REVIEW_CODE_FAMILIES | DEFAULT_AUXILIARY_CODE_FAMILIES
    )

    # Dictionary rows already carry their own headers.
    if all(isinstance(row, Mapping) for row in materialized):
        for offset, row in enumerate(materialized):
            assert isinstance(row, Mapping)
            raw = _mapping_record(row)
            codes = find_material_codes_in_cells(
                list(row.values()),
                review_families=material_review_families,
            )
            explicit_matches = find_material_codes(
                raw.get("mt_code"),
                review_families=material_review_families,
            )
            code_match = (
                explicit_matches[0] if explicit_matches else (codes[0][0] if codes else None)
            )
            if code_match:
                location = f"{sheet_name}!row={start_row + offset}"
                specs.append(
                    _make_spec(
                        raw,
                        code=code_match.normalized_code,
                        raw_code=code_match.raw_code,
                        code_family=code_match.family,
                        source_file_id=source_file_id,
                        source_location=location,
                    )
                )
        return _annotate_conflicts(specs)

    table = [_row_values(row) for row in materialized]

    # Parse conventional tabular sections.  A header is credible only when it
    # contains a material-code column and at least one descriptive field.
    consumed_rows: set[int] = set()
    active_header: dict[int, str] | None = None
    for offset, cells in enumerate(table):
        candidate = _header_map(cells)
        nonempty_count = sum(bool(normalize_text(cell)) for cell in cells)
        is_header_row = (
            "mt_code" in candidate.values()
            and len(set(candidate.values())) >= 2
            and len(candidate) / max(1, nonempty_count) >= 0.75
        )
        if is_header_row:
            active_header = candidate
            consumed_rows.add(offset)
            continue
        if not active_header:
            continue
        raw = {
            field: cells[index]
            for index, field in active_header.items()
            if index < len(cells) and _clean_value(cells[index]) is not None
        }
        code_matches = find_material_codes(
            raw.get("mt_code"),
            review_families=material_review_families,
        )
        code_match = code_matches[0] if code_matches else None
        if code_match:
            location = f"{sheet_name}!row={start_row + offset}"
            specs.append(
                _make_spec(
                    raw,
                    code=code_match.normalized_code,
                    raw_code=code_match.raw_code,
                    code_family=code_match.family,
                    source_file_id=source_file_id,
                    source_location=location,
                )
            )
            consumed_rows.add(offset)
        elif any(_field_for_label(cell) for cell in cells):
            active_header = None

    # Parse key/value material cards.  Each primary code row starts a block
    # ending immediately before the next primary code row.
    starts: list[tuple[int, MaterialCodeMatch]] = []
    for offset, cells in enumerate(table):
        if offset in consumed_rows:
            continue
        codes = find_material_codes_in_cells(
            cells,
            review_families=material_review_families,
        )
        if not codes:
            continue
        primary = _choose_primary_code(cells, codes)
        # Cross references inside a free-text description must not start a card.
        has_number_label = any(_field_for_label(cell) == "mt_code" for cell in cells)
        has_code_cell = any(
            find_material_codes(
                cell,
                review_families=material_review_families,
            )
            and _MATERIAL_CODE_RE.fullmatch(normalize_text(cell))
            for cell in cells
        )
        if has_number_label or has_code_cell:
            starts.append((offset, primary))

    for index, (offset, code_match) in enumerate(starts):
        end = starts[index + 1][0] if index + 1 < len(starts) else len(table)
        block = table[offset:end]
        raw = _extract_pairs(block)
        # A compact row can contain code + free description without labels.
        if not raw.get("name"):
            for cell in block[0]:
                text = normalize_text(cell)
                if (
                    text
                    and not find_material_codes(
                        text,
                        review_families=material_review_families,
                    )
                    and _field_for_label(text) is None
                ):
                    if re.search(r"不锈钢|金属|铝板|钢板|玻璃|铜|铁", text):
                        raw["name"] = cell
                        break
        location = f"{sheet_name}!rows={start_row + offset}:{start_row + end - 1}"
        specs.append(
            _make_spec(
                raw,
                code=code_match.normalized_code,
                raw_code=code_match.raw_code,
                code_family=code_match.family,
                source_file_id=source_file_id,
                source_location=location,
            )
        )

    # Remove only exact duplicate records.  Separate conflicting source rows are
    # retained so that conflict annotation remains auditable.
    unique: dict[str, MaterialSpec] = {}
    for spec in specs:
        unique.setdefault(spec.id, spec)
    return _annotate_conflicts(list(unique.values()))


def _word_tag(local_name: str) -> str:
    return f"{{{_OOXML_WORD_NS}}}{local_name}"


def _docx_member_is_safe(info: zipfile.ZipInfo) -> bool:
    """Return whether an OOXML member name is safe without extracting it."""

    name = info.filename
    if not name or "\\" in name or name.startswith(("/", "//")):
        return False
    path = PurePosixPath(name)
    if path.is_absolute() or any(part in {"", ".", ".."} for part in path.parts):
        return False
    if re.match(r"^[A-Za-z]:", path.parts[0]):
        return False
    unix_mode = (info.external_attr >> 16) & 0o170000
    return not stat.S_ISLNK(unix_mode)


def _read_docx_document_xml(path: Path) -> tuple[bytes, str]:
    """Read only ``word/document.xml`` after validating the OOXML container."""

    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    document_sha256 = digest.hexdigest()

    with zipfile.ZipFile(path) as archive:
        members: dict[str, zipfile.ZipInfo] = {}
        normalized_names: set[str] = set()
        for info in archive.infolist():
            if not _docx_member_is_safe(info):
                raise ValueError(f"unsafe DOCX member path: {info.filename!r}")
            normalized = info.filename.casefold()
            if normalized in normalized_names:
                raise ValueError(f"duplicate DOCX member path: {info.filename!r}")
            normalized_names.add(normalized)
            members[normalized] = info
        info = members.get(_DOCX_DOCUMENT_MEMBER.casefold())
        if info is None:
            raise ValueError("DOCX has no word/document.xml")
        if info.file_size > _DOCX_MAX_DOCUMENT_XML_BYTES:
            raise ValueError("DOCX word/document.xml exceeds the read-only parser size limit")
        ratio = info.file_size / max(1, info.compress_size)
        if ratio > _DOCX_MAX_XML_COMPRESSION_RATIO:
            raise ValueError("DOCX word/document.xml exceeds the compression-ratio limit")
        payload = archive.read(info)
    if len(payload) != info.file_size:
        raise ValueError("DOCX word/document.xml size changed while reading")
    upper_prefix = payload[:4096].upper()
    if b"<!DOCTYPE" in upper_prefix or b"<!ENTITY" in upper_prefix:
        raise ValueError("DOCX document XML must not contain a DTD or entity declaration")
    return payload, document_sha256


def _docx_paragraph_text(paragraph: ElementTree.Element) -> str:
    pieces: list[str] = []
    for element in paragraph.iter():
        if element.tag == _word_tag("t") and element.text:
            pieces.append(element.text)
        elif element.tag == _word_tag("tab"):
            pieces.append("\t")
        elif element.tag in {_word_tag("br"), _word_tag("cr")}:
            pieces.append("\n")
    return "".join(pieces).strip()


def _docx_cell_text(cell: ElementTree.Element) -> str:
    values = [_docx_paragraph_text(paragraph) for paragraph in cell.findall(f".//{_word_tag('p')}")]
    return " ".join(value for value in values if value).strip()


def _docx_description(value: Any) -> str | None:
    text = normalize_text(value).strip("|｜:：;；,-—–")
    if (
        len(text) < 2
        or _field_for_label(text) is not None
        or find_material_codes(text)
        or _DOCX_NON_DESCRIPTION_RE.fullmatch(text)
    ):
        return None
    return text


def _docx_table_candidate(
    cells: Sequence[str],
    *,
    stainless_families: Iterable[str],
    review_families: Iterable[str],
) -> tuple[MaterialCodeMatch, str, int, tuple[int, ...]] | None:
    codes = find_material_codes_in_cells(
        cells,
        stainless_families=stainless_families,
        review_families=review_families,
    )
    if not codes:
        return None
    code_match = _choose_primary_code(cells, codes)
    code_indexes = next(indexes for candidate, indexes in codes if candidate == code_match)
    ordered_indexes = [
        *range(max(code_indexes) + 1, len(cells)),
        *range(0, min(code_indexes)),
    ]
    for index in ordered_indexes:
        if not normalize_text(cells[index]) or _field_for_label(cells[index]) is not None:
            continue
        description = _docx_description(cells[index])
        return (code_match, description, index, code_indexes) if description is not None else None
    return None


def _docx_paragraph_candidate(
    raw_text: str,
    *,
    stainless_families: Iterable[str],
    review_families: Iterable[str],
) -> tuple[MaterialCodeMatch, str] | None:
    normalized = normalize_text(raw_text)
    codes = find_material_codes(
        normalized,
        stainless_families=stainless_families,
        review_families=review_families,
    )
    if len(codes) != 1:
        return None
    code_match = codes[0]
    span: tuple[int, int] | None = None
    for match in _MATERIAL_CODE_RE.finditer(normalized.upper()):
        family = normalize_material_code_family(match.group("prefix"))
        number_text = match.group("number")
        if family is None:
            continue
        width = max(2, len(number_text))
        if f"{family}-{int(number_text):0{width}d}" == code_match.normalized_code:
            span = match.span("raw")
            break
    if span is None:
        return None
    remainder = f"{normalized[: span[0]]} {normalized[span[1] :]}".strip()
    separated = [
        normalize_text(part) for part in re.split(r"[|｜\t;；]+", remainder) if normalize_text(part)
    ]
    if len(separated) == 1:
        separated = [
            normalize_text(part)
            for part in re.split(r"\s{2,}", raw_text)
            if normalize_text(part)
            and code_match.normalized_code
            not in {
                value.normalized_code
                for value in find_material_codes(
                    part,
                    stainless_families=stainless_families,
                    review_families=review_families,
                )
            }
        ] or separated
    for part in separated:
        usage_split = _DOCX_USAGE_RE.split(part, maxsplit=1)[0]
        description = _docx_description(usage_split) or _docx_description(part)
        if description is not None:
            return code_match, description
    return None


def _dedupe_docx_specs(specs: Sequence[MaterialSpec]) -> list[MaterialSpec]:
    """Merge exact repeated document records while retaining all locators."""

    merged: dict[tuple[Any, ...], MaterialSpec] = {}
    for spec in specs:
        identity = (
            spec.mt_code,
            normalize_text(spec.name).casefold(),
            normalize_text(spec.grade).casefold(),
            spec.thickness_mm,
            normalize_text(spec.finish).casefold(),
            normalize_text(spec.process).casefold(),
            spec.source_file_id,
            spec.source_sha256,
        )
        existing = merged.get(identity)
        if existing is None:
            merged[identity] = spec
            continue
        evidence = sorted(
            {
                *existing.source_evidence,
                *spec.source_evidence,
                *(value for value in (existing.source_location, spec.source_location) if value),
            }
        )
        primary_location = min(evidence) if evidence else existing.source_location
        payload = {
            **existing.model_dump(mode="python", exclude={"id"}),
            "source_location": primary_location,
            "source_evidence": evidence,
        }
        merged[identity] = MaterialSpec(
            id=_stable_id("material", payload),
            **payload,
        )
    return _annotate_conflicts(list(merged.values()))


def load_docx_material_specs(
    path: str | Path,
    *,
    source_file_id: str | None = None,
    expected_sha256: str | None = None,
    stainless_families: Iterable[str] | None = None,
    review_families: Iterable[str] | None = None,
) -> list[MaterialSpec]:
    """Read explicit material-code/description mappings from DOCX OOXML.

    The archive is never extracted. Only the main Word document XML is read, and
    every accepted mapping keeps its document hash plus table-cell or paragraph
    locator. Records remain REVIEW candidates and never become commercial PASS.
    """

    document_path = Path(path).expanduser().resolve()
    payload, document_sha256 = _read_docx_document_xml(document_path)
    if expected_sha256 and document_sha256.casefold() != expected_sha256.casefold():
        raise ValueError("DOCX SHA-256 does not match the ingested source manifest")
    effective_review_families = (
        DEFAULT_REVIEW_CODE_FAMILIES | DEFAULT_AUXILIARY_CODE_FAMILIES
        if review_families is None
        else review_families
    )
    stainless, review = _family_sets(stainless_families, effective_review_families)
    try:
        root = ElementTree.fromstring(payload)
    except ElementTree.ParseError as exc:
        raise ValueError(f"invalid DOCX document XML: {exc}") from exc
    body = root.find(f".//{_word_tag('body')}")
    if body is None:
        raise ValueError("DOCX document XML has no Word body")

    specs: list[MaterialSpec] = []
    table_index = 0
    paragraph_index = 0
    for child in body:
        if child.tag == _word_tag("p"):
            paragraph_index += 1
            raw_text = _docx_paragraph_text(child)
            candidate = _docx_paragraph_candidate(
                raw_text,
                stainless_families=stainless,
                review_families=review,
            )
            if candidate is None:
                continue
            code_match, description = candidate
            locator = f"docx:sha256={document_sha256}:paragraph={paragraph_index}"
            specs.append(
                _make_spec(
                    {"name": description},
                    code=code_match.normalized_code,
                    raw_code=code_match.raw_code,
                    code_family=code_match.family,
                    source_file_id=source_file_id,
                    source_location=locator,
                    source_type="docx_material_book",
                    source_sha256=document_sha256,
                    source_evidence=[locator],
                )
            )
        elif child.tag == _word_tag("tbl"):
            table_index += 1
            for row_index, row in enumerate(child.findall(f"./{_word_tag('tr')}"), 1):
                cells = row.findall(f"./{_word_tag('tc')}")
                values = [_docx_cell_text(cell) for cell in cells]
                candidate = _docx_table_candidate(
                    values,
                    stainless_families=stainless,
                    review_families=review,
                )
                if candidate is None:
                    continue
                code_match, description, description_index, code_indexes = candidate
                row_locator = f"docx:sha256={document_sha256}:table={table_index}:row={row_index}"
                evidence = [
                    f"{row_locator}:cell={cell_index + 1}"
                    for cell_index in sorted({*code_indexes, description_index})
                ]
                specs.append(
                    _make_spec(
                        {"name": description},
                        code=code_match.normalized_code,
                        raw_code=code_match.raw_code,
                        code_family=code_match.family,
                        source_file_id=source_file_id,
                        source_location=row_locator,
                        source_type="docx_material_book",
                        source_sha256=document_sha256,
                        source_evidence=evidence,
                    )
                )
    return _dedupe_docx_specs(specs)


def load_material_specs(
    path: str | Path,
    *,
    source_file_id: str | None = None,
    sheet_names: Iterable[str] | None = None,
) -> list[MaterialSpec]:
    """Read ``.xlsx`` or BIFF ``.xls`` and parse all selected worksheets."""

    workbook_path = Path(path)
    selected = set(sheet_names or [])
    specs: list[MaterialSpec] = []
    suffix = workbook_path.suffix.lower()

    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                if selected and worksheet.title not in selected:
                    continue
                specs.extend(
                    parse_material_rows(
                        worksheet.iter_rows(values_only=True),
                        source_file_id=source_file_id,
                        sheet_name=worksheet.title,
                    )
                )
        finally:
            workbook.close()
    elif suffix == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(workbook_path, on_demand=True)
        try:
            for worksheet in workbook.sheets():
                if selected and worksheet.name not in selected:
                    continue
                rows = (
                    [worksheet.cell_value(row, column) for column in range(worksheet.ncols)]
                    for row in range(worksheet.nrows)
                )
                specs.extend(
                    parse_material_rows(
                        rows,
                        source_file_id=source_file_id,
                        sheet_name=worksheet.name,
                    )
                )
        finally:
            workbook.release_resources()
    else:
        raise ValueError(f"unsupported material workbook: {workbook_path.suffix or '<none>'}")

    return _annotate_conflicts(specs)


__all__ = [
    "DEFAULT_REVIEW_CODE_FAMILIES",
    "DEFAULT_STAINLESS_CODE_FAMILIES",
    "MaterialCodeMatch",
    "annotate_material_conflicts",
    "find_material_codes",
    "find_material_codes_in_cells",
    "find_mt_codes",
    "find_mt_codes_in_cells",
    "load_docx_material_specs",
    "load_material_specs",
    "material_code_disposition",
    "normalize_material_code",
    "normalize_material_code_family",
    "normalize_mt_code",
    "normalize_text",
    "parse_cad_material_specs",
    "parse_material_rows",
]
