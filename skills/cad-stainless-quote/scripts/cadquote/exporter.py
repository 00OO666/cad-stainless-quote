"""Create auditable quotation workbooks with the open-source XlsxWriter backend."""

from __future__ import annotations

import json
import math
import os
import tempfile
from collections.abc import Mapping
from datetime import UTC, date, datetime
from decimal import Decimal
from enum import Enum
from pathlib import Path
from typing import Any

import xlsxwriter
from openpyxl import load_workbook
from xlsxwriter.exceptions import XlsxInputError
from xlsxwriter.image import Image as XlsxWriterImage

from .calculation import calculate_item, engineering_quantity_expression_to_excel
from .io import write_json_atomic
from .models import EvidenceEdge, MeasurementCandidate, ReviewStatus, RunIssue, TakeoffItem

QUOTE_HEADERS = [
    "序号",
    "名称",
    "MT编号",
    "材料",
    "平面图位置",
    "对应立面",
    "对应节点",
    "展开规格",
    "宽",
    "长度",
    "数量",
    "工程量",
    "单位",
    "计价方式",
    "单价",
    "金额",
    "备注",
]

_SHEET_NAMES = ["报价表", "来源追踪", "待确认", "运行信息"]
_EVIDENCE_SHEET_NAME = "截图证据"
_EVIDENCE_HEADERS = [
    "序号",
    "MT",
    "构件ID",
    "阶段",
    "状态",
    "图号",
    "来源文件",
    "CAD bbox",
    "实体ID",
    "DXF Handle",
    "定位图",
    "放大图",
    "缺图原因",
]
_FORMULA_ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
_EVIDENCE_IMAGE_MAX_WIDTH_PX = 480.0
_EVIDENCE_IMAGE_MAX_HEIGHT_PX = 340.0
_EVIDENCE_ROW_MIN_HEIGHT_PX = 240.0


def _enum_value(value: Any) -> Any:
    return value.value if isinstance(value, Enum) else value


def _safe_text(value: str) -> str:
    """Force untrusted CAD/workbook text to remain literal spreadsheet text."""

    if value.startswith(("=", "+", "-", "@", "\t", "\r", "\n")):
        return "'" + value
    return value


def _json_text(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


def _first_record_value(record: Mapping[str, Any], *names: str) -> Any:
    for name in names:
        if name in record and record[name] is not None:
            return record[name]
    return None


def _record_mapping(record: Any) -> dict[str, Any]:
    if isinstance(record, Mapping):
        return dict(record)
    model_dump = getattr(record, "model_dump", None)
    if not callable(model_dump):
        raise TypeError("evidence_records entries must be mappings or expose model_dump()")
    try:
        payload = model_dump(mode="json")
    except TypeError:
        payload = model_dump()
    if not isinstance(payload, Mapping):
        raise TypeError("evidence record model_dump() must return a mapping")
    return dict(payload)


def _joined_record_value(value: Any) -> Any:
    if isinstance(value, (list, tuple, set)):
        return "；".join(str(_enum_value(item)) for item in value)
    return value


def _image_source(value: Any) -> Any:
    if isinstance(value, Mapping):
        return _first_record_value(value, "absolute_path", "path", "file")
    return value


def _normalize_evidence_record(record: Any, sequence: int) -> dict[str, Any]:
    payload = _record_mapping(record)
    images = payload.get("images")
    image_payload = images if isinstance(images, Mapping) else {}
    base_dir = _first_record_value(payload, "image_root", "base_dir", "evidence_root")

    location_image = _first_record_value(
        payload,
        "location_image",
        "location_image_path",
        "context_image",
        "context_image_path",
        "locator_image",
        "overview_image",
        "overview_image_path",
        "panel_image",
        "panel_file",
        "定位图",
    )
    if location_image is None:
        location_image = _first_record_value(
            image_payload,
            "location",
            "locator",
            "overview",
            "panel",
            "定位图",
        )
    zoom_image = _first_record_value(
        payload,
        "zoom_image",
        "zoom_image_path",
        "detail_image",
        "detail_image_path",
        "crop_image",
        "crop_image_path",
        "crop_path",
        "absolute_path",
        "放大图",
    )
    if zoom_image is None:
        zoom_image = _first_record_value(
            image_payload,
            "zoom",
            "detail",
            "crop",
            "放大图",
        )

    record_sequence = _first_record_value(payload, "sequence", "序号", "index")
    return {
        "sequence": sequence if record_sequence is None else record_sequence,
        "mt": _first_record_value(payload, "mt_code", "mt", "MT", "MT编号"),
        "component_id": _first_record_value(payload, "component_id", "构件ID"),
        "stage": _first_record_value(payload, "stage", "阶段"),
        "status": _first_record_value(
            payload,
            "evidence_state",
            "render_state",
            "state",
            "status",
            "状态",
        ),
        "drawing_number": _first_record_value(
            payload,
            "drawing_number",
            "sheet_number",
            "sheet_code",
            "图号",
        ),
        "source_file": _first_record_value(
            payload,
            "source_file",
            "source_file_id",
            "source",
            "source_path",
            "来源文件",
        ),
        "cad_bbox": _first_record_value(
            payload,
            "cad_bbox",
            "focus_bbox",
            "detail_bbox",
            "context_bbox",
            "bbox",
            "CAD bbox",
        ),
        "entity_ids": _joined_record_value(
            _first_record_value(payload, "entity_ids", "entity_id", "实体ID")
        ),
        "dxf_handles": _joined_record_value(
            _first_record_value(
                payload,
                "dxf_handles",
                "entity_handles",
                "dxf_handle",
                "handles",
                "handle",
                "DXF Handle",
            )
        ),
        "location_image": _image_source(location_image),
        "zoom_image": _image_source(zoom_image),
        "missing_reason": _first_record_value(
            payload,
            "missing_reason",
            "missing_image_reason",
            "render_reason",
            "reason",
            "缺图原因",
        ),
        "base_dir": _image_source(base_dir),
    }


def _cell_value(value: Any) -> Any:
    value = _enum_value(value)
    if value is None:
        return None
    if isinstance(value, str):
        return _safe_text(value)
    if isinstance(value, bool):
        return value
    if isinstance(value, Decimal):
        value = float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not math.isfinite(value):
            return _safe_text(str(value))
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (dict, list, tuple, set)):
        return _safe_text(_json_text(value))
    return _safe_text(str(value))


def _write_cell(
    worksheet: Any,
    row: int,
    column: int,
    value: Any,
    cell_format: Any | None = None,
) -> None:
    """Write a typed value without XlsxWriter's string formula/URL inference."""

    value = _cell_value(value)
    if value is None:
        worksheet.write_blank(row, column, None, cell_format)
    elif isinstance(value, bool):
        worksheet.write_boolean(row, column, value, cell_format)
    elif isinstance(value, (int, float)):
        worksheet.write_number(row, column, value, cell_format)
    else:
        worksheet.write_string(row, column, value, cell_format)


def _write_row(
    worksheet: Any,
    row: int,
    values: list[Any],
    cell_format: Any | None = None,
) -> None:
    for column, value in enumerate(values):
        _write_cell(worksheet, row, column, value, cell_format)


def _status_text(item: TakeoffItem) -> str:
    return str(_enum_value(item.status))


def _quantity_and_amount_cache(item: TakeoffItem) -> tuple[float | None, float | None]:
    """Return cached formula results consistent with deterministic takeoff JSON."""

    calculated = calculate_item(item)
    if item.engineering_quantity_expression:
        quantity = (
            None
            if _status_text(item) == "BLOCK" and item.engineering_quantity is None
            else calculated.engineering_quantity
        )
        amount = calculated.amount
    else:
        quantity = (
            item.engineering_quantity
            if item.engineering_quantity is not None
            else calculated.engineering_quantity
        )
        amount = item.amount if item.amount is not None else calculated.amount
    if _status_text(item) != "PASS":
        amount = None
    return quantity, amount


def _effective_export_items(
    items: list[TakeoffItem], edges: list[EvidenceEdge]
) -> list[TakeoffItem]:
    """Recalculate rows and require portable proof for every custom formula."""

    material_evidence_rows: dict[str, set[int]] = {}
    material_evidence_components: dict[str, set[str]] = {}
    for row_index, source in enumerate(items):
        assembly = source.composite_assembly
        if assembly is None:
            continue
        owner = source.component_id or f"<missing-component:row-{row_index + 2}>"
        for reference in assembly.included_materials:
            for evidence_id in reference.evidence_ids:
                normalized_id = evidence_id.strip()
                if not normalized_id:
                    continue
                material_evidence_rows.setdefault(normalized_id, set()).add(row_index)
                material_evidence_components.setdefault(normalized_id, set()).add(owner)

    # Treat every component-to-material edge as an ownership declaration, not
    # only PASS edges.  Otherwise a forged REVIEW/malformed proof edge could
    # hide that the same CAD entity was claimed by another component.
    for edge in edges:
        if edge.relation != "component_to_material":
            continue
        owner = edge.source_id.strip() or "<missing-edge-component>"
        for value in edge.basis:
            if not value.startswith("evidence:"):
                continue
            evidence_id = value.removeprefix("evidence:").strip()
            if evidence_id:
                material_evidence_components.setdefault(evidence_id, set()).add(owner)

    duplicate_material_evidence: dict[int, set[str]] = {}
    for evidence_id, row_indexes in material_evidence_rows.items():
        component_owners = material_evidence_components.get(evidence_id, set())
        if len(row_indexes) <= 1 and len(component_owners) <= 1:
            continue
        for row_index in row_indexes:
            duplicate_material_evidence.setdefault(row_index, set()).add(evidence_id)

    proved_targets: set[tuple[str, str, str, str]] = set()
    proved_materials: dict[tuple[str, str, str, str], set[str]] = {}
    proved_dimensions: set[tuple[str, str]] = set()
    for edge in edges:
        if (
            edge.relation == "component_to_dimension"
            and _enum_value(edge.status) == "PASS"
        ):
            proved_dimensions.add((edge.source_id, edge.target_id))
        if (
            edge.relation == "component_to_engineering_quantity_evidence"
            and _enum_value(edge.status) == "PASS"
        ):
            expressions = [
                value.removeprefix("expression:")
                for value in edge.basis
                if value.startswith("expression:")
            ]
            bases = [
                value.removeprefix("basis:")
                for value in edge.basis
                if value.startswith("basis:")
            ]
            if len(expressions) == 1 and len(bases) == 1:
                proved_targets.add(
                    (edge.source_id, edge.target_id, expressions[0], bases[0])
                )
        if edge.relation == "component_to_material" and _enum_value(edge.status) == "PASS":
            roles = [
                value.removeprefix("role:")
                for value in edge.basis
                if value.startswith("role:")
            ]
            codes = [
                value.removeprefix("material_code:")
                for value in edge.basis
                if value.startswith("material_code:")
            ]
            evidence_ids = {
                value.removeprefix("evidence:")
                for value in edge.basis
                if value.startswith("evidence:")
            }
            if len(roles) == 1 and len(codes) == 1 and evidence_ids:
                key = (edge.source_id, edge.target_id, roles[0], codes[0])
                proved_materials.setdefault(key, set()).update(evidence_ids)

    effective: list[TakeoffItem] = []
    for row_index, source in enumerate(items):
        item = calculate_item(source)
        if item.engineering_quantity_expression and _status_text(item) != "BLOCK":
            proof_identity = (
                item.component_id or "",
                item.engineering_quantity_expression,
                item.engineering_quantity_basis or "",
            )
            missing = sorted(
                evidence_id
                for evidence_id in set(item.engineering_quantity_evidence_ids)
                if (
                    proof_identity[0],
                    evidence_id,
                    proof_identity[1],
                    proof_identity[2],
                )
                not in proved_targets
            )
            if not item.component_id or missing:
                detail = (
                    "缺少构件ID"
                    if not item.component_id
                    else "缺少PASS工程量证据边：" + "，".join(missing)
                )
                item = item.model_copy(
                    update={
                        "status": ReviewStatus.BLOCK,
                        "block_reason": detail,
                        "engineering_quantity": None,
                        "amount": None,
                    }
                )
        assembly = item.composite_assembly
        if assembly is not None:
            duplicate_ids = sorted(duplicate_material_evidence.get(row_index, set()))
            if duplicate_ids:
                detail = "复合材料证据重复占用：" + "，".join(duplicate_ids)
                existing_note = item.note or ""
                item = item.model_copy(
                    update={
                        "status": (
                            ReviewStatus.BLOCK
                            if _status_text(item) == "BLOCK"
                            else ReviewStatus.REVIEW
                        ),
                        "note": "；".join(
                            value for value in (existing_note, detail) if value
                        ),
                        "unit_price": None,
                        "price_entry_id": None,
                        "amount": None,
                    }
                )
            missing_material_proof: list[str] = []
            for label, candidate_id in (
                ("整樘投影宽轴", assembly.projection_width_candidate_id),
                ("整樘投影长轴", assembly.projection_length_candidate_id),
            ):
                if (
                    not item.component_id
                    or not candidate_id
                    or (item.component_id, candidate_id) not in proved_dimensions
                ):
                    missing_material_proof.append(label)
            if (
                not assembly.projection_axis_basis
                or not assembly.projection_axis_evidence_ids
                or not assembly.projection_component_entity_id
                or assembly.projection_component_entity_id
                not in assembly.projection_axis_evidence_ids
            ):
                missing_material_proof.append("整樘投影轴CAD依据")
            if not any(
                reference.role == "glass_infill"
                for reference in assembly.included_materials
            ):
                missing_material_proof.append("glass_infill")
            for reference in assembly.included_materials:
                proof_key = (
                    item.component_id or "",
                    reference.material_spec_id,
                    reference.role,
                    reference.material_code,
                )
                proved_evidence = proved_materials.get(proof_key, set())
                if (
                    _enum_value(reference.status) != "PASS"
                    or not reference.evidence_ids
                    or not set(reference.evidence_ids).issubset(proved_evidence)
                ):
                    missing_material_proof.append(
                        f"{reference.role}:{reference.material_code or reference.material_spec_id}"
                    )
            if not item.component_id or not assembly.included_materials:
                missing_material_proof.append("复合材料构件绑定")
            if missing_material_proof:
                detail = "缺少PASS复合材料证据边：" + "，".join(
                    sorted(set(missing_material_proof))
                )
                existing_note = item.note or ""
                downgraded_status = (
                    ReviewStatus.BLOCK
                    if _status_text(item) == "BLOCK"
                    else ReviewStatus.REVIEW
                )
                item = item.model_copy(
                    update={
                        "status": downgraded_status,
                        "note": "；".join(value for value in (existing_note, detail) if value),
                        "unit_price": None,
                        "price_entry_id": None,
                        "amount": None,
                    }
                )
        effective.append(item)
    return effective


def _quantity_formula(
    item: TakeoffItem,
    *,
    excel_row: int,
    engineering_cache: float | None,
) -> str:
    if item.engineering_quantity_expression and engineering_cache is not None:
        return engineering_quantity_expression_to_excel(
            item.engineering_quantity_expression,
            row=excel_row,
        )
    if item.engineering_quantity_expression:
        return '=""'
    return (
        f'=IF(M{excel_row}="㎡",I{excel_row}*J{excel_row}*K{excel_row}/1000000,'
        f'IF(M{excel_row}="m",J{excel_row}*K{excel_row}/1000,'
        f'IF(OR(M{excel_row}="件",M{excel_row}="套"),K{excel_row},"")))'
    )


def _quote_note(item: TakeoffItem) -> str:
    notes = [f"[{_status_text(item)}]"]
    composite_note = None
    if item.composite_assembly is not None:
        glass_names = [
            f"{reference.material_code}{reference.material_name}"
            for reference in item.composite_assembly.included_materials
            if reference.role == "glass_infill"
        ]
        included_glass = "、".join(glass_names) or "玻璃材料待确认"
        composite_note = (
            f"含{included_glass}；不锈钢框架与玻璃不拆分，整樘按立面投影面积计量"
        )
    reasons = list(
        dict.fromkeys(
            value for value in (composite_note, item.note, item.block_reason) if value
        )
    )
    if reasons:
        reason = "；".join(reasons)
        if len(reason) > 260:
            reason = f"{reason[:260]}…详见“待确认”表"
        notes.append(reason)
    return "；".join(notes)


def _material_display(item: TakeoffItem) -> str | None:
    if item.composite_assembly is None:
        return item.material
    primary = (
        f"{item.mt_code}｜{item.material}"
        if item.material
        else item.mt_code
    )
    included = [
        f"{reference.material_code}｜{reference.material_name}"
        for reference in item.composite_assembly.included_materials
        if reference.material_code or reference.material_name
    ]
    return "；".join([primary, *included])


def _write_quote_sheet(workbook: Any, items: list[TakeoffItem]) -> dict[str, Any]:
    sheet = workbook.add_worksheet("报价表")
    sheet.hide_gridlines(2)
    sheet.freeze_panes(1, 0)

    navy = "#1F4E78"
    pale_blue = "#DCE6F1"
    light_border = "#D9E2F3"
    header = workbook.add_format(
        {
            "bg_color": navy,
            "font_color": "#FFFFFF",
            "bold": True,
            "align": "center",
            "valign": "vcenter",
            "text_wrap": True,
            "border": 1,
            "border_color": navy,
        }
    )
    body = workbook.add_format(
        {
            "valign": "vcenter",
            "text_wrap": True,
            "bottom": 1,
            "bottom_color": light_border,
        }
    )
    numeric = workbook.add_format(
        {
            "valign": "vcenter",
            "align": "right",
            "num_format": "#,##0.00",
            "bottom": 1,
            "bottom_color": light_border,
        }
    )
    centered = workbook.add_format(
        {
            "valign": "vcenter",
            "align": "center",
            "text_wrap": True,
            "bottom": 1,
            "bottom_color": light_border,
        }
    )
    status_formats = {
        "PASS": workbook.add_format(
            {
                "bg_color": "#E2F0D9",
                "align": "center",
                "valign": "vcenter",
                "bottom": 1,
                "bottom_color": light_border,
            }
        ),
        "BLOCK": workbook.add_format(
            {
                "bg_color": "#F4CCCC",
                "align": "center",
                "valign": "vcenter",
                "bottom": 1,
                "bottom_color": light_border,
            }
        ),
        "REVIEW": workbook.add_format(
            {
                "bg_color": "#FFF2CC",
                "align": "center",
                "valign": "vcenter",
                "bottom": 1,
                "bottom_color": light_border,
            }
        ),
    }
    total_label = workbook.add_format(
        {
            "bg_color": pale_blue,
            "bold": True,
            "bottom": 6,
            "bottom_color": navy,
        }
    )
    total_number = workbook.add_format(
        {
            "bg_color": pale_blue,
            "bold": True,
            "num_format": "#,##0.00",
            "bottom": 6,
            "bottom_color": navy,
        }
    )

    _write_row(sheet, 0, QUOTE_HEADERS, header)
    sheet.set_row(0, 30)
    caches: list[tuple[float | None, float | None]] = []
    for index, item in enumerate(items):
        excel_row = index + 2
        row = index + 1
        status = _status_text(item)
        engineering_cache, amount_cache = _quantity_and_amount_cache(item)
        caches.append((engineering_cache, amount_cache))
        values = [
            item.sequence,
            item.name,
            item.mt_code,
            _material_display(item),
            item.plan_location,
            item.elevation,
            item.detail,
            item.unfolded_spec,
            item.width_mm,
            item.length_mm,
            item.quantity,
            None,
            item.unit,
            item.pricing_method,
            item.unit_price,
            None,
            _quote_note(item),
        ]
        for column, value in enumerate(values):
            if column == 0:
                cell_format = status_formats.get(status, status_formats["REVIEW"])
            elif column == 2:
                cell_format = centered
            elif 8 <= column <= 15:
                cell_format = numeric
            else:
                cell_format = body
            _write_cell(sheet, row, column, value, cell_format)

        quantity_formula = _quantity_formula(
            item,
            excel_row=excel_row,
            engineering_cache=engineering_cache,
        )
        amount_formula = (
            f'=IF(LEFT(Q{excel_row},6)<>"[PASS]","",'
            f'IF(OR(L{excel_row}="",O{excel_row}=""),"",'
            f'ROUND(L{excel_row}*O{excel_row},2)))'
        )
        sheet.write_formula(
            row,
            11,
            quantity_formula,
            numeric,
            "" if engineering_cache is None else engineering_cache,
        )
        sheet.write_formula(
            row,
            15,
            amount_formula,
            numeric,
            "" if amount_cache is None else amount_cache,
        )
        if item.engineering_quantity_expression:
            formula_text = item.engineering_quantity_expression
        elif item.unit == "㎡":
            formula_text = (
                f"{item.width_mm if item.width_mm is not None else '?'}×"
                f"{item.length_mm if item.length_mm is not None else '?'}×"
                f"{item.quantity if item.quantity is not None else '?'}÷1,000,000"
            )
        elif item.unit == "m":
            formula_text = (
                f"{item.length_mm if item.length_mm is not None else '?'}×"
                f"{item.quantity if item.quantity is not None else '?'}÷1,000"
            )
        else:
            formula_text = f"{item.quantity if item.quantity is not None else '?'}"
        quantity_evidence = sorted(
            set(item.evidence_ids) | set(item.engineering_quantity_evidence_ids)
        )
        sheet.write_comment(
            row,
            11,
            f"构件ID：{item.component_id or '—'}\n"
            f"计算：{formula_text}\n"
            f"依据：{item.engineering_quantity_basis or '标准单位公式'}\n"
            f"证据：{'，'.join(quantity_evidence) or '—'}",
            {"author": "User", "width": 420, "height": 150},
        )

    if items:
        sheet.add_table(
            0,
            0,
            len(items),
            len(QUOTE_HEADERS) - 1,
            {
                "name": "QuoteItemsTable",
                "style": "Table Style Medium 2",
                "columns": [{"header": value} for value in QUOTE_HEADERS],
            },
        )
    else:
        sheet.autofilter(0, 0, 0, len(QUOTE_HEADERS) - 1)

    total_row = len(items) + 1
    _write_cell(sheet, total_row, 14, "已确认合计", total_label)
    total_cache = sum(
        Decimal(str(amount))
        for item, (_, amount) in zip(items, caches, strict=True)
        if _status_text(item) == "PASS" and amount is not None
    )
    total_formula = f"=SUM(P2:P{len(items) + 1})" if items else "=0"
    sheet.write_formula(total_row, 15, total_formula, total_number, float(total_cache))

    widths = [7, 20, 12, 20, 24, 15, 15, 22, 10, 12, 10, 12, 9, 18, 12, 14, 38]
    for column, width in enumerate(widths):
        sheet.set_column(column, column, width)
    sheet.set_landscape()
    sheet.fit_to_pages(1, 0)
    sheet.repeat_rows(0)
    sheet.set_margins(0.3, 0.3, 0.5, 0.5)

    return {
        "total_excel_row": total_row + 1,
        "engineering_cache": [value[0] for value in caches],
        "amount_cache": [value[1] for value in caches],
        "total_cache": float(total_cache),
    }


def _subheader_format(workbook: Any) -> Any:
    return workbook.add_format(
        {
            "bg_color": "#DCE6F1",
            "font_color": "#1F1F1F",
            "bold": True,
            "valign": "vcenter",
            "text_wrap": True,
            "border": 1,
            "border_color": "#D9E2F3",
        }
    )


def _write_trace_sheet(
    workbook: Any,
    edges: list[EvidenceEdge],
    measurements: list[MeasurementCandidate],
) -> None:
    sheet = workbook.add_worksheet("来源追踪")
    sheet.hide_gridlines(2)
    sheet.freeze_panes(1, 0)
    headers = [
        "关系ID",
        "关系",
        "源ID",
        "目标ID",
        "候选角色",
        "原始值",
        "数值",
        "单位",
        "来源文件ID",
        "图纸ID",
        "实体ID",
        "依据",
        "置信度",
        "状态",
    ]
    _write_row(sheet, 0, headers, _subheader_format(workbook))
    body = workbook.add_format({"valign": "vcenter", "text_wrap": True})
    numeric = workbook.add_format(
        {"valign": "vcenter", "align": "right", "num_format": "#,##0.00"}
    )
    confidence = workbook.add_format(
        {"valign": "vcenter", "align": "right", "num_format": "0.000"}
    )
    measurement_by_id = {value.id: value for value in measurements}
    for row, edge in enumerate(edges, start=1):
        measurement = measurement_by_id.get(edge.target_id)
        engineering_fields = {
            value.split(":", 1)[0]: value.split(":", 1)[1]
            for value in edge.basis
            if edge.relation == "component_to_engineering_quantity_evidence" and ":" in value
        }
        values = [
            edge.id,
            edge.relation,
            edge.source_id,
            edge.target_id,
            (
                measurement.role
                if measurement
                else "engineering_quantity_evidence" if engineering_fields else None
            ),
            (
                measurement.raw_value
                if measurement
                else engineering_fields.get("expression")
            ),
            measurement.numeric_value if measurement else None,
            measurement.unit if measurement else None,
            (
                measurement.source_file_id
                if measurement
                else engineering_fields.get("source_file")
            ),
            measurement.sheet_id if measurement else engineering_fields.get("sheet"),
            (
                "；".join(measurement.entity_ids)
                if measurement
                else edge.target_id if engineering_fields else None
            ),
            "；".join(edge.basis),
            edge.confidence,
            edge.status,
        ]
        for column, value in enumerate(values):
            cell_format = numeric if column == 6 else confidence if column == 12 else body
            _write_cell(sheet, row, column, value, cell_format)
    widths = [30, 24, 30, 30, 14, 24, 12, 10, 28, 28, 38, 48, 12, 12]
    for column, width in enumerate(widths):
        sheet.set_column(column, column, width)
    if edges:
        sheet.autofilter(0, 0, len(edges), len(headers) - 1)


def _write_pending_sheet(
    workbook: Any,
    items: list[TakeoffItem],
    issues: list[RunIssue],
) -> int:
    sheet = workbook.add_worksheet("待确认")
    sheet.hide_gridlines(2)
    sheet.freeze_panes(1, 0)
    headers = ["类型", "序号/来源", "严重度/状态", "原因", "建议动作"]
    _write_row(sheet, 0, headers, _subheader_format(workbook))
    body = workbook.add_format({"valign": "vcenter", "text_wrap": True})
    rows: list[list[Any]] = []
    for item in items:
        if _status_text(item) != "PASS":
            rows.append(
                [
                    "算量项",
                    f"{item.sequence} / {item.component_id or '无构件ID'}",
                    item.status,
                    item.block_reason or item.note,
                    "回图复核并选择证据候选",
                ]
            )
    for issue in issues:
        rows.append(
            [
                "运行问题",
                issue.source_id,
                issue.severity,
                issue.message,
                issue.suggested_action,
            ]
        )
    for row, values in enumerate(rows, start=1):
        _write_row(sheet, row, values, body)
    for column, width in enumerate([14, 26, 14, 62, 42]):
        sheet.set_column(column, column, width)
    if rows:
        sheet.autofilter(0, 0, len(rows), len(headers) - 1)
    return len(rows)


def _write_run_sheet(
    workbook: Any,
    items: list[TakeoffItem],
    metadata: dict[str, Any],
    generated_at: str,
) -> None:
    sheet = workbook.add_worksheet("运行信息")
    sheet.hide_gridlines(2)
    sheet.freeze_panes(1, 0)
    _write_row(sheet, 0, ["键", "值"], _subheader_format(workbook))
    body = workbook.add_format({"valign": "vcenter", "text_wrap": True})
    statuses = [_status_text(item) for item in items]
    rows: list[list[Any]] = [
        ["generated_at", generated_at],
        ["item_count", len(items)],
        ["pass_count", statuses.count("PASS")],
        ["review_count", statuses.count("REVIEW")],
        ["block_count", statuses.count("BLOCK")],
    ]
    for key, value in sorted(metadata.items(), key=lambda pair: str(pair[0])):
        rows.append([str(key), value if isinstance(value, str) else _json_text(value)])
    for row, values in enumerate(rows, start=1):
        _write_row(sheet, row, values, body)
    sheet.set_column(0, 0, 28)
    sheet.set_column(1, 1, 78)


def _resolve_evidence_image(
    value: Any,
    base_dir: Any,
    label: str,
) -> tuple[Path | None, XlsxWriterImage | None, str | None]:
    if value is None or not str(value).strip():
        return None, None, f"缺图：未提供{label}"
    try:
        path = Path(os.fspath(value)).expanduser()
    except TypeError:
        return None, None, f"缺图：{label}路径格式无效"
    if not path.is_absolute() and base_dir is not None and str(base_dir).strip():
        try:
            path = Path(os.fspath(base_dir)).expanduser() / path
        except TypeError:
            return None, None, f"缺图：{label}的基准目录格式无效"
    path = path.resolve()
    if not path.is_file():
        return path, None, f"缺图：{label}文件不存在（{path}）"
    try:
        image = XlsxWriterImage(path)
    except (OSError, TypeError, ValueError, XlsxInputError) as exc:
        return path, None, f"缺图：{label}不可读取（{type(exc).__name__}: {exc}）"
    if image.width <= 0 or image.height <= 0:
        return path, None, f"缺图：{label}尺寸无效（{path}）"
    return path, image, None


def _insert_evidence_image(
    sheet: Any,
    row: int,
    column: int,
    value: Any,
    base_dir: Any,
    label: str,
) -> tuple[bool, float, float | None, str | None]:
    path, image, missing_reason = _resolve_evidence_image(value, base_dir, label)
    if image is None:
        return False, 0.0, None, missing_reason
    scale = min(
        1.0,
        _EVIDENCE_IMAGE_MAX_WIDTH_PX / float(image.width),
        _EVIDENCE_IMAGE_MAX_HEIGHT_PX / float(image.height),
    )
    try:
        result = sheet.insert_image(
            row,
            column,
            image,
            {
                "x_offset": 4,
                "y_offset": 4,
                "x_scale": scale,
                "y_scale": scale,
                "object_position": 1,
                "description": f"{label}：{path.name if path is not None else 'CAD证据'}",
            },
        )
    except (OSError, TypeError, ValueError, XlsxInputError) as exc:
        return (
            False,
            0.0,
            None,
            f"缺图：{label}无法写入工作簿（{type(exc).__name__}: {exc}）",
        )
    if result != 0:
        return False, 0.0, None, f"缺图：{label}无法写入工作簿（错误码 {result}）"
    return True, float(image.height) * scale, scale, None


def _write_evidence_sheet(
    workbook: Any,
    records: list[dict[str, Any]],
) -> dict[str, Any]:
    sheet = workbook.add_worksheet(_EVIDENCE_SHEET_NAME)
    sheet.hide_gridlines(2)
    sheet.freeze_panes(1, 0)
    sheet.set_zoom(80)
    _write_row(sheet, 0, _EVIDENCE_HEADERS, _subheader_format(workbook))
    sheet.set_row(0, 30)

    body = workbook.add_format(
        {
            "valign": "top",
            "text_wrap": True,
            "bottom": 1,
            "bottom_color": "#D9E2F3",
        }
    )
    centered = workbook.add_format(
        {
            "align": "center",
            "valign": "top",
            "text_wrap": True,
            "bottom": 1,
            "bottom_color": "#D9E2F3",
        }
    )
    missing = workbook.add_format(
        {
            "bg_color": "#FCE8E6",
            "font_color": "#C00000",
            "bold": True,
            "valign": "top",
            "text_wrap": True,
            "bottom": 1,
            "bottom_color": "#D9E2F3",
        }
    )

    image_count = 0
    missing_rows: list[int] = []
    missing_cells: list[str] = []
    image_scales: list[float] = []
    for index, record in enumerate(records):
        row = index + 1
        excel_row = row + 1
        values = [
            record["sequence"],
            record["mt"],
            record["component_id"],
            record["stage"],
            record["status"],
            record["drawing_number"],
            record["source_file"],
            record["cad_bbox"],
            record["entity_ids"],
            record["dxf_handles"],
        ]
        for column, value in enumerate(values):
            _write_cell(sheet, row, column, value, centered if column in {0, 1, 4, 5} else body)
        _write_cell(sheet, row, 10, None, body)
        _write_cell(sheet, row, 11, None, body)

        row_image_height = 0.0
        row_missing_reasons: list[str] = []
        for column, key, label in (
            (10, "location_image", "定位图"),
            (11, "zoom_image", "放大图"),
        ):
            inserted, rendered_height, scale, reason = _insert_evidence_image(
                sheet,
                row,
                column,
                record[key],
                record["base_dir"],
                label,
            )
            if inserted:
                image_count += 1
                row_image_height = max(row_image_height, rendered_height)
                if scale is not None:
                    image_scales.append(scale)
            else:
                message = reason or f"缺图：{label}不可用"
                _write_cell(sheet, row, column, message, missing)
                row_missing_reasons.append(message)
                missing_cells.append(f"{'K' if column == 10 else 'L'}{excel_row}")

        if row_missing_reasons:
            explicit_reason = record["missing_reason"]
            reason_parts = []
            if explicit_reason is not None and str(explicit_reason).strip():
                reason_parts.append(str(explicit_reason).strip())
            reason_parts.extend(row_missing_reasons)
            combined_reason = "；".join(reason_parts)
            _write_cell(sheet, row, 12, combined_reason, missing)
            missing_rows.append(excel_row)
        else:
            _write_cell(sheet, row, 12, None, body)
        row_height_px = max(
            _EVIDENCE_ROW_MIN_HEIGHT_PX,
            min(_EVIDENCE_IMAGE_MAX_HEIGHT_PX, row_image_height) + 12.0,
        )
        sheet.set_row(row, row_height_px * 0.75)

    widths = [7, 12, 28, 14, 12, 16, 34, 28, 34, 24, 68, 68, 48]
    for column, width in enumerate(widths):
        sheet.set_column(column, column, width)
    if records:
        sheet.autofilter(0, 0, len(records), len(_EVIDENCE_HEADERS) - 1)
    sheet.set_landscape()
    sheet.fit_to_pages(1, 0)
    sheet.repeat_rows(0)
    sheet.set_margins(0.3, 0.3, 0.5, 0.5)
    return {
        "record_count": len(records),
        "image_count": image_count,
        "missing_rows": missing_rows,
        "missing_cells": missing_cells,
        "image_scales": image_scales,
    }


def _build_xlsx(
    destination: Path,
    items: list[TakeoffItem],
    edges: list[EvidenceEdge],
    measurements: list[MeasurementCandidate],
    issues: list[RunIssue],
    metadata: dict[str, Any],
    generated_at: str,
    evidence_records: list[dict[str, Any]],
) -> dict[str, Any]:
    workbook = xlsxwriter.Workbook(
        str(destination),
        {
            "strings_to_formulas": False,
            "strings_to_urls": False,
            "nan_inf_to_errors": False,
        },
    )
    try:
        quote_state = _write_quote_sheet(workbook, items)
        _write_trace_sheet(workbook, edges, measurements)
        pending_count = _write_pending_sheet(workbook, items, issues)
        _write_run_sheet(workbook, items, metadata, generated_at)
        evidence_state = (
            _write_evidence_sheet(workbook, evidence_records) if evidence_records else None
        )
    finally:
        workbook.close()
    return {
        **quote_state,
        "pending_count": pending_count,
        "evidence": evidence_state,
    }


def _same_cached_value(actual: Any, expected: float | None) -> bool:
    if expected is None:
        return actual in (None, "")
    if not isinstance(actual, (int, float)):
        return False
    return math.isclose(float(actual), float(expected), rel_tol=1e-9, abs_tol=1e-9)


def _has_red_font(cell: Any) -> bool:
    color = cell.font.color
    return bool(
        color is not None
        and color.type == "rgb"
        and str(color.rgb).upper().endswith("C00000")
    )


def _validate_xlsx(path: Path, items: list[TakeoffItem], state: dict[str, Any]) -> dict[str, Any]:
    evidence_state = state.get("evidence")
    expected_sheets = [*_SHEET_NAMES, _EVIDENCE_SHEET_NAME] if evidence_state else _SHEET_NAMES
    formula_book = load_workbook(path, read_only=False, data_only=False)
    try:
        actual_sheets = formula_book.sheetnames
        quote = formula_book["报价表"] if "报价表" in actual_sheets else None
        actual_headers = (
            [cell.value for cell in next(quote.iter_rows(min_row=1, max_row=1))]
            if quote
            else []
        )
        quantity_formulas = (
            [quote.cell(row=index + 2, column=12).value for index in range(len(items))]
            if quote
            else []
        )
        amount_formulas = (
            [quote.cell(row=index + 2, column=16).value for index in range(len(items))]
            if quote
            else []
        )
        total_formula = (
            quote.cell(row=state["total_excel_row"], column=16).value if quote else None
        )
        key_rows = []
        if quote:
            for cells in quote.iter_rows(
                min_row=1,
                max_row=min(state["total_excel_row"], 22),
                min_col=1,
                max_col=len(QUOTE_HEADERS),
            ):
                key_rows.append([cell.value for cell in cells])

        evidence_headers: list[Any] = []
        evidence_image_count = 0
        evidence_missing_rows: list[int] = []
        evidence_missing_cells: list[str] = []
        evidence_missing_red = True
        if evidence_state and _EVIDENCE_SHEET_NAME in actual_sheets:
            evidence_sheet = formula_book[_EVIDENCE_SHEET_NAME]
            evidence_headers = [cell.value for cell in evidence_sheet[1]]
            evidence_image_count = len(getattr(evidence_sheet, "_images", ()))
            for excel_row in range(2, evidence_state["record_count"] + 2):
                if evidence_sheet.cell(row=excel_row, column=13).value not in (None, ""):
                    evidence_missing_rows.append(excel_row)
                for column_letter in ("K", "L"):
                    cell = evidence_sheet[f"{column_letter}{excel_row}"]
                    if isinstance(cell.value, str) and cell.value.startswith("缺图："):
                        evidence_missing_cells.append(cell.coordinate)
            red_coordinates = [
                *evidence_state["missing_cells"],
                *(f"M{row}" for row in evidence_state["missing_rows"]),
            ]
            evidence_missing_red = all(
                formula_book[_EVIDENCE_SHEET_NAME][coordinate].value not in (None, "")
                and _has_red_font(formula_book[_EVIDENCE_SHEET_NAME][coordinate])
                for coordinate in red_coordinates
            )
    finally:
        formula_book.close()

    value_book = load_workbook(path, read_only=True, data_only=True)
    formula_errors: list[dict[str, Any]] = []
    try:
        quote_values = (
            value_book["报价表"] if "报价表" in value_book.sheetnames else None
        )
        quantity_values = (
            [
                quote_values.cell(row=index + 2, column=12).value
                for index in range(len(items))
            ]
            if quote_values
            else []
        )
        amount_values = (
            [
                quote_values.cell(row=index + 2, column=16).value
                for index in range(len(items))
            ]
            if quote_values
            else []
        )
        total_value = (
            quote_values.cell(row=state["total_excel_row"], column=16).value
            if quote_values
            else None
        )
        for sheet in value_book.worksheets:
            for cells in sheet.iter_rows():
                for cell in cells:
                    if cell.value in _FORMULA_ERROR_VALUES:
                        formula_errors.append(
                            {"sheet": sheet.title, "cell": cell.coordinate, "value": cell.value}
                        )
    finally:
        value_book.close()

    quantity_matches = all(
        _same_cached_value(actual, expected)
        for actual, expected in zip(
            quantity_values, state["engineering_cache"], strict=True
        )
    )
    amount_matches = all(
        _same_cached_value(actual, expected)
        for actual, expected in zip(amount_values, state["amount_cache"], strict=True)
    )
    checks = {
        "sheet_order": actual_sheets == expected_sheets,
        "quote_headers": actual_headers == QUOTE_HEADERS,
        "quantity_formulas": quantity_formulas
        == [
            _quantity_formula(
                item,
                excel_row=index + 2,
                engineering_cache=state["engineering_cache"][index],
            )
            for index, item in enumerate(items)
        ],
        "amount_formulas": all(
            isinstance(value, str) and value.startswith("=IF(") for value in amount_formulas
        ),
        "total_formula": isinstance(total_formula, str) and total_formula.startswith("="),
        "quantity_cache": quantity_matches,
        "amount_cache": amount_matches,
        "total_cache": _same_cached_value(total_value, state["total_cache"]),
        "formula_errors": not formula_errors,
    }
    if evidence_state:
        checks.update(
            {
                "evidence_headers": evidence_headers == _EVIDENCE_HEADERS,
                "evidence_image_count": (
                    evidence_image_count == evidence_state["image_count"]
                ),
                "evidence_missing_rows": (
                    evidence_missing_rows == evidence_state["missing_rows"]
                    and evidence_missing_cells == evidence_state["missing_cells"]
                ),
                "evidence_missing_red": evidence_missing_red,
                "evidence_image_scale": all(
                    0 < value <= 1 for value in evidence_state["image_scales"]
                ),
            }
        )
    report = {
        "backend": "XlsxWriter",
        "openpyxl_readable": True,
        "checks": checks,
        "sheet_names": actual_sheets,
        "quote_headers": actual_headers,
        "item_count": len(items),
        "formula_count": len(quantity_formulas) + len(amount_formulas) + 1,
        "formula_errors": formula_errors,
        "key_range": key_rows,
    }
    if evidence_state:
        report["evidence"] = {
            "record_count": evidence_state["record_count"],
            "image_count": evidence_image_count,
            "missing_rows": evidence_missing_rows,
        }
    failed = [name for name, passed in checks.items() if not passed]
    if failed:
        raise RuntimeError(f"XlsxWriter workbook validation failed: {', '.join(failed)}")
    return report


def _record_preview_limitation(preview_dir: str | Path) -> dict[str, Any]:
    directory = Path(preview_dir).expanduser().resolve()
    directory.mkdir(parents=True, exist_ok=True)
    note = directory / "PREVIEW_NOT_GENERATED.txt"
    note.write_text(
        "未生成工作簿图片预览：开源 XlsxWriter/openpyxl 导出后端不包含渲染器。\n"
        "报价工作簿本身已通过 openpyxl 结构、公式和缓存值校验。\n",
        encoding="utf-8",
    )
    return {
        "requested": True,
        "generated": False,
        "directory": str(directory),
        "reason": "XlsxWriter/openpyxl backend does not include an image renderer",
        "note": str(note),
    }


def build_quote_workbook(
    items: list[TakeoffItem],
    output: str | Path,
    *,
    edges: list[EvidenceEdge] | None = None,
    measurements: list[MeasurementCandidate] | None = None,
    issues: list[RunIssue] | None = None,
    metadata: dict[str, Any] | None = None,
    evidence_records: list[Any] | tuple[Any, ...] | None = None,
    preview_dir: str | Path | None = None,
    verification_report: str | Path | None = None,
) -> Path:
    """Create and validate a portable formula-driven quotation workbook."""

    edge_rows = edges or []
    effective_items = _effective_export_items(items, edge_rows)
    destination = Path(output).expanduser().resolve()
    destination.parent.mkdir(parents=True, exist_ok=True)
    generated_at = datetime.now(UTC).isoformat()
    measurement_rows = measurements or []
    issue_rows = issues or []
    run_metadata = metadata or {}
    evidence_rows = [
        _normalize_evidence_record(record, index)
        for index, record in enumerate(evidence_records or (), start=1)
    ]

    with tempfile.TemporaryDirectory(
        prefix=".cadquote-xlsxwriter-",
        dir=destination.parent,
    ) as temp:
        temporary_output = Path(temp) / destination.name
        state = _build_xlsx(
            temporary_output,
            effective_items,
            edge_rows,
            measurement_rows,
            issue_rows,
            run_metadata,
            generated_at,
            evidence_rows,
        )
        report = _validate_xlsx(temporary_output, effective_items, state)
        if not temporary_output.is_file() or temporary_output.stat().st_size == 0:
            raise RuntimeError("XlsxWriter reported success but produced no workbook")
        os.replace(temporary_output, destination)

    preview = (
        _record_preview_limitation(preview_dir)
        if preview_dir is not None
        else {"requested": False, "generated": False}
    )
    report.update(
        {
            "output": str(destination),
            "generated_at": generated_at,
            "edge_count": len(edge_rows),
            "measurement_count": len(measurement_rows),
            "pending_count": state["pending_count"],
            "evidence_record_count": len(evidence_rows),
            "preview": preview,
        }
    )
    if verification_report is not None:
        write_json_atomic(Path(verification_report).expanduser().resolve(), report)
    return destination
