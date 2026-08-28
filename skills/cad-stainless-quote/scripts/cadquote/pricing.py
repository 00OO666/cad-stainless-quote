"""Versioned, exact-by-default price book loading and matching."""

from __future__ import annotations

import json
import re
import unicodedata
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from .models import MaterialSpec, PriceBook, PriceEntry, TakeoffItem


def _norm(value: object | None) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", unicodedata.normalize("NFKC", str(value))).casefold()


def _bool(value: object) -> bool:
    return _norm(value) in {"1", "true", "yes", "y", "是", "已批准", "approved"}


def _code_list(value: object | None) -> list[str]:
    if value is None:
        return []
    raw_values = (
        value
        if isinstance(value, (list, tuple, set))
        else re.split(r"[,，;；|｜]+", str(value))
    )
    result: list[str] = []
    seen: set[str] = set()
    for raw in raw_values:
        code = str(raw).strip()
        identity = _norm(code)
        if code and identity not in seen:
            result.append(code)
            seen.add(identity)
    return result


def load_price_book(path: str | Path) -> PriceBook:
    source = Path(path)
    if source.suffix.lower() == ".json":
        return PriceBook.model_validate(json.loads(source.read_text(encoding="utf-8")))
    if source.suffix.lower() != ".xlsx":
        raise ValueError("price book must be .json or .xlsx")

    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook["价格库"] if "价格库" in workbook.sheetnames else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("empty price book")
    aliases = {
        "id": {"id", "价格id", "priceid"},
        "version": {"版本", "version"},
        "approved": {"批准", "已批准", "approved"},
        "mt_code": {"mt编号", "mtcode", "材料编号"},
        "material": {"材料", "材质", "material"},
        "grade": {"牌号", "grade"},
        "thickness_mm": {"厚度", "厚度mm", "thickness"},
        "finish": {"表面处理", "finish"},
        "process": {"加工方式", "工艺", "process"},
        "pricing_method": {"计价方式", "pricingmethod"},
        "unit": {"单位", "unit"},
        "unit_price": {"单价", "unitprice"},
        "currency": {"币种", "currency"},
        "tax_included": {"含税", "taxincluded"},
        "valid_from": {"生效日期", "validfrom"},
        "valid_to": {"失效日期", "validto"},
        "source": {"价格来源", "来源", "source"},
        "note": {"备注", "note"},
        "included_material_codes": {
            "包含材料编号",
            "随附材料编号",
            "includedmaterialcodes",
        },
        "composite_billing_basis": {
            "复合计量口径",
            "复合计价口径",
            "compositebillingbasis",
        },
    }
    header_map: dict[str, int] = {}
    for index, raw in enumerate(rows[0]):
        normalized = _norm(raw)
        for field, names in aliases.items():
            if normalized in {_norm(name) for name in names}:
                header_map[field] = index
                break
    # These are not optional commercial dimensions: match_price() intentionally
    # refuses wildcard material/process/context fields.  Fail at workbook load
    # time so a structurally incomplete sheet cannot look like a usable price
    # book and then silently miss every item.
    required = {
        "version",
        "approved",
        "mt_code",
        "material",
        "grade",
        "thickness_mm",
        "finish",
        "process",
        "pricing_method",
        "unit",
        "unit_price",
        "currency",
        "tax_included",
        "source",
    }
    missing = required - header_map.keys()
    if missing:
        raise ValueError(f"price book missing columns: {sorted(missing)}")

    def cell(row: tuple[object, ...], field: str, default: object = None) -> object:
        index = header_map.get(field)
        return row[index] if index is not None and index < len(row) else default

    entries: list[PriceEntry] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in row):
            continue
        version = str(cell(row, "version", "unversioned"))
        entries.append(
            PriceEntry(
                id=str(cell(row, "id", f"row:{row_number}")),
                version=version,
                approved=_bool(cell(row, "approved", False)),
                mt_code=str(cell(row, "mt_code")),
                material=cell(row, "material"),
                grade=cell(row, "grade"),
                thickness_mm=cell(row, "thickness_mm"),
                finish=cell(row, "finish"),
                process=cell(row, "process"),
                pricing_method=str(cell(row, "pricing_method")),
                unit=str(cell(row, "unit")),
                unit_price=float(cell(row, "unit_price")),
                currency=str(cell(row, "currency", "CNY")),
                tax_included=(
                    _bool(cell(row, "tax_included"))
                    if cell(row, "tax_included") not in (None, "")
                    else None
                ),
                valid_from=str(cell(row, "valid_from")) if cell(row, "valid_from") else None,
                valid_to=str(cell(row, "valid_to")) if cell(row, "valid_to") else None,
                source=str(cell(row, "source", source)),
                note=cell(row, "note"),
                included_material_codes=_code_list(
                    cell(row, "included_material_codes")
                ),
                composite_billing_basis=(
                    str(cell(row, "composite_billing_basis")).strip()
                    if cell(row, "composite_billing_basis") not in (None, "")
                    else None
                ),
            )
        )
    versions = sorted({entry.version for entry in entries})
    version = versions[0] if len(versions) == 1 else "+".join(versions)
    approved = bool(entries) and all(entry.approved for entry in entries)
    return PriceBook(version=version, approved=approved, source=str(source), entries=entries)


def match_price(
    item: TakeoffItem,
    material: MaterialSpec | None,
    book: PriceBook,
    *,
    quote_date: date | str | None = None,
    currency: str = "CNY",
    tax_included: bool | None = None,
) -> tuple[PriceEntry | None, list[str]]:
    issues: list[str] = []
    if not book.approved:
        return None, ["价格库未整体批准"]
    versions = {entry.version.strip() for entry in book.entries if entry.version.strip()}
    if not book.version.strip() or book.version == "unversioned" or len(versions) != 1:
        return None, ["价格库必须具有唯一、明确的批准版本"]
    if material is None:
        return None, ["缺少唯一材料定义，不能匹配单价"]
    required_material = {
        "材料": material.name,
        "牌号": material.grade,
        "厚度": material.thickness_mm,
        "表面处理": material.finish,
        "加工方式": material.process,
    }
    missing_material = [label for label, value in required_material.items() if value is None]
    if missing_material:
        return None, [f"材料证据缺少：{', '.join(missing_material)}"]
    if material.conflicts:
        return None, ["材料定义存在冲突"]

    assembly = item.composite_assembly
    included_code_set: set[str] = set()
    if assembly is not None:
        if item.unit != "㎡":
            issues.append("复合屏风必须以㎡计量")
        if assembly.billing_basis != "whole_elevation_projection":
            issues.append("复合屏风必须按整樘立面投影计量")
        if (
            not assembly.projection_width_candidate_id
            or not assembly.projection_length_candidate_id
            or not assembly.projection_component_entity_id
            or not assembly.projection_axis_basis
            or not assembly.projection_axis_evidence_ids
        ):
            issues.append("复合屏风缺少整樘投影轴及CAD证据")
        present_roles = {reference.role for reference in assembly.included_materials}
        if "glass_infill" not in present_roles:
            issues.append("screen_with_glass复合构件缺少glass_infill材料")
        missing_roles = sorted(set(assembly.required_material_roles) - present_roles)
        if missing_roles:
            issues.append("复合构件缺少必需材料角色：" + ", ".join(missing_roles))
        for reference in assembly.included_materials:
            if (
                reference.status.value != "PASS"
                or not reference.material_code.strip()
                or not reference.material_name.strip()
                or not reference.evidence_ids
            ):
                issues.append(
                    f"复合材料{reference.role}缺少已确认编号、名称或同构件证据"
                )
            included_code_set.add(_norm(reference.material_code))
        if not assembly.included_materials:
            issues.append("复合构件没有随附材料")
        if issues:
            return None, issues

    if quote_date is None:
        effective_date = date.today()
    elif isinstance(quote_date, date):
        effective_date = quote_date
    else:
        try:
            effective_date = date.fromisoformat(str(quote_date)[:10])
        except ValueError:
            return None, [f"报价基准日期无效：{quote_date}"]

    def parsed_date(raw: str | None, label: str) -> tuple[date | None, str | None]:
        if not raw:
            return None, None
        try:
            return date.fromisoformat(str(raw)[:10]), None
        except ValueError:
            return None, f"{label}无效：{raw}"

    candidates: list[PriceEntry] = []
    rejected_context: list[str] = []
    for entry in book.entries:
        if not entry.approved:
            continue
        if _norm(entry.mt_code) != _norm(item.mt_code):
            continue
        if _norm(entry.pricing_method) != _norm(item.pricing_method):
            continue
        if _norm(entry.unit) != _norm(item.unit):
            continue
        entry_included_codes = {
            _norm(code) for code in entry.included_material_codes if _norm(code)
        }
        if assembly is None:
            if entry.composite_billing_basis is not None or entry_included_codes:
                continue
        elif (
            _norm(entry.composite_billing_basis) != _norm(assembly.billing_basis)
            or entry_included_codes != included_code_set
        ):
            continue
        required_entry = (
            entry.material,
            entry.grade,
            entry.thickness_mm,
            entry.finish,
            entry.process,
        )
        if any(value is None for value in required_entry):
            rejected_context.append(f"{entry.id}: 价格规格字段不完整")
            continue
        text_pairs = (
            (entry.material, material.name),
            (entry.grade, material.grade),
            (entry.finish, material.finish),
            (entry.process, material.process),
        )
        if any(_norm(expected) != _norm(actual) for expected, actual in text_pairs):
            continue
        assert entry.thickness_mm is not None and material.thickness_mm is not None
        if abs(entry.thickness_mm - material.thickness_mm) > 0.01:
            continue
        if _norm(entry.currency) != _norm(currency):
            rejected_context.append(f"{entry.id}: 币种{entry.currency}不匹配{currency}")
            continue
        if entry.tax_included is None:
            rejected_context.append(f"{entry.id}: 未明确含税口径")
            continue
        if tax_included is not None and entry.tax_included != tax_included:
            rejected_context.append(f"{entry.id}: 含税口径不匹配")
            continue
        valid_from, from_error = parsed_date(entry.valid_from, "生效日期")
        valid_to, to_error = parsed_date(entry.valid_to, "失效日期")
        if from_error or to_error:
            rejected_context.extend(value for value in (from_error, to_error) if value)
            continue
        if valid_from and effective_date < valid_from:
            rejected_context.append(f"{entry.id}: 价格尚未生效")
            continue
        if valid_to and effective_date > valid_to:
            rejected_context.append(f"{entry.id}: 价格已过期")
            continue
        candidates.append(entry)
    if not candidates:
        issues.append("没有满足MT、规格、工艺和计价方式的已批准单价")
        issues.extend(sorted(set(rejected_context)))
        return None, issues
    if len(candidates) > 1:
        issues.append(f"存在{len(candidates)}个同等价格匹配，不能自动选价")
        return None, issues
    return candidates[0], issues


def apply_price(
    item: TakeoffItem,
    material: MaterialSpec | None,
    book: PriceBook,
    *,
    quote_date: date | str | None = None,
    currency: str = "CNY",
    tax_included: bool | None = None,
) -> tuple[TakeoffItem, list[str]]:
    entry, issues = match_price(
        item,
        material,
        book,
        quote_date=quote_date,
        currency=currency,
        tax_included=tax_included,
    )
    if entry is None:
        return (
            item.model_copy(
                update={
                    "unit_price": None,
                    "price_entry_id": None,
                    "amount": None,
                }
            ),
            issues,
        )
    return item.model_copy(
        update={"unit_price": entry.unit_price, "price_entry_id": entry.id}
    ), issues
