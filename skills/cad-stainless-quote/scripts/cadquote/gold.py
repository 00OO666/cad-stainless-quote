"""Import and audit human-authored quote workbooks as reviewable gold data.

The importer deliberately treats a workbook as *candidate* gold.  Imported rows
remain REVIEW (or BLOCK when an audit finds a material problem); this module never
turns an unreviewed spreadsheet row into PASS.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter
from datetime import date, datetime, time
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Literal

import xlrd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, ConfigDict, Field

from .calculation import evaluate_numeric_expression, infer_unfolded_width, infer_unit
from .models import ReviewStatus, TakeoffItem

CANONICAL_FIELDS = (
    "sequence",
    "name",
    "mt_code",
    "material_code",
    "material",
    "plan_location",
    "elevation",
    "detail",
    "unfolded_spec",
    "width_mm",
    "length_mm",
    "quantity",
    "engineering_quantity",
    "unit",
    "pricing_method",
    "unit_price",
    "amount",
    "note",
)

_CANONICAL_HEADERS = {
    "序号": "sequence",
    "名称": "name",
    "mt编号": "mt_code",
    "材料": "material",
    "平面图位置": "plan_location",
    "对应立面": "elevation",
    "对应节点": "detail",
    "展开规格": "unfolded_spec",
    "宽": "width_mm",
    "长度": "length_mm",
    "数量": "quantity",
    "工程量": "engineering_quantity",
    "单位": "unit",
    "计价方式": "pricing_method",
    "单价": "unit_price",
    "金额": "amount",
    "备注": "note",
}

_HEADER_ALIASES = {
    **_CANONICAL_HEADERS,
    "编号": "sequence",
    "项次": "sequence",
    "项目名称": "name",
    "项目": "name",
    "材料名称": "name",
    "部位": "name",
    "部位名称": "name",
    "部位/名称": "name",
    "mt号": "mt_code",
    "mt编码": "mt_code",
    "材料代号": "material_code",
    "材料编号": "material_code",
    "材质": "material",
    "材质表面处理": "material",
    "材料特征描述": "material",
    "材料表面": "material",
    "位置": "plan_location",
    "位置图号": "plan_location",
    "使用位置": "plan_location",
    "立面": "elevation",
    "立面编号": "elevation",
    "立面图": "elevation",
    "立面示意图": "elevation",
    "图号": "elevation",
    "节点": "detail",
    "节点编号": "detail",
    "大样": "detail",
    "大样编号": "detail",
    "大样图": "detail",
    "节点示意图": "detail",
    "计算式附图": "unfolded_spec",
    "计算式": "unfolded_spec",
    "展开尺寸": "unfolded_spec",
    "规格尺寸": "unfolded_spec",
    "规格": "unfolded_spec",
    "造型规格mm": "unfolded_spec",
    "展开宽": "width_mm",
    "展开宽度": "width_mm",
    "宽展开尺寸": "width_mm",
    "宽度mm": "width_mm",
    "长": "length_mm",
    "高长": "length_mm",
    "长度mm": "length_mm",
    "件数": "quantity",
    "计量": "quantity",
    "计算方式": "pricing_method",
    "计量单位": "unit",
    "材料单价": "unit_price",
    "含税单价": "unit_price",
    "含税总价": "amount",
    "合计": "amount",
    "小计": "amount",
    "总价": "amount",
    "明细备注": "note",
}

_MT_RE = re.compile(
    r"(?<![A-Z0-9])MT\s*[-_]?\s*(\d{1,3})(?:\s*[-_]\s*(\d{1,3}))?(?![A-Z0-9])",
    re.I,
)
_NUMBER_FORMAT_DECIMALS = re.compile(r"\.([0#?]+)")
_XLRD_CELL_TYPES = {
    xlrd.XL_CELL_EMPTY: "empty",
    xlrd.XL_CELL_TEXT: "text",
    xlrd.XL_CELL_NUMBER: "number",
    xlrd.XL_CELL_DATE: "date",
    xlrd.XL_CELL_BOOLEAN: "boolean",
    xlrd.XL_CELL_ERROR: "error",
    xlrd.XL_CELL_BLANK: "blank",
}


class _StrictModel(BaseModel):
    model_config = ConfigDict(extra="forbid")


class GoldCellEvidence(_StrictModel):
    """Loss-minimised evidence for one source workbook cell."""

    sheet_name: str
    sheet_index: int = Field(ge=0)
    row: int = Field(ge=1)
    column: int = Field(ge=1)
    coordinate: str
    raw_value: Any = None
    value_repr: str
    formula: str | None = None
    formula_kind: Literal["normal", "array"] | None = None
    formula_range: str | None = None
    data_type: str
    number_format: str | None = None


class GoldImageEvidence(_StrictModel):
    """A workbook image or image-cell formula anchored to an auditable cell."""

    id: str
    sheet_name: str
    sheet_index: int = Field(ge=0)
    source_type: Literal["embedded", "cell_formula"]
    category: Literal["elevation", "detail", "drawing", "rendering", "image"]
    anchor_coordinate: str
    anchor_row: int = Field(ge=1)
    anchor_column: int = Field(ge=1)
    end_coordinate: str | None = None
    end_row: int | None = Field(default=None, ge=1)
    end_column: int | None = Field(default=None, ge=1)
    media_format: str | None = None
    formula: str | None = None
    reference_id: str | None = None


class GoldSheetEvidence(_StrictModel):
    sheet_name: str
    sheet_index: int = Field(ge=0)
    max_row: int = Field(ge=0)
    max_column: int = Field(ge=0)
    header_rows: list[int] = Field(default_factory=list)
    header_cells: list[GoldCellEvidence] = Field(default_factory=list)
    field_columns: dict[str, int] = Field(default_factory=dict)
    images: list[GoldImageEvidence] = Field(default_factory=list)
    imported_row_count: int = Field(default=0, ge=0)


class GoldAuditIssue(_StrictModel):
    id: str
    code: Literal[
        "QUANTITY_FORMULA_MISMATCH",
        "QUANTITY_ROUNDING_DEVIATION",
        "QUANTITY_NOT_CALCULABLE",
        "QUANTITY_DERIVED_FROM_ENGINEERING",
        "AMOUNT_FORMULA_MISMATCH",
        "AMOUNT_NOT_CALCULABLE",
        "MISSING_REQUIRED_FIELD",
        "INVALID_NUMERIC_VALUE",
        "UNRECOGNIZED_UNIT",
    ]
    severity: Literal["REVIEW", "BLOCK"]
    message: str
    sheet_name: str
    row: int = Field(ge=1)
    field: str
    actual: str | None = None
    expected: str | None = None
    difference: str | None = None
    tolerance: str | None = None
    evidence_cells: list[str] = Field(default_factory=list)


class GoldRow(_StrictModel):
    id: str
    sheet_name: str
    sheet_index: int = Field(ge=0)
    row: int = Field(ge=1)
    item: TakeoffItem
    raw_cells: list[GoldCellEvidence]
    field_cells: dict[str, list[str]] = Field(default_factory=dict)
    source_material_code: str | None = None
    mt_code_source: Literal["mt_code", "material_code", "material"] | None = None
    image_evidence: list[GoldImageEvidence] = Field(default_factory=list)
    recalculated_engineering_quantity: float | None = None
    reported_engineering_quantity: float | None = None
    reported_quantity: float | None = None
    effective_quantity: float | None = None
    quantity_source: Literal["reported", "derived_from_engineering_quantity", "missing"] = (
        "reported"
    )
    quantity_derivation_basis: str | None = None
    quantity_difference: float | None = None
    quantity_tolerance: float | None = None
    quantity_audit: Literal["MATCH", "MISMATCH", "DERIVED", "NOT_CALCULABLE"]
    audit_issue_ids: list[str] = Field(default_factory=list)


class GoldImportSummary(_StrictModel):
    sheet_count: int = Field(ge=0)
    row_count: int = Field(ge=0)
    mt_distribution: dict[str, int]
    unit_distribution: dict[str, int]
    audit_issue_count: int = Field(ge=0)
    quantity_mismatch_count: int = Field(ge=0)
    quantity_material_mismatch_count: int = Field(ge=0)
    quantity_rounding_deviation_count: int = Field(ge=0)
    quantity_not_calculable_count: int = Field(ge=0)
    quantity_derived_from_engineering_count: int = Field(ge=0)
    amount_mismatch_count: int = Field(ge=0)
    amount_not_calculable_count: int = Field(ge=0)
    pass_count: int = Field(ge=0)
    review_count: int = Field(ge=0)
    block_count: int = Field(ge=0)


class GoldImportResult(_StrictModel):
    """Serializable import result, including row-level source evidence."""

    schema_version: str = "1.1"
    source_path: str
    source_sha256: str
    workbook_format: Literal["xls", "xlsx", "xlsm"]
    sheets: list[GoldSheetEvidence]
    rows: list[GoldRow]
    issues: list[GoldAuditIssue]
    summary: GoldImportSummary

    def to_json(self, *, indent: int | None = 2) -> str:
        return json.dumps(self.model_dump(mode="json"), ensure_ascii=False, indent=indent)

    def write_json(self, output: str | Path, *, indent: int | None = 2) -> Path:
        destination = Path(output)
        destination.parent.mkdir(parents=True, exist_ok=True)
        destination.write_text(self.to_json(indent=indent), encoding="utf-8")
        return destination


class _Cell:
    def __init__(
        self,
        *,
        sheet_name: str,
        sheet_index: int,
        row: int,
        column: int,
        value: Any,
        formula: str | None,
        formula_kind: Literal["normal", "array"] | None,
        formula_range: str | None,
        data_type: str,
        number_format: str | None,
    ) -> None:
        self.sheet_name = sheet_name
        self.sheet_index = sheet_index
        self.row = row
        self.column = column
        self.value = _json_scalar(value)
        self.formula = formula
        self.formula_kind = formula_kind
        self.formula_range = formula_range
        self.data_type = data_type
        self.number_format = number_format

    @property
    def coordinate(self) -> str:
        return f"{get_column_letter(self.column)}{self.row}"

    def evidence(self) -> GoldCellEvidence:
        return GoldCellEvidence(
            sheet_name=self.sheet_name,
            sheet_index=self.sheet_index,
            row=self.row,
            column=self.column,
            coordinate=self.coordinate,
            raw_value=self.value,
            value_repr=_value_repr(self.value),
            formula=self.formula,
            formula_kind=self.formula_kind,
            formula_range=self.formula_range,
            data_type=self.data_type,
            number_format=self.number_format,
        )


class _ImageAnchor:
    def __init__(
        self,
        *,
        source_type: Literal["embedded", "cell_formula"],
        row: int,
        column: int,
        end_row: int | None = None,
        end_column: int | None = None,
        media_format: str | None = None,
        formula: str | None = None,
        reference_id: str | None = None,
    ) -> None:
        self.source_type = source_type
        self.row = row
        self.column = column
        self.end_row = end_row
        self.end_column = end_column
        self.media_format = media_format
        self.formula = formula
        self.reference_id = reference_id


class _Sheet:
    def __init__(
        self,
        name: str,
        index: int,
        rows: list[list[_Cell]],
        images: list[_ImageAnchor] | None = None,
    ) -> None:
        self.name = name
        self.index = index
        self.rows = rows
        self.images = images or []

    @property
    def max_row(self) -> int:
        return len(self.rows)

    @property
    def max_column(self) -> int:
        return max((len(row) for row in self.rows), default=0)

    def cell(self, row: int, column: int) -> _Cell | None:
        if row < 1 or column < 1 or row > len(self.rows):
            return None
        cells = self.rows[row - 1]
        return cells[column - 1] if column <= len(cells) else None


def _json_scalar(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def _value_repr(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _formula_details(
    value: Any, data_type: Any
) -> tuple[str | None, Literal["normal", "array"] | None, str | None]:
    if str(data_type or "") != "f" or value is None:
        return None, None, None
    if isinstance(value, str):
        return value, "normal", None
    text = getattr(value, "text", None)
    reference = getattr(value, "ref", None)
    return (
        str(text) if text is not None else str(value),
        "array",
        str(reference) if reference is not None else None,
    )


def _image_anchor(image: Any) -> tuple[int, int, int | None, int | None] | None:
    anchor = getattr(image, "anchor", None)
    if isinstance(anchor, str):
        match = re.fullmatch(r"([A-Za-z]+)(\d+)", anchor)
        if not match:
            return None
        column = 0
        for char in match.group(1).upper():
            column = column * 26 + ord(char) - 64
        return int(match.group(2)), column, None, None
    start = getattr(anchor, "_from", None)
    if start is None:
        return None
    end = getattr(anchor, "to", None)
    return (
        int(start.row) + 1,
        int(start.col) + 1,
        int(end.row) + 1 if end is not None else None,
        int(end.col) + 1 if end is not None else None,
    )


_DISPIMG_RE = re.compile(r"(?i)DISPIMG\s*\(\s*[\"']([^\"']+)[\"']")


def _read_xlsx(path: Path) -> list[_Sheet]:
    values_book = load_workbook(path, data_only=True, read_only=False)
    formulas_book = load_workbook(path, data_only=False, read_only=False)
    sheets: list[_Sheet] = []
    for index, value_ws in enumerate(values_book.worksheets):
        formula_ws = formulas_book[value_ws.title]
        images: list[_ImageAnchor] = []
        for image in formula_ws._images:
            anchor = _image_anchor(image)
            if anchor is None:
                continue
            row, column, end_row, end_column = anchor
            images.append(
                _ImageAnchor(
                    source_type="embedded",
                    row=row,
                    column=column,
                    end_row=end_row,
                    end_column=end_column,
                    media_format=(str(image.format) if getattr(image, "format", None) else None),
                )
            )

        value_coordinates = {
            coordinate for coordinate, cell in value_ws._cells.items() if cell.value is not None
        }
        formula_coordinates = {
            coordinate for coordinate, cell in formula_ws._cells.items() if cell.value is not None
        }
        populated = value_coordinates | formula_coordinates
        max_row = max((row for row, _ in populated), default=0)
        max_column = max((column for _, column in populated), default=0)
        for merged in formula_ws.merged_cells.ranges:
            max_row = max(max_row, int(merged.max_row))
            max_column = max(max_column, int(merged.max_col))
        for image in images:
            max_row = max(max_row, image.end_row or image.row)
            max_column = max(max_column, image.end_column or image.column)

        rows: list[list[_Cell]] = []
        for row_index in range(1, max_row + 1):
            row: list[_Cell] = []
            for column in range(1, max_column + 1):
                value_cell = value_ws.cell(row_index, column)
                formula_cell = formula_ws.cell(row_index, column)
                formula, formula_kind, formula_range = _formula_details(
                    formula_cell.value, formula_cell.data_type
                )
                value = value_cell.value
                if formula is not None:
                    match = _DISPIMG_RE.search(formula)
                    if match:
                        images.append(
                            _ImageAnchor(
                                source_type="cell_formula",
                                row=row_index,
                                column=column,
                                formula=formula,
                                reference_id=match.group(1),
                            )
                        )
                row.append(
                    _Cell(
                        sheet_name=value_ws.title,
                        sheet_index=index,
                        row=row_index,
                        column=column,
                        value=value,
                        formula=formula,
                        formula_kind=formula_kind,
                        formula_range=formula_range,
                        data_type=str(formula_cell.data_type or value_cell.data_type or "unknown"),
                        number_format=str(formula_cell.number_format or "General"),
                    )
                )
            rows.append(row)
        sheets.append(_Sheet(value_ws.title, index, rows, images))
    values_book.close()
    formulas_book.close()
    return sheets


def _read_xls(path: Path) -> list[_Sheet]:
    book = xlrd.open_workbook(path, formatting_info=True)
    sheets: list[_Sheet] = []
    for index, worksheet in enumerate(book.sheets()):
        rows: list[list[_Cell]] = []
        for row_index in range(worksheet.nrows):
            row: list[_Cell] = []
            for column in range(worksheet.ncols):
                source = worksheet.cell(row_index, column)
                number_format = None
                try:
                    xf = book.xf_list[source.xf_index]
                    number_format = book.format_map[xf.format_key].format_str
                except (AttributeError, IndexError, KeyError):
                    pass
                row.append(
                    _Cell(
                        sheet_name=worksheet.name,
                        sheet_index=index,
                        row=row_index + 1,
                        column=column + 1,
                        value=source.value if source.ctype != xlrd.XL_CELL_EMPTY else None,
                        formula=None,
                        formula_kind=None,
                        formula_range=None,
                        data_type=_XLRD_CELL_TYPES.get(source.ctype, "unknown"),
                        number_format=number_format,
                    )
                )
            rows.append(row)
        sheets.append(_Sheet(worksheet.name, index, rows))
    return sheets


def _normalise_header(value: Any) -> str:
    text = _value_repr(value).strip().lower()
    return re.sub(r"[\s\n\r\t/／()（）:_：·-]+", "", text)


def _normalise_mt(value: Any) -> str | None:
    text = _value_repr(value).strip().upper().replace("_", "-")
    match = _MT_RE.search(text)
    if not match:
        return None
    suffix = f"-{int(match.group(2)):02d}" if match.group(2) else ""
    return f"MT-{int(match.group(1)):02d}{suffix}"


def _is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _to_decimal(value: Any) -> Decimal | None:
    if _is_empty(value):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    text = str(value).strip().replace(",", "").replace("，", "")
    text = re.sub(r"(?i)\s*(?:mm|m|㎡|m2|件|套)\s*$", "", text)
    try:
        return evaluate_numeric_expression(text)
    except ValueError:
        return None


def _to_float(value: Any) -> float | None:
    number = _to_decimal(value)
    return float(number) if number is not None else None


def _to_sequence(value: Any, fallback: int) -> int:
    number = _to_decimal(value)
    if number is None or number < 1:
        return fallback
    return int(number)


def _normalise_unit(value: Any, pricing_method: Any = None) -> str | None:
    text = _value_repr(value).strip().lower().replace("²", "2")
    aliases = {
        "m": "m",
        "米": "m",
        "延米": "m",
        "lm": "m",
        "㎡": "㎡",
        "m2": "㎡",
        "m^2": "㎡",
        "平方米": "㎡",
        "平米": "㎡",
        "件": "件",
        "套": "套",
    }
    return aliases.get(text) or infer_unit(_value_repr(pricing_method))


def _default_pricing_method(unit: str | None) -> str | None:
    return {"m": "按米", "㎡": "按展开面积", "件": "按件", "套": "按套"}.get(unit)


def _format_quantum(number_format: str | None) -> Decimal:
    """Return half of the visible cell quantum for comparison tolerance."""
    if not number_format or number_format.lower() == "general":
        return Decimal("0.0000005")
    section = number_format.split(";", 1)[0]
    match = _NUMBER_FORMAT_DECIMALS.search(section)
    decimals = len(match.group(1)) if match else 0
    return Decimal(1).scaleb(-decimals) / Decimal(2)


def _formula_quantity(
    *,
    unit: str | None,
    width: Decimal | None,
    length: Decimal | None,
    quantity: Decimal | None,
    unfolded_spec: Any,
) -> tuple[Decimal | None, list[str]]:
    missing: list[str] = []
    if quantity is None:
        missing.append("数量")
    if unit == "m":
        if length is None:
            missing.append("长度")
        if missing:
            return None, missing
        return length * quantity / Decimal("1000"), []  # type: ignore[operator]
    if unit == "㎡":
        if width is None:
            width = infer_unfolded_width(_value_repr(unfolded_spec))
        if width is None:
            missing.append("宽/展开规格")
        if length is None:
            missing.append("长度")
        if missing:
            return None, missing
        return width * length * quantity / Decimal("1000000"), []  # type: ignore[operator]
    if unit in ("件", "套"):
        return (quantity, []) if quantity is not None else (None, missing)
    return None, ["单位"]


def _quantity_from_authoritative_engineering(
    *,
    unit: str | None,
    width: Decimal | None,
    length: Decimal | None,
    reported_engineering: Decimal | None,
    unfolded_spec: Any,
    display_tolerance: Decimal,
) -> tuple[Decimal | None, str | None]:
    """Derive a missing/default quantity from an authoritative human total.

    This is a *candidate-gold repair* rule, not a production CAD inference.
    It applies only when the reported engineering quantity resolves to one
    positive integer multiplier under the standard unit formula. Production
    takeoff must still prove that multiplier from physical CAD topology.

    For linear pricing only the displayed length axis is considered here. A
    workbook that bills the width axis is an audited billing-axis convention,
    not evidence that the quantity field should be changed.
    """

    if reported_engineering is None or reported_engineering <= 0:
        return None, None
    base: Decimal | None = None
    basis: str | None = None
    if unit == "m" and length is not None and length > 0:
        base = length / Decimal("1000")
        basis = "工程量÷显示长度（m）得到有效数量"
    elif unit == "㎡":
        if width is None:
            inferred = infer_unfolded_width(_value_repr(unfolded_spec))
            width = Decimal(str(inferred)) if inferred is not None else None
        if width is not None and width > 0 and length is not None and length > 0:
            base = width * length / Decimal("1000000")
            basis = "工程量÷单件展开面积得到有效数量"
    elif unit in {"件", "套"}:
        base = Decimal(1)
        basis = "工程量即有效件数/套数"
    if base is None or base <= 0:
        return None, None

    ratio = reported_engineering / base
    integer = ratio.to_integral_value(rounding=ROUND_HALF_UP)
    if integer <= 0 or integer > Decimal("100000"):
        return None, None
    reconstructed = base * integer
    tolerance = max(display_tolerance, abs(reported_engineering) * Decimal("0.0001"))
    if abs(reconstructed - reported_engineering) > tolerance:
        return None, None
    return integer, basis


def _header_field(header: str) -> str | None:
    field = _HEADER_ALIASES.get(header)
    if field:
        return field
    if header.startswith(("单价", "材料单价", "含税单价")):
        return "unit_price"
    if header.startswith(("金额", "合计", "小计", "总价", "含税总价")):
        return "amount"
    if header.startswith("宽度") and header.endswith(("mm", "毫米")):
        return "width_mm"
    if header.startswith("长度") and header.endswith(("mm", "毫米")):
        return "length_mm"
    if header.startswith("造型规格"):
        return "unfolded_spec"
    return None


def _mapping_for_header_rows(sheet: _Sheet, start: int, depth: int) -> dict[str, int]:
    observations: list[tuple[int, str, str]] = []
    for column in range(1, sheet.max_column + 1):
        for row_number in range(start, start + depth):
            cell = sheet.cell(row_number, column)
            header = _normalise_header(cell.value if cell else None)
            if not header:
                continue
            field = "mt_code" if header == "颜色" else _header_field(header)
            if field:
                observations.append((column, header, field))

    mapping: dict[str, int] = {}
    count_specific = [
        observation for observation in observations if observation[1] in {"件数", "计量"}
    ]
    quantity_named = [observation for observation in observations if observation[1] == "数量"]
    engineering_named = [
        observation for observation in observations if observation[2] == "engineering_quantity"
    ]

    for column, _header, field in observations:
        if field in {"quantity", "engineering_quantity"}:
            continue
        mapping.setdefault(field, column)

    if engineering_named:
        mapping["engineering_quantity"] = engineering_named[0][0]
        candidates = count_specific or quantity_named
        if candidates:
            mapping["quantity"] = candidates[0][0]
    elif count_specific and quantity_named:
        # Several established takeoff templates label the physical count as
        # "计量" below a dimensions group, while the calculated billable value
        # is headed "数量".  Keep both semantics rather than overwriting one.
        mapping["quantity"] = count_specific[0][0]
        mapping["engineering_quantity"] = quantity_named[0][0]
    elif count_specific:
        mapping["quantity"] = count_specific[0][0]
    elif quantity_named:
        mapping["quantity"] = quantity_named[0][0]

    return mapping


def _header_mapping(sheet: _Sheet) -> tuple[list[int], dict[str, int]] | None:
    best: tuple[int, int, int, dict[str, int]] | None = None
    for row_number in range(1, min(sheet.max_row, 30) + 1):
        for depth in (1, 2):
            if row_number + depth - 1 > sheet.max_row:
                continue
            mapping = _mapping_for_header_rows(sheet, row_number, depth)
            identity_score = sum(
                field in mapping
                for field in ("sequence", "name", "mt_code", "material_code", "material")
            )
            measurement_score = sum(
                field in mapping
                for field in (
                    "unfolded_spec",
                    "width_mm",
                    "length_mm",
                    "quantity",
                    "engineering_quantity",
                    "unit",
                )
            )
            score = len(mapping) * 2 + identity_score * 3 + measurement_score
            # Prefer the shallower header when both interpretations recognise
            # exactly the same fields.  A real second header row must add value.
            candidate = (score, -depth, -row_number, mapping)
            if best is None or candidate[:3] > best[:3]:
                best = candidate
    if best is None or best[0] < 12:
        return None
    score, negative_depth, negative_row, mapping = best
    del score
    row_number = -negative_row
    depth = -negative_depth
    return list(range(row_number, row_number + depth)), mapping


def _looks_like_data(values: dict[str, Any]) -> bool:
    sequence = _to_decimal(values.get("sequence"))
    mt_code = _normalise_mt(values.get("mt_code"))
    name = _value_repr(values.get("name")).strip()
    summary_text = " ".join(
        _value_repr(values.get(field)).strip()
        for field in ("sequence", "name")
        if not _is_empty(values.get(field))
    )
    if re.search(r"(?:^|\s)(?:合计|小计|总计|总合计|本页合计|税金)(?:$|\s)", summary_text):
        return False
    business_signal = any(
        not _is_empty(values.get(field))
        for field in (
            "name",
            "mt_code",
            "material_code",
            "material",
            "unfolded_spec",
            "width_mm",
            "length_mm",
            "quantity",
            "engineering_quantity",
        )
    )
    if sequence is not None and sequence >= 1 and business_signal:
        return True
    if mt_code and _MT_RE.search(mt_code) and name:
        return True
    return False


def _issue_id(source_sha256: str, sheet: str, row: int, code: str) -> str:
    payload = f"{source_sha256}|{sheet}|{row}|{code}".encode()
    return f"gold-issue:{hashlib.sha256(payload).hexdigest()[:20]}"


def _row_id(source_sha256: str, sheet: str, row: int) -> str:
    payload = f"{source_sha256}|{sheet}|{row}".encode()
    return f"gold-row:{hashlib.sha256(payload).hexdigest()[:20]}"


def _field_value(sheet: _Sheet, row: int, mapping: dict[str, int], field: str) -> Any:
    column = mapping.get(field)
    cell = sheet.cell(row, column) if column is not None else None
    return cell.value if cell else None


_CELL_REFERENCE_RE = re.compile(r"(?<![A-Z0-9_])\$?([A-Z]{1,3})\$?(\d+)", re.I)


def _canonical_amount_formula(sheet: _Sheet, row: int, mapping: dict[str, int]) -> bool:
    amount_column = mapping.get("amount")
    engineering_column = mapping.get("engineering_quantity")
    unit_price_column = mapping.get("unit_price")
    if not amount_column or not engineering_column or not unit_price_column:
        return False
    amount_cell = sheet.cell(row, amount_column)
    if amount_cell is None or amount_cell.formula is None:
        return True
    references = {
        f"{column.upper()}{reference_row}"
        for column, reference_row in _CELL_REFERENCE_RE.findall(amount_cell.formula)
    }
    expected = {
        f"{get_column_letter(engineering_column)}{row}",
        f"{get_column_letter(unit_price_column)}{row}",
    }
    return references == expected


def _image_category(
    sheet: _Sheet, header_rows: list[int], column: int
) -> Literal["elevation", "detail", "drawing", "rendering", "image"]:
    text = "".join(
        _normalise_header(
            sheet.cell(row_number, column).value
            if sheet.cell(row_number, column) is not None
            else None
        )
        for row_number in header_rows
    )
    if any(token in text for token in ("节点", "大样")):
        return "detail"
    if "立面" in text:
        return "elevation"
    if any(token in text for token in ("效果", "图样效果")):
        return "rendering"
    if any(token in text for token in ("图纸", "图样", "附图")):
        return "drawing"
    return "image"


def _sheet_image_evidence(
    source_sha256: str, sheet: _Sheet, header_rows: list[int]
) -> list[GoldImageEvidence]:
    evidence: list[GoldImageEvidence] = []
    for image_index, image in enumerate(sheet.images, start=1):
        anchor_coordinate = f"{get_column_letter(image.column)}{image.row}"
        end_coordinate = (
            f"{get_column_letter(image.end_column)}{image.end_row}"
            if image.end_row is not None and image.end_column is not None
            else None
        )
        payload = (
            f"{source_sha256}|{sheet.name}|{image.source_type}|{image_index}|"
            f"{anchor_coordinate}|{end_coordinate or ''}|{image.reference_id or ''}"
        ).encode()
        evidence.append(
            GoldImageEvidence(
                id=f"gold-image:{hashlib.sha256(payload).hexdigest()[:20]}",
                sheet_name=sheet.name,
                sheet_index=sheet.index,
                source_type=image.source_type,
                category=_image_category(sheet, header_rows, image.column),
                anchor_coordinate=anchor_coordinate,
                anchor_row=image.row,
                anchor_column=image.column,
                end_coordinate=end_coordinate,
                end_row=image.end_row,
                end_column=image.end_column,
                media_format=image.media_format,
                formula=image.formula,
                reference_id=image.reference_id,
            )
        )
    return evidence


def import_gold_workbook(path: str | Path) -> GoldImportResult:
    """Import an ``.xls``/``.xlsx`` quote and audit every detected data row."""
    source = Path(path).expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(source)
    suffix = source.suffix.lower()
    if suffix == ".xls":
        workbook_format: Literal["xls", "xlsx", "xlsm"] = "xls"
        source_sheets = _read_xls(source)
    elif suffix in {".xlsx", ".xlsm"}:
        workbook_format = "xlsm" if suffix == ".xlsm" else "xlsx"
        source_sheets = _read_xlsx(source)
    else:
        raise ValueError(f"unsupported gold workbook format: {source.suffix}")

    source_hash = _sha256(source)
    sheets: list[GoldSheetEvidence] = []
    rows: list[GoldRow] = []
    issues: list[GoldAuditIssue] = []
    fallback_sequence = 1

    for sheet in source_sheets:
        detected = _header_mapping(sheet)
        if detected is None:
            sheets.append(
                GoldSheetEvidence(
                    sheet_name=sheet.name,
                    sheet_index=sheet.index,
                    max_row=sheet.max_row,
                    max_column=sheet.max_column,
                    images=_sheet_image_evidence(source_hash, sheet, []),
                )
            )
            continue
        header_rows, mapping = detected
        sheet_images = _sheet_image_evidence(source_hash, sheet, header_rows)
        header_cells = [
            cell.evidence() for row_number in header_rows for cell in sheet.rows[row_number - 1]
        ]
        imported_before = len(rows)
        for row_number in range(header_rows[-1] + 1, sheet.max_row + 1):
            values = {
                field: _field_value(sheet, row_number, mapping, field) for field in CANONICAL_FIELDS
            }
            if not _looks_like_data(values):
                continue

            raw_cells = [cell.evidence() for cell in sheet.rows[row_number - 1]]
            field_cells = {
                field: [sheet.cell(row_number, column).coordinate]  # type: ignore[union-attr]
                for field, column in mapping.items()
                if sheet.cell(row_number, column) is not None
            }
            source_material_code = _value_repr(values["material_code"]).strip() or None
            material = _value_repr(values["material"]).strip() or None
            mt_code = _normalise_mt(values["mt_code"])
            mt_code_source: Literal["mt_code", "material_code", "material"] | None = None
            if mt_code:
                mt_code_source = "mt_code"
            elif (derived_mt := _normalise_mt(source_material_code)) is not None:
                mt_code = derived_mt
                mt_code_source = "material_code"
            elif (derived_mt := _normalise_mt(material)) is not None:
                mt_code = derived_mt
                mt_code_source = "material"
            if mt_code_source and mt_code_source != "mt_code":
                field_cells["mt_code"] = list(field_cells.get(mt_code_source, []))
            mt_code = mt_code or ""
            name = _value_repr(values["name"]).strip()
            sequence = _to_sequence(values["sequence"], fallback_sequence)
            fallback_sequence = max(fallback_sequence + 1, sequence + 1)
            unit = _normalise_unit(values["unit"], values["pricing_method"])
            pricing_method = _value_repr(
                values["pricing_method"]
            ).strip() or _default_pricing_method(unit)
            width = _to_decimal(values["width_mm"])
            length = _to_decimal(values["length_mm"])
            quantity = _to_decimal(values["quantity"])
            reported = _to_decimal(values["engineering_quantity"])
            unit_price = _to_decimal(values["unit_price"])
            amount = _to_decimal(values["amount"])
            numeric_values = {
                "width_mm": width,
                "length_mm": length,
                "quantity": quantity,
                "engineering_quantity": reported,
                "unit_price": unit_price,
                "amount": amount,
            }
            invalid_numeric_fields = [
                field for field, value in numeric_values.items() if value is not None and value < 0
            ]
            width_for_calculation = width if width is None or width >= 0 else None
            length_for_calculation = length if length is None or length >= 0 else None
            quantity_for_calculation = quantity if quantity is None or quantity >= 0 else None
            reported_for_item = reported if reported is None or reported >= 0 else None
            unit_price_for_item = unit_price if unit_price is None or unit_price >= 0 else None
            amount_for_item = amount if amount is None or amount >= 0 else None
            engineering_cell = sheet.cell(row_number, mapping.get("engineering_quantity", 0))
            tolerance = _format_quantum(
                engineering_cell.number_format if engineering_cell else None
            )
            derived_quantity: Decimal | None = None
            quantity_derivation_basis: str | None = None
            if quantity is None or quantity >= 0:
                derived_quantity, quantity_derivation_basis = (
                    _quantity_from_authoritative_engineering(
                        unit=unit,
                        width=width_for_calculation,
                        length=length_for_calculation,
                        reported_engineering=reported_for_item,
                        unfolded_spec=values["unfolded_spec"],
                        display_tolerance=tolerance,
                    )
                )
            quantity_was_derived = (
                derived_quantity is not None and derived_quantity != quantity_for_calculation
            )
            if quantity_was_derived:
                quantity_for_calculation = derived_quantity
            else:
                quantity_derivation_basis = None
            recalculated, missing = _formula_quantity(
                unit=unit,
                width=width_for_calculation,
                length=length_for_calculation,
                quantity=quantity_for_calculation,
                unfolded_spec=values["unfolded_spec"],
            )
            issue_ids: list[str] = []
            row_status = ReviewStatus.REVIEW
            quantity_audit: Literal[
                "MATCH", "MISMATCH", "DERIVED", "NOT_CALCULABLE"
            ] = "MATCH"
            difference: Decimal | None = None
            quantity_source: Literal[
                "reported", "derived_from_engineering_quantity", "missing"
            ] = "missing" if quantity is None else "reported"
            if quantity_was_derived:
                quantity_source = "derived_from_engineering_quantity"

            if invalid_numeric_fields:
                code = "INVALID_NUMERIC_VALUE"
                issue = GoldAuditIssue(
                    id=_issue_id(source_hash, sheet.name, row_number, code),
                    code=code,
                    severity="BLOCK",
                    message="数值字段不能为负数：" + "、".join(invalid_numeric_fields),
                    sheet_name=sheet.name,
                    row=row_number,
                    field="/".join(invalid_numeric_fields),
                    evidence_cells=sum(
                        (field_cells.get(field, []) for field in invalid_numeric_fields), []
                    ),
                )
                issues.append(issue)
                issue_ids.append(issue.id)
                row_status = ReviewStatus.BLOCK

            if not _is_empty(values["unit"]) and unit is None:
                code = "UNRECOGNIZED_UNIT"
                issue = GoldAuditIssue(
                    id=_issue_id(source_hash, sheet.name, row_number, code),
                    code=code,
                    severity="BLOCK",
                    message=f"无法识别计价单位：{_value_repr(values['unit']).strip()}",
                    sheet_name=sheet.name,
                    row=row_number,
                    field="unit",
                    actual=_value_repr(values["unit"]).strip(),
                    evidence_cells=field_cells.get("unit", []),
                )
                issues.append(issue)
                issue_ids.append(issue.id)
                row_status = ReviewStatus.BLOCK

            if quantity_was_derived:
                quantity_audit = "DERIVED"
                code = "QUANTITY_DERIVED_FROM_ENGINEERING"
                issue = GoldAuditIssue(
                    id=_issue_id(source_hash, sheet.name, row_number, code),
                    code=code,
                    severity="REVIEW",
                    message=(
                        "按已确认业务口径以人工工程量为准，反推出数量字段存在漏写或默认值"
                    ),
                    sheet_name=sheet.name,
                    row=row_number,
                    field="quantity",
                    actual=str(quantity) if quantity is not None else None,
                    expected=str(derived_quantity),
                    evidence_cells=sum(
                        (
                            field_cells.get(field, [])
                            for field in (
                                "width_mm",
                                "length_mm",
                                "quantity",
                                "engineering_quantity",
                                "unit",
                            )
                        ),
                        [],
                    ),
                )
                issues.append(issue)
                issue_ids.append(issue.id)

            missing_required = [
                label for label, value in (("名称", name), ("MT编号", mt_code)) if not value
            ]
            if missing_required:
                code = "MISSING_REQUIRED_FIELD"
                issue = GoldAuditIssue(
                    id=_issue_id(source_hash, sheet.name, row_number, code),
                    code=code,
                    severity="BLOCK",
                    message=f"缺少必填字段：{'、'.join(missing_required)}",
                    sheet_name=sheet.name,
                    row=row_number,
                    field="/".join(missing_required),
                    evidence_cells=sum(
                        (
                            field_cells.get(field, [])
                            for field in ("name", "mt_code", "material_code", "material")
                        ),
                        [],
                    ),
                )
                issues.append(issue)
                issue_ids.append(issue.id)
                row_status = ReviewStatus.BLOCK

            if recalculated is None:
                quantity_audit = "NOT_CALCULABLE"
                code = "QUANTITY_NOT_CALCULABLE"
                issue = GoldAuditIssue(
                    id=_issue_id(source_hash, sheet.name, row_number, code),
                    code=code,
                    severity="BLOCK",
                    message=f"无法复算工程量，缺少或不支持：{'、'.join(missing)}",
                    sheet_name=sheet.name,
                    row=row_number,
                    field="engineering_quantity",
                    actual=str(reported) if reported is not None else None,
                    evidence_cells=sum(
                        (
                            field_cells.get(field, [])
                            for field in ("width_mm", "length_mm", "quantity", "unit")
                        ),
                        [],
                    ),
                )
                issues.append(issue)
                issue_ids.append(issue.id)
                row_status = ReviewStatus.BLOCK
            elif reported is None:
                quantity_audit = "NOT_CALCULABLE"
                code = "QUANTITY_NOT_CALCULABLE"
                issue = GoldAuditIssue(
                    id=_issue_id(source_hash, sheet.name, row_number, code),
                    code=code,
                    severity="BLOCK",
                    message="原表缺少工程量，无法与复算值核对",
                    sheet_name=sheet.name,
                    row=row_number,
                    field="engineering_quantity",
                    expected=str(recalculated),
                    evidence_cells=field_cells.get("engineering_quantity", []),
                )
                issues.append(issue)
                issue_ids.append(issue.id)
                row_status = ReviewStatus.BLOCK
            else:
                difference = reported - recalculated
                if abs(difference) > tolerance:
                    quantity_audit = "MISMATCH"
                    material_threshold = max(
                        tolerance * Decimal(2), abs(recalculated) * Decimal("0.01")
                    )
                    is_material_mismatch = abs(difference) > material_threshold
                    code = (
                        "QUANTITY_FORMULA_MISMATCH"
                        if is_material_mismatch
                        else "QUANTITY_ROUNDING_DEVIATION"
                    )
                    issue = GoldAuditIssue(
                        id=_issue_id(source_hash, sheet.name, row_number, code),
                        code=code,
                        severity="BLOCK" if is_material_mismatch else "REVIEW",
                        message=(
                            "原表工程量与计价公式复算值存在实质偏差"
                            if is_material_mismatch
                            else "原表工程量超出单元格显示精度允许的半单位舍入差"
                        ),
                        sheet_name=sheet.name,
                        row=row_number,
                        field="engineering_quantity",
                        actual=str(reported),
                        expected=str(recalculated),
                        difference=str(difference),
                        tolerance=str(tolerance),
                        evidence_cells=sum(
                            (
                                field_cells.get(field, [])
                                for field in (
                                    "width_mm",
                                    "length_mm",
                                    "quantity",
                                    "engineering_quantity",
                                    "unit",
                                )
                            ),
                            [],
                        ),
                    )
                    issues.append(issue)
                    issue_ids.append(issue.id)
                    if is_material_mismatch:
                        row_status = ReviewStatus.BLOCK

            if unit_price is not None and reported is not None:
                amount_column = mapping.get("amount")
                amount_cell = sheet.cell(row_number, amount_column or 0)
                canonical_amount_formula = _canonical_amount_formula(sheet, row_number, mapping)
                if (
                    amount_cell is not None
                    and amount_cell.formula
                    and (amount is None or not canonical_amount_formula)
                ):
                    code = "AMOUNT_NOT_CALCULABLE"
                    issue = GoldAuditIssue(
                        id=_issue_id(source_hash, sheet.name, row_number, code),
                        code=code,
                        severity="REVIEW",
                        message=(
                            "金额公式缺少缓存显示值，无法核对"
                            if amount is None
                            else "金额公式包含安装费、税费或其他价格分量，不能按工程量×材料单价核对"
                        ),
                        sheet_name=sheet.name,
                        row=row_number,
                        field="amount",
                        actual=str(amount) if amount is not None else None,
                        evidence_cells=sum(
                            (
                                field_cells.get(field, [])
                                for field in (
                                    "engineering_quantity",
                                    "unit_price",
                                    "amount",
                                )
                            ),
                            [],
                        ),
                    )
                    issues.append(issue)
                    issue_ids.append(issue.id)
                elif amount is not None:
                    expected_amount = unit_price * reported
                    amount_difference = amount - expected_amount
                    if abs(amount_difference) > Decimal("0.005"):
                        code = "AMOUNT_FORMULA_MISMATCH"
                        issue = GoldAuditIssue(
                            id=_issue_id(source_hash, sheet.name, row_number, code),
                            code=code,
                            severity="BLOCK",
                            message="原表金额与工程量×单价不一致",
                            sheet_name=sheet.name,
                            row=row_number,
                            field="amount",
                            actual=str(amount),
                            expected=str(expected_amount),
                            difference=str(amount_difference),
                            tolerance="0.005",
                            evidence_cells=sum(
                                (
                                    field_cells.get(field, [])
                                    for field in (
                                        "engineering_quantity",
                                        "unit_price",
                                        "amount",
                                    )
                                ),
                                [],
                            ),
                        )
                        issues.append(issue)
                        issue_ids.append(issue.id)
                        row_status = ReviewStatus.BLOCK

            item = TakeoffItem(
                sequence=sequence,
                name=name,
                mt_code=mt_code,
                material=material,
                plan_location=_value_repr(values["plan_location"]).strip() or None,
                elevation=_value_repr(values["elevation"]).strip() or None,
                detail=_value_repr(values["detail"]).strip() or None,
                unfolded_spec=_value_repr(values["unfolded_spec"]).strip() or None,
                width_mm=(
                    float(width_for_calculation) if width_for_calculation is not None else None
                ),
                length_mm=(
                    float(length_for_calculation) if length_for_calculation is not None else None
                ),
                quantity=(
                    float(quantity_for_calculation)
                    if quantity_for_calculation is not None
                    else None
                ),
                engineering_quantity=(
                    float(reported_for_item) if reported_for_item is not None else None
                ),
                unit=unit,  # type: ignore[arg-type]
                pricing_method=pricing_method or None,
                unit_price=(
                    float(unit_price_for_item) if unit_price_for_item is not None else None
                ),
                amount=float(amount_for_item) if amount_for_item is not None else None,
                note=_value_repr(values["note"]).strip() or None,
                evidence_ids=[
                    f"{sheet.name}!{cell.coordinate}"
                    for cell in raw_cells
                    if not _is_empty(cell.raw_value)
                ]
                + [
                    image.id
                    for image in sheet_images
                    if image.anchor_row <= row_number <= (image.end_row or image.anchor_row)
                ],
                status=row_status,
                block_reason=(
                    "人工金标准导入审计发现阻断项" if row_status == ReviewStatus.BLOCK else None
                ),
            )
            rows.append(
                GoldRow(
                    id=_row_id(source_hash, sheet.name, row_number),
                    sheet_name=sheet.name,
                    sheet_index=sheet.index,
                    row=row_number,
                    item=item,
                    raw_cells=raw_cells,
                    field_cells=field_cells,
                    source_material_code=source_material_code,
                    mt_code_source=mt_code_source,
                    image_evidence=[
                        image
                        for image in sheet_images
                        if image.anchor_row <= row_number <= (image.end_row or image.anchor_row)
                    ],
                    recalculated_engineering_quantity=(
                        float(recalculated) if recalculated is not None else None
                    ),
                    reported_engineering_quantity=(
                        float(reported) if reported is not None else None
                    ),
                    reported_quantity=float(quantity) if quantity is not None else None,
                    effective_quantity=(
                        float(quantity_for_calculation)
                        if quantity_for_calculation is not None
                        else None
                    ),
                    quantity_source=quantity_source,
                    quantity_derivation_basis=quantity_derivation_basis,
                    quantity_difference=float(difference) if difference is not None else None,
                    quantity_tolerance=float(tolerance) if tolerance is not None else None,
                    quantity_audit=quantity_audit,
                    audit_issue_ids=issue_ids,
                )
            )

        sheets.append(
            GoldSheetEvidence(
                sheet_name=sheet.name,
                sheet_index=sheet.index,
                max_row=sheet.max_row,
                max_column=sheet.max_column,
                header_rows=header_rows,
                header_cells=header_cells,
                field_columns=mapping,
                images=sheet_images,
                imported_row_count=len(rows) - imported_before,
            )
        )

    issue_counts = Counter(issue.code for issue in issues)
    status_counts = Counter(row.item.status.value for row in rows)
    summary = GoldImportSummary(
        sheet_count=sum(sheet.imported_row_count > 0 for sheet in sheets),
        row_count=len(rows),
        mt_distribution=dict(sorted(Counter(row.item.mt_code or "未识别" for row in rows).items())),
        unit_distribution=dict(sorted(Counter(row.item.unit or "未识别" for row in rows).items())),
        audit_issue_count=len(issues),
        quantity_mismatch_count=(
            issue_counts["QUANTITY_FORMULA_MISMATCH"] + issue_counts["QUANTITY_ROUNDING_DEVIATION"]
        ),
        quantity_material_mismatch_count=issue_counts["QUANTITY_FORMULA_MISMATCH"],
        quantity_rounding_deviation_count=issue_counts["QUANTITY_ROUNDING_DEVIATION"],
        quantity_not_calculable_count=issue_counts["QUANTITY_NOT_CALCULABLE"],
        quantity_derived_from_engineering_count=issue_counts[
            "QUANTITY_DERIVED_FROM_ENGINEERING"
        ],
        amount_mismatch_count=issue_counts["AMOUNT_FORMULA_MISMATCH"],
        amount_not_calculable_count=issue_counts["AMOUNT_NOT_CALCULABLE"],
        pass_count=status_counts[ReviewStatus.PASS.value],
        review_count=status_counts[ReviewStatus.REVIEW.value],
        block_count=status_counts[ReviewStatus.BLOCK.value],
    )
    return GoldImportResult(
        source_path=str(source),
        source_sha256=source_hash,
        workbook_format=workbook_format,
        sheets=sheets,
        rows=rows,
        issues=issues,
        summary=summary,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="导入并审计人工报价清单金标准")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", "-o", type=Path)
    args = parser.parse_args(argv)
    result = import_gold_workbook(args.workbook)
    if args.output:
        result.write_json(args.output)
    else:
        print(result.to_json())
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
