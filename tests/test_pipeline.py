import hashlib
import json
from pathlib import Path

import cadquote.pipeline as pipeline_module
import ezdxf
import pytest
from cad_quote import build_parser
from cadquote.models import (
    ComponentInstance,
    EvidenceEdge,
    MaterialSpec,
    MtOccurrence,
    ReviewStatus,
    RunIssue,
    Severity,
    Sheet,
    TakeoffItem,
)
from cadquote.pipeline import (
    ConfirmationBundle,
    _build_review_pack,
    _gate_pass_items_by_excel_evidence,
    _load_manifest_confirmations,
    _overall_status,
    _price_evidence_edges,
    _price_items,
    _snapshot_index_issues,
    load_confirmation_bundle,
    resume_pipeline,
    run_pipeline,
)
from cadquote.takeoff import TakeoffBuildResult
from openpyxl import load_workbook


def _write_price_book(path: Path, version: str = "price-v1") -> None:
    path.write_text(
        json.dumps(
            {
                "version": version,
                "approved": True,
                "source": f"controlled-{version}",
                "entries": [],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


def _write_minimal_dxf(path: Path) -> None:
    drawing = ezdxf.new("R2018")
    modelspace = drawing.modelspace()
    modelspace.add_text("平面布置图", dxfattribs={"insert": (0, 100), "height": 5})
    modelspace.add_text("MT-01", dxfattribs={"insert": (20, 20), "height": 5})
    drawing.saveas(path)


def test_review_pack_bounds_dense_relation_candidates_without_hiding_global_edges():
    plan = Sheet(id="plan", source_file_id="source", kind="plan")
    elevation_sheets = [
        Sheet(id=f"elevation-{index}", source_file_id="source", kind="elevation")
        for index in range(100)
    ]
    detail_sheets = [
        Sheet(id=f"detail-{index}", source_file_id="source", kind="detail")
        for index in range(3)
    ]
    occurrence = MtOccurrence(
        id="plan-occurrence",
        mt_code="MT-01",
        source_file_id="source",
        sheet_id="plan",
    )
    component = ComponentInstance(
        id="component:1",
        mt_code="MT-01",
        plan_occurrence_ids=[occurrence.id],
    )
    plan_edges = [
        EvidenceEdge(
            id=f"plan-edge-{index}",
            relation="plan_to_elevation",
            source_id="plan",
            target_id=f"elevation-{index}",
            confidence=0.8,
        )
        for index in range(100)
    ]
    detail_edges = [
        EvidenceEdge(
            id=f"detail-edge-{elevation}-{detail}",
            relation="elevation_to_detail",
            source_id=f"elevation-{elevation}",
            target_id=f"detail-{detail}",
            confidence=0.7,
        )
        for elevation in range(100)
        for detail in range(3)
    ]
    edges = [*plan_edges, *detail_edges]
    pack = _build_review_pack(
        TakeoffBuildResult(components=[component], evidence_edges=edges),
        [],
        [plan, *elevation_sheets, *detail_sheets],
        [],
        [occurrence],
        edges,
        ConfirmationBundle(),
    )

    group = pack["components"][0]
    assert len(group["relation_edge_candidates"]) == 48
    assert group["relation_candidate_truncation"] == {
        "elevation_to_detail": {"total": 300, "kept": 24},
        "plan_to_elevation": {"total": 100, "kept": 24},
    }
    assert len(pack["unassigned_relation_edges"]) == 352


def test_pricing_accepts_duplicate_identical_material_evidence(tmp_path: Path):
    price_book = tmp_path / "approved.json"
    price_book.write_text(
        json.dumps(
            {
                "version": "v1",
                "approved": True,
                "source": "采购审批单",
                "entries": [
                    {
                        "id": "price:1",
                        "version": "v1",
                        "approved": True,
                        "mt_code": "MT-01",
                        "material": "古铜色不锈钢",
                        "grade": "304",
                        "thickness_mm": 1.2,
                        "finish": "PVD",
                        "process": "折弯",
                        "pricing_method": "按实际展开面积计算",
                        "unit": "㎡",
                        "unit_price": 500,
                        "currency": "CNY",
                        "tax_included": True,
                        "source": "采购审批单第1行",
                    }
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    item = TakeoffItem(
        sequence=1,
        name="收边",
        mt_code="MT-01",
        material="古铜色不锈钢",
        width_mm=1_000,
        length_mm=2_000,
        quantity=1,
        engineering_quantity=2,
        unit="㎡",
        pricing_method="按实际展开面积计算",
        status=ReviewStatus.PASS,
    )
    materials = [
        MaterialSpec(
            id=f"material:{index}",
            mt_code="MT-01",
            name="古铜色不锈钢",
            grade="304",
            thickness_mm=1.2,
            finish="PVD",
            process="折弯",
        )
        for index in range(2)
    ]

    priced, issues, _ = _price_items(
        [item],
        materials,
        price_book,
        quote_date="2026-08-26",
        currency="CNY",
        tax_included=True,
    )

    assert not issues
    assert priced[0].status == ReviewStatus.PASS
    assert priced[0].unit_price == 500
    assert priced[0].amount == 1000


def test_pipeline_single_dxf_produces_auditable_review_package(tmp_path: Path):
    drawing = ezdxf.new("R2018")
    modelspace = drawing.modelspace()
    modelspace.add_text("平面布置图", dxfattribs={"insert": (0, 100), "height": 5})
    modelspace.add_text("MT-01", dxfattribs={"insert": (20, 20), "height": 5})
    source = tmp_path / "一层平面图.dxf"
    drawing.saveas(source)

    result = run_pipeline(source, tmp_path / "run", render_evidence=False)

    assert result.status.value == "BLOCK"
    assert result.counts["source_files"] == 1
    assert result.counts["mt_occurrences"] >= 1
    assert Path(result.quote_path).is_file()
    assert Path(result.manifest_path).is_file()
    assert Path(result.paths["review_pack"]).is_file()
    assert Path(result.paths["vector_quantity_probes"]).is_file()
    assert Path(result.paths["drawing_catalog"]).is_file()
    assert Path(result.paths["drawing_catalog_sqlite"]).is_file()
    workbook = load_workbook(result.quote_path)
    assert workbook.sheetnames == ["报价表", "来源追踪", "待确认", "运行信息"]
    assert workbook["待确认"].max_row >= 2

    index_path = Path(result.paths["index_json"])
    index_mtime = index_path.stat().st_mtime_ns
    first_review_pack = json.loads(Path(result.paths["review_pack"]).read_text(encoding="utf-8"))
    assert first_review_pack["summary"]["vector_quantity_review_candidate_count"] == 0
    assert first_review_pack["vector_quantity_evidence"]["policy"]["auto_quantity"] is False
    entity_id = first_review_pack["components"][0]["sources"]["entity_ids"][0]
    entity_record = first_review_pack["evidence_catalog"]["entities"][entity_id]
    source_record = first_review_pack["evidence_catalog"]["source_files"][
        entity_record["source_file_id"]
    ]
    assert source_record["relative_path"] == "一层平面图.dxf"
    assert {"handle", "layer", "space", "bbox", "text"} <= set(entity_record)
    component_id = first_review_pack["components"][0]["component"]["id"]
    confirmations = tmp_path / "confirmations.json"
    confirmations.write_text(
        json.dumps(
            {
                "schema_version": "1.0",
                "components": {
                    component_id: {
                        "selected": {"merge_component_ids": ["component:not-real"]},
                        "reviewer": "测试审核员",
                        "reviewed_at": "2026-08-26T10:00:00+08:00",
                        "reason": "保留审计元数据",
                    }
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    resumed = resume_pipeline(
        tmp_path / "run",
        confirmations=confirmations,
        render_evidence=False,
    )
    assert resumed.status == ReviewStatus.BLOCK
    assert index_path.stat().st_mtime_ns == index_mtime
    review_pack = json.loads(Path(resumed.paths["review_pack"]).read_text(encoding="utf-8"))
    assert review_pack["metadata"]["run_mode"] == "resume"
    assert review_pack["summary"]["component_count"] >= 1
    assert review_pack["components"][0]["component"]["id"].startswith("component:")
    assert review_pack["confirmation_audit"][component_id]["reviewer"] == "测试审核员"
    reused = _load_manifest_confirmations(tmp_path / "run" / "manifest.json")
    assert reused.audit[component_id]["reviewer"] == "测试审核员"
    assert reused.selections[component_id]["merge_component_ids"] == ["component:not-real"]
    assert "measurement_candidates" in review_pack["components"][0]
    assert "vector_quantity_probes" in review_pack["components"][0]
    assert "relation_edge_candidates" in review_pack["components"][0]
    assert review_pack["confirmation_template"]["optional_selection_examples"][
        "merge_component_ids"
    ]
    template = review_pack["confirmation_template"]["components"][component_id]
    assert "reviewed_at" in template
    assert "timestamp" not in template


def test_excel_evidence_gate_never_crosses_sequence_or_component_for_same_mt():
    items = [
        TakeoffItem(
            sequence=1,
            name="构件一",
            mt_code="MT-01",
            component_id="component:1",
            amount=100,
            status=ReviewStatus.PASS,
        ),
        TakeoffItem(
            sequence=2,
            name="构件二",
            mt_code="MT-01",
            component_id="component:2",
            amount=200,
            status=ReviewStatus.PASS,
        ),
    ]
    records = [
        {
            "id": f"record:{sequence}:{component_id}:{stage}",
            "sequence": sequence,
            "component_id": component_id,
            "mt_code": "MT-01",
            "stage": stage,
            "render_state": "RENDERED",
            "context_image": "context.png",
            "detail_image": "detail.png",
        }
        for sequence, component_id in ((1, "component:1"), (1, "component:2"), (2, "component:1"))
        for stage in ("plan", "elevation", "detail")
    ]

    gated, issues = _gate_pass_items_by_excel_evidence(items, records)

    assert gated[0].status == ReviewStatus.PASS
    assert gated[0].amount == 100
    assert gated[1].status == ReviewStatus.REVIEW
    assert gated[1].amount is None
    assert len(issues) == 1
    assert issues[0].source_id == "component:2"


def test_excel_evidence_render_exception_writes_three_explicit_failed_rows(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    item = TakeoffItem(
        sequence=1,
        name="构件",
        mt_code="MT-01",
        component_id="component:1",
        amount=100,
        status=ReviewStatus.PASS,
    )
    takeoff = TakeoffBuildResult(
        components=[ComponentInstance(id="component:1", mt_code="MT-01")]
    )

    def fail_render(*_args, **_kwargs):
        raise RuntimeError("synthetic renderer failure")

    monkeypatch.setattr(pipeline_module, "render_excel_evidence", fail_render)
    issues: list[RunIssue] = []
    payload, counts, index_path = pipeline_module._prepare_excel_evidence(
        [item],
        takeoff,
        [],
        [],
        [],
        {},
        tmp_path / "excel_evidence",
        issues,
    )

    assert {record["stage"] for record in payload} == {"plan", "elevation", "detail"}
    assert all(record["render_state"] == "FAILED" for record in payload)
    assert all(record["image_root"] == str(tmp_path / "excel_evidence") for record in payload)
    assert counts == {
        "excel_evidence_records": 3,
        "excel_evidence_rendered": 0,
        "excel_evidence_missing": 0,
        "excel_evidence_failed": 3,
    }
    assert index_path.is_file()
    assert "EXCEL_EVIDENCE_PIPELINE_FAILED" in {issue.code for issue in issues}


def test_full_and_resume_export_excel_evidence_sheet(tmp_path: Path):
    source = tmp_path / "evidence.dxf"
    _write_minimal_dxf(source)

    first = run_pipeline(source, tmp_path / "run")

    assert Path(first.paths["excel_evidence"]).is_file()
    assert first.counts["excel_evidence_records"] >= 3
    assert (
        first.counts["excel_evidence_rendered"]
        + first.counts["excel_evidence_missing"]
        + first.counts["excel_evidence_failed"]
        == first.counts["excel_evidence_records"]
    )
    assert "截图证据" in load_workbook(first.quote_path).sheetnames

    resumed = resume_pipeline(tmp_path / "run")

    assert Path(resumed.paths["excel_evidence"]).is_file()
    assert resumed.counts["excel_evidence_records"] >= 3
    assert "截图证据" in load_workbook(resumed.quote_path).sheetnames


def test_confirmation_parser_preserves_reviewer_audit(tmp_path: Path):
    path = tmp_path / "confirmations.json"
    path.write_text(
        json.dumps(
            {
                "schema_version": "1.0",
                "components": {
                    "component:1": {
                        "selected": {"length": "measurement:1"},
                        "reviewer": "张工",
                        "timestamp": "2026-08-26T03:00:00+08:00",
                        "reason": "与立面尺寸一致",
                    }
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    bundle = load_confirmation_bundle(path)
    assert bundle.selections == {"component:1": {"length": "measurement:1"}}
    assert bundle.audit["component:1"]["reviewer"] == "张工"
    assert bundle.audit["component:1"]["reviewed_at"] == "2026-08-26T03:00:00+08:00"
    assert "timestamp" not in bundle.audit["component:1"]
    assert bundle.audit["component:1"]["reason"] == "与立面尺寸一致"


def test_confirmation_parser_preserves_auditable_derived_measurement(tmp_path: Path):
    path = tmp_path / "derived-confirmations.json"
    path.write_text(
        json.dumps(
            {
                "schema_version": "1.0",
                "components": {
                    "component:1": {
                        "selected": {
                            "length": {
                                "kind": "derived_measurement",
                                "expression": "left+middle+right",
                                "terms": [
                                    {"symbol": "left", "candidate_id": "measurement:1"},
                                    {"symbol": "middle", "candidate_id": "measurement:2"},
                                    {"symbol": "right", "candidate_id": "measurement:3"},
                                ],
                                "unit": "mm",
                                "basis": "同一立面连续三段尺寸链",
                            }
                        },
                        "reviewer": "张工",
                        "reviewed_at": "2026-08-26T03:00:00+08:00",
                        "reason": "逐段核对实体句柄",
                    }
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    bundle = load_confirmation_bundle(path)

    derived = bundle.selections["component:1"]["length"]
    assert derived["kind"] == "derived_measurement"
    assert derived["expression"] == "left+middle+right"
    assert derived["terms"][1] == {
        "symbol": "middle",
        "candidate_id": "measurement:2",
    }
    assert bundle.audit["component:1"]["selected"]["length"] == derived


def test_confirmation_parser_preserves_audited_engineering_quantity_expression(
    tmp_path: Path,
):
    path = tmp_path / "engineering-quantity-confirmations.json"
    path.write_text(
        json.dumps(
            {
                "schema_version": "1.0",
                "components": {
                    "component:1": {
                        "selected": {
                            "unit": "m",
                            "engineering_quantity": {
                                "kind": "engineering_quantity_expression",
                                "expression": "length_mm*2/1000",
                                "basis": "立面证实两条独立实体线",
                                "evidence_ids": [" entity:left ", "entity:right"],
                            }
                        },
                        "reviewer": "张工",
                        "reviewed_at": "2026-08-26T03:00:00+08:00",
                        "reason": "核对构件拓扑与计价轴",
                    }
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    bundle = load_confirmation_bundle(path)

    assert bundle.selections["component:1"]["engineering_quantity"] == {
        "kind": "engineering_quantity_expression",
        "expression": "length_mm*2/1000",
        "basis": "立面证实两条独立实体线",
        "evidence_ids": ["entity:left", "entity:right"],
    }


def test_confirmation_parser_preserves_reference_only_composite_assembly(
    tmp_path: Path,
) -> None:
    path = tmp_path / "composite-confirmations.json"
    path.write_text(
        json.dumps(
            {
                "schema_version": "1.0",
                "components": {
                    "component:1": {
                        "selected": {
                            "unit": "㎡",
                            "composite_assembly": {
                                "kind": "single_line_composite",
                                "assembly_type": "screen_with_glass",
                                "billing_basis": "whole_elevation_projection",
                                "required_material_roles": ["glass_infill"],
                                "included_materials": [
                                    {
                                        "role": "glass_infill",
                                        "material_spec_id": " material:glass ",
                                    }
                                ],
                                "basis": "框架与玻璃属于同一成品总成",
                                "evidence_ids": [" entity:glass ", "entity:glass"],
                                "projection_width_candidate_id": " measurement:width ",
                                "projection_length_candidate_id": " measurement:length ",
                                "projection_component_entity_id": " entity:screen-frame ",
                                "projection_axis_basis": " 已核对整樘立面外轮廓 ",
                            },
                        },
                        "reviewer": "张工",
                        "reviewed_at": "2026-08-26T03:00:00+08:00",
                        "reason": "核对屏风立面和玻璃标注",
                    }
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    bundle = load_confirmation_bundle(path)
    assembly = bundle.selections["component:1"]["composite_assembly"]
    assert assembly["included_materials"] == [
        {"role": "glass_infill", "material_spec_id": "material:glass"}
    ]
    assert assembly["evidence_ids"] == ["entity:glass"]
    assert assembly["projection_width_candidate_id"] == "measurement:width"
    assert assembly["projection_length_candidate_id"] == "measurement:length"
    assert assembly["projection_component_entity_id"] == "entity:screen-frame"
    assert assembly["projection_axis_basis"] == "已核对整樘立面外轮廓"

    invalid_roles = tmp_path / "invalid-composite-roles.json"
    invalid_roles_payload = json.loads(path.read_text(encoding="utf-8"))
    invalid_roles_payload["components"]["component:1"]["selected"][
        "composite_assembly"
    ]["required_material_roles"] = ["included_other"]
    invalid_roles.write_text(
        json.dumps(invalid_roles_payload, ensure_ascii=False),
        encoding="utf-8",
    )
    with pytest.raises(ValueError, match="must include 'glass_infill'"):
        load_confirmation_bundle(invalid_roles)

    assembly["included_materials"][0]["material_code"] = "GC-GL-01"
    invalid = tmp_path / "invalid-composite-confirmations.json"
    payload = json.loads(path.read_text(encoding="utf-8"))
    payload["components"]["component:1"]["selected"]["composite_assembly"] = assembly
    invalid.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    with pytest.raises(ValueError, match="unsupported keys"):
        load_confirmation_bundle(invalid)


def test_confirmation_parser_rejects_target_value_engineering_quantity(
    tmp_path: Path,
):
    path = tmp_path / "target-value-engineering-quantity.json"
    path.write_text(
        json.dumps(
            {
                "components": {
                    "component:1": {
                        "selected": {
                            "unit": "m",
                            "engineering_quantity": {
                                "kind": "engineering_quantity_expression",
                                "expression": "7.25",
                                "basis": "直接复制目标值",
                                "evidence_ids": ["entity:1"],
                            },
                        }
                    }
                }
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="must reference a CAD-backed field"):
        load_confirmation_bundle(path)


def test_confirmation_parser_rejects_unsafe_engineering_quantity_expression(
    tmp_path: Path,
):
    path = tmp_path / "unsafe-engineering-quantity-confirmations.json"
    path.write_text(
        json.dumps(
            {
                "components": {
                    "component:1": {
                        "selected": {
                            "engineering_quantity": {
                                "kind": "engineering_quantity_expression",
                                "expression": "SUM(length_mm)",
                                "basis": "人工输入",
                                "evidence_ids": ["entity:1"],
                            }
                        }
                    }
                }
            }
        ),
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="unsupported engineering quantity expression"):
        load_confirmation_bundle(path)


def test_confirmation_parser_preserves_audited_not_applicable_detail(tmp_path: Path):
    path = tmp_path / "no-detail-confirmations.json"
    path.write_text(
        json.dumps(
            {
                "schema_version": "1.0",
                "components": {
                    "component:1": {
                        "selected": {
                            "detail_requirement": {
                                "kind": "not_applicable",
                                "basis": "立面已包含全部计量尺寸，且无节点索引",
                                "searched_sheet_ids": [" elevation ", "elevation"],
                                "reference_entity_ids": ["entity:1"],
                            }
                        },
                        "reviewer": "张工",
                        "reviewed_at": "2026-08-26T03:00:00+08:00",
                        "reason": "复核立面及索引范围",
                    }
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    bundle = load_confirmation_bundle(path)

    assert bundle.selections["component:1"]["detail_requirement"] == {
        "kind": "not_applicable",
        "basis": "立面已包含全部计量尺寸，且无节点索引",
        "searched_sheet_ids": ["elevation"],
        "reference_entity_ids": ["entity:1"],
    }


def test_confirmation_parser_rejects_not_applicable_detail_without_search_scope(
    tmp_path: Path,
):
    path = tmp_path / "invalid-no-detail-confirmations.json"
    path.write_text(
        json.dumps(
            {
                "schema_version": "1.0",
                "components": {
                    "component:1": {
                        "selected": {
                            "detail_requirement": {
                                "kind": "not_applicable",
                                "basis": "未发现节点",
                                "searched_sheet_ids": [],
                            }
                        }
                    }
                },
            }
        ),
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="searched_sheet_ids"):
        load_confirmation_bundle(path)


def test_confirmation_parser_rejects_derived_measurement_without_source_terms(
    tmp_path: Path,
):
    path = tmp_path / "invalid-derived-confirmations.json"
    path.write_text(
        json.dumps(
            {
                "schema_version": "1.0",
                "components": {
                    "component:1": {
                        "selected": {
                            "length": {
                                "kind": "derived_measurement",
                                "expression": "7080",
                                "terms": [],
                                "basis": "人工填值",
                            }
                        }
                    }
                },
            }
        ),
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="terms"):
        load_confirmation_bundle(path)


def test_confirmation_parser_accepts_legacy_mapping(tmp_path: Path):
    path = tmp_path / "legacy.json"
    path.write_text(
        json.dumps({"component:1": {"quantity": "measurement:2"}}),
        encoding="utf-8",
    )
    bundle = load_confirmation_bundle(path)
    assert bundle.selections["component:1"]["quantity"] == "measurement:2"
    assert bundle.audit["component:1"]["format"] == "legacy"


@pytest.mark.parametrize(
    ("audit", "message"),
    [
        (
            {
                "reviewed_at": "2026-08-26T10:00:00+08:00",
                "reason": "已核图",
            },
            "reviewer",
        ),
        (
            {
                "reviewer": "张工",
                "reviewed_at": "2026-08-26T10:00:00+08:00",
            },
            "reason",
        ),
        (
            {
                "reviewer": "张工",
                "reviewed_at": "2026-08-26T10:00:00",
                "reason": "已核图",
            },
            "timezone",
        ),
        (
            {
                "reviewer": "张工",
                "reviewed_at": "不是时间",
                "reason": "已核图",
            },
            "ISO 8601",
        ),
    ],
)
def test_confirmation_parser_rejects_incomplete_audit_for_effective_selection(
    tmp_path: Path,
    audit: dict[str, str],
    message: str,
):
    path = tmp_path / "invalid-audit.json"
    path.write_text(
        json.dumps(
            {
                "schema_version": "1.0",
                "components": {
                    "component:1": {
                        **audit,
                        "selected": {"length": "measurement:1"},
                    }
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match=message):
        load_confirmation_bundle(path)


def test_empty_structured_confirmation_does_not_require_audit_metadata(tmp_path: Path):
    path = tmp_path / "empty-selection.json"
    path.write_text(
        json.dumps(
            {
                "schema_version": "1.0",
                "components": {"component:1": {"selected": {}}},
            }
        ),
        encoding="utf-8",
    )

    bundle = load_confirmation_bundle(path)
    assert bundle.selections == {"component:1": {}}


def test_full_and_resume_runs_block_legacy_unaudited_confirmations(tmp_path: Path):
    drawing = ezdxf.new("R2018")
    drawing.modelspace().add_text("MT-01", dxfattribs={"insert": (20, 20), "height": 5})
    source = tmp_path / "legacy-plan.dxf"
    drawing.saveas(source)
    legacy = tmp_path / "legacy-confirmations.json"
    legacy.write_text(
        json.dumps({"component:unknown": {"quantity": "measurement:unknown"}}),
        encoding="utf-8",
    )

    full = run_pipeline(
        source,
        tmp_path / "legacy-run",
        confirmations=legacy,
        render_evidence=False,
    )
    full_issue = next(
        issue for issue in full.issues if issue.code == "LEGACY_CONFIRMATIONS_UNAUDITED"
    )
    assert full_issue.severity == Severity.BLOCK
    assert full.status != ReviewStatus.PASS
    otherwise_commercial_pass = TakeoffItem(
        sequence=1,
        name="已核价构件",
        mt_code="MT-01",
        amount=100,
        unit_price=100,
        status=ReviewStatus.PASS,
    )
    assert _overall_status([otherwise_commercial_pass], [full_issue]) == ReviewStatus.REVIEW

    resumed = resume_pipeline(tmp_path / "legacy-run", render_evidence=False)
    resume_issue = next(
        issue for issue in resumed.issues if issue.code == "LEGACY_CONFIRMATIONS_UNAUDITED"
    )
    assert resume_issue.severity == Severity.BLOCK
    assert resumed.status != ReviewStatus.PASS


def test_confirmation_parser_accepts_only_merge_component_id_arrays(tmp_path: Path):
    path = tmp_path / "merge-confirmations.json"
    path.write_text(
        json.dumps(
            {
                "schema_version": "1.0",
                "components": {
                    "component:target": {
                        "reviewer": "审核员",
                        "reviewed_at": "2026-08-26T10:00:00+08:00",
                        "reason": "确认两个来源属于同一物理构件",
                        "selected": {
                            "merge_component_ids": [
                                " component:source-1 ",
                                "component:source-2",
                                "component:source-1",
                            ],
                            "unit": "件",
                        }
                    }
                },
            }
        ),
        encoding="utf-8",
    )
    bundle = load_confirmation_bundle(path)
    assert bundle.selections["component:target"] == {
        "merge_component_ids": ["component:source-1", "component:source-2"],
        "unit": "件",
    }

    invalid = tmp_path / "invalid-confirmations.json"
    invalid.write_text(
        json.dumps(
            {
                "schema_version": "1.0",
                "components": {"component:target": {"selected": {"quantity": ["measurement:1"]}}},
            }
        ),
        encoding="utf-8",
    )
    with pytest.raises(ValueError, match="quantity"):
        load_confirmation_bundle(invalid)


def test_manifest_confirmation_parser_preserves_merge_arrays_for_resume(
    tmp_path: Path,
):
    manifest = tmp_path / "manifest.json"
    manifest.write_text(
        json.dumps(
            {
                "metadata": {
                    "confirmation_schema_version": "1.0",
                    "confirmation_audit": {
                        "component:target": {
                            "reviewer": "审核员",
                            "timestamp": "2026-08-26T10:00:00+08:00",
                            "reason": "复核归并来源与目标",
                            "selected": {"merge_component_ids": ["component:source"]},
                        }
                    },
                }
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    bundle = _load_manifest_confirmations(manifest)
    assert bundle.selections == {"component:target": {"merge_component_ids": ["component:source"]}}
    assert bundle.audit["component:target"]["reviewer"] == "审核员"
    assert bundle.audit["component:target"]["reviewed_at"] == "2026-08-26T10:00:00+08:00"
    assert "timestamp" not in bundle.audit["component:target"]


def test_overall_status_blocks_all_blocked_and_gates_severe_issues():
    blocked = TakeoffItem(
        sequence=1,
        name="构件",
        mt_code="MT-01",
        status=ReviewStatus.BLOCK,
    )
    assert _overall_status([blocked]) == ReviewStatus.BLOCK

    passed = blocked.model_copy(
        update={"status": ReviewStatus.PASS, "amount": 100, "unit_price": 100}
    )
    severe = RunIssue(
        stage="index",
        severity=Severity.ERROR,
        code="INDEX_ERROR",
        message="索引不完整",
    )
    assert _overall_status([passed]) == ReviewStatus.PASS
    assert _overall_status([passed], [severe]) == ReviewStatus.REVIEW


def test_index_snapshot_audit_signals_become_gate_issues():
    issues = _snapshot_index_issues(
        [
            {
                "source_file_id": "file:1",
                "source_path": "drawing.dxf",
                "audit_error_count": 2,
                "audit_fix_count": 3,
                "recovered": True,
            }
        ]
    )
    by_code = {issue.code: issue for issue in issues}
    assert by_code["DXF_AUDIT_ERRORS"].severity == Severity.ERROR
    assert by_code["DXF_RECOVERED_INPUT"].severity == Severity.ERROR
    assert by_code["DXF_AUDIT_FIXES_APPLIED"].severity == Severity.WARNING


def test_exact_price_selection_adds_pass_evidence_edge():
    priced = TakeoffItem(
        sequence=1,
        name="构件",
        mt_code="MT-01",
        component_id="component:1",
        price_entry_id="price:1",
        unit_price=500,
    )
    edges = _price_evidence_edges(
        [priced],
        {
            "price_book": "approved.json",
            "price_book_version": "v1",
            "quote_date": "2026-08-26",
            "currency": "CNY",
            "tax_included": True,
        },
    )
    assert len(edges) == 1
    assert edges[0].relation == "component_to_price"
    assert edges[0].status == ReviewStatus.PASS
    assert edges[0].target_id == "price:1"


def test_full_manifest_records_reproducible_pricing_context_and_resume_inherits_it(
    tmp_path: Path,
):
    source = tmp_path / "pricing-context.dxf"
    price_book = tmp_path / "approved-price-book.json"
    _write_minimal_dxf(source)
    _write_price_book(price_book)

    run = run_pipeline(
        source,
        tmp_path / "run",
        price_book=price_book,
        quote_date="2026-08-01",
        currency="USD",
        tax_included=False,
        render_evidence=False,
    )
    first_manifest = json.loads(Path(run.manifest_path).read_text(encoding="utf-8"))
    first_price = first_manifest["metadata"]["price"]
    expected_sha256 = hashlib.sha256(price_book.read_bytes()).hexdigest()
    assert first_price == {
        "price_book": str(price_book.resolve()),
        "price_book_sha256": expected_sha256,
        "price_book_version": "price-v1",
        "price_book_approved": True,
        "price_book_integrity": "VERIFIED",
        "quote_date": "2026-08-01",
        "currency": "USD",
        "tax_included": False,
    }

    resumed = resume_pipeline(tmp_path / "run", render_evidence=False)
    resumed_manifest = json.loads(Path(resumed.manifest_path).read_text(encoding="utf-8"))
    resumed_price = resumed_manifest["metadata"]["price"]
    assert resumed_price == first_price
    audit = resumed_manifest["metadata"]["pricing_context_audit"][-1]
    assert audit["explicit_fields"] == []
    assert audit["changes"] == {}
    assert audit["price_book_integrity"]["status"] == "VERIFIED"


def test_resume_preserves_material_snapshot_and_does_not_duplicate_recovery_issue(
    tmp_path: Path,
):
    source = tmp_path / "materials-resume.dxf"
    _write_minimal_dxf(source)
    drawing = ezdxf.readfile(source)
    drawing.modelspace().add_text(
        "MT-01 304不锈钢 1.2mm 拉丝",
        dxfattribs={"insert": (40, 20), "height": 5},
    )
    drawing.saveas(source)

    first = run_pipeline(source, tmp_path / "run", render_evidence=False)
    materials_path = tmp_path / "run" / "analysis" / "materials.json"
    first_bytes = materials_path.read_bytes()
    first_mtime = materials_path.stat().st_mtime_ns
    first_recovery_count = sum(
        issue.code == "CAD_MATERIAL_SPECS_RECOVERED" for issue in first.issues
    )
    assert first_recovery_count == 1

    resumed = resume_pipeline(tmp_path / "run", render_evidence=False)

    assert materials_path.read_bytes() == first_bytes
    assert materials_path.stat().st_mtime_ns == first_mtime
    assert sum(
        issue.code == "CAD_MATERIAL_SPECS_RECOVERED" for issue in resumed.issues
    ) == first_recovery_count


@pytest.mark.parametrize("mutation", ["missing", "changed"])
def test_resume_blocks_when_inherited_price_book_cannot_be_verified(
    tmp_path: Path,
    mutation: str,
):
    source = tmp_path / "price-integrity.dxf"
    price_book = tmp_path / "approved-price-book.json"
    _write_minimal_dxf(source)
    _write_price_book(price_book)
    run_pipeline(
        source,
        tmp_path / "run",
        price_book=price_book,
        quote_date="2026-08-01",
        render_evidence=False,
    )
    expected_sha256 = hashlib.sha256(price_book.read_bytes()).hexdigest()
    if mutation == "missing":
        price_book.unlink()
        expected_code = "PRICE_BOOK_ORIGINAL_MISSING"
    else:
        _write_price_book(price_book, version="silently-replaced")
        expected_code = "PRICE_BOOK_HASH_MISMATCH"

    resumed = resume_pipeline(tmp_path / "run", render_evidence=False)

    assert resumed.status == ReviewStatus.BLOCK
    assert expected_code in {issue.code for issue in resumed.issues}
    assert all(
        item["status"] == "BLOCK"
        for item in json.loads(Path(resumed.paths["takeoff"]).read_text(encoding="utf-8"))
    )
    manifest = json.loads(Path(resumed.manifest_path).read_text(encoding="utf-8"))
    assert manifest["metadata"]["price"]["price_book_sha256"] == expected_sha256
    assert manifest["metadata"]["price"]["price_book_integrity"] == "FAILED"
    assert manifest["metadata"]["pricing_context_audit"][-1]["price_book_integrity"][
        "status"
    ] == "FAILED"


def test_resume_explicit_repricing_is_allowed_and_audited(tmp_path: Path):
    source = tmp_path / "explicit-reprice.dxf"
    first_book = tmp_path / "price-v1.json"
    second_book = tmp_path / "price-v2.json"
    _write_minimal_dxf(source)
    _write_price_book(first_book, version="v1")
    _write_price_book(second_book, version="v2")
    run_pipeline(
        source,
        tmp_path / "run",
        price_book=first_book,
        quote_date="2026-08-01",
        currency="CNY",
        tax_included=False,
        render_evidence=False,
    )

    resumed = resume_pipeline(
        tmp_path / "run",
        price_book=second_book,
        quote_date="2026-08-15",
        currency="USD",
        tax_included=True,
        render_evidence=False,
    )

    assert "PRICE_BOOK_HASH_MISMATCH" not in {issue.code for issue in resumed.issues}
    manifest = json.loads(Path(resumed.manifest_path).read_text(encoding="utf-8"))
    price = manifest["metadata"]["price"]
    assert price["price_book"] == str(second_book.resolve())
    assert price["price_book_version"] == "v2"
    assert price["quote_date"] == "2026-08-15"
    assert price["currency"] == "USD"
    assert price["tax_included"] is True
    audit = manifest["metadata"]["pricing_context_audit"][-1]
    assert audit["explicit_fields"] == [
        "currency",
        "price_book",
        "quote_date",
        "tax_included",
    ]
    assert set(audit["changes"]) >= {
        "price_book",
        "price_book_sha256",
        "price_book_version",
        "quote_date",
        "currency",
        "tax_included",
    }
    assert audit["price_book_integrity"]["status"] == "REPLACED_EXPLICITLY"
    assert "PRICING_CONTEXT_EXPLICIT_OVERRIDE" in {
        issue.code for issue in resumed.issues
    }


def test_cli_run_has_defaults_but_resume_can_distinguish_omitted_context():
    parser = build_parser()
    run_args = parser.parse_args(["run", "input.dxf", "--out", "run"])
    resume_args = parser.parse_args(["resume", "run"])
    explicit_resume_args = parser.parse_args(
        [
            "resume",
            "run",
            "--currency",
            "USD",
            "--quote-date",
            "2026-08-15",
            "--tax-included",
            "false",
        ]
    )

    assert run_args.currency == "CNY"
    assert run_args.quote_date is None
    assert run_args.tax_included is None
    assert resume_args.currency is None
    assert resume_args.quote_date is None
    assert resume_args.tax_included is None
    assert explicit_resume_args.currency == "USD"
    assert explicit_resume_args.quote_date == "2026-08-15"
    assert explicit_resume_args.tax_included is False
