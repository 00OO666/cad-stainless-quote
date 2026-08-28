import json
import struct
import zlib
from argparse import Namespace
from pathlib import Path

import pytest
from cad_quote import command_quote
from cadquote.calculation import (
    calculate_item,
    engineering_quantity_expression_to_excel,
    evaluate_numeric_expression,
)
from cadquote.evaluation import evaluate_takeoff
from cadquote.exporter import QUOTE_HEADERS, build_quote_workbook
from cadquote.models import (
    EvidenceEdge,
    MaterialSpec,
    MeasurementCandidate,
    PriceBook,
    PriceEntry,
    ReviewStatus,
    RunIssue,
    Severity,
    TakeoffItem,
)
from cadquote.pricing import apply_price, load_price_book
from openpyxl import Workbook, load_workbook


def item(**updates):
    values = {
        "sequence": 1,
        "name": "不锈钢收口线",
        "mt_code": "MT-01",
        "unfolded_spec": "10+180+10",
        "length_mm": 5000,
        "quantity": 4,
        "unit": "㎡",
        "pricing_method": "按实际展开面积计算",
        "status": ReviewStatus.REVIEW,
    }
    values.update(updates)
    return TakeoffItem(**values)


def _write_test_png(path: Path, width: int = 600, height: int = 300) -> None:
    def chunk(kind: bytes, data: bytes) -> bytes:
        payload = kind + data
        return (
            struct.pack(">I", len(data))
            + payload
            + struct.pack(">I", zlib.crc32(payload) & 0xFFFFFFFF)
        )

    scanlines = b"".join(b"\x00" + b"\xE8\xEE\xF5" * width for _ in range(height))
    path.write_bytes(
        b"\x89PNG\r\n\x1a\n"
        + chunk(b"IHDR", struct.pack(">IIBBBBB", width, height, 8, 2, 0, 0, 0))
        + chunk(b"IDAT", zlib.compress(scanlines))
        + chunk(b"IEND", b"")
    )


def test_xlsx_price_book_rejects_commercially_incomplete_columns(tmp_path: Path):
    path = tmp_path / "incomplete-price-book.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "价格库"
    sheet.append(["MT编号", "计价方式", "单位", "单价"])
    sheet.append(["MT-01", "按米", "m", 100])
    workbook.save(path)

    with pytest.raises(ValueError, match="price book missing columns"):
        load_price_book(path)


def test_safe_expression_and_quantity_calculation():
    assert evaluate_numeric_expression("5+10+120+10") == 145
    with pytest.raises(ValueError):
        evaluate_numeric_expression("__import__('os').system('whoami')")
    for invalid in (True, float("nan"), float("inf"), 1e16):
        with pytest.raises(ValueError):
            evaluate_numeric_expression(invalid)
    calculated = calculate_item(item())
    assert calculated.width_mm == 200
    assert calculated.engineering_quantity == 4.0
    assert calculated.status == ReviewStatus.REVIEW


def test_linear_and_piece_calculation():
    linear = calculate_item(
        item(unfolded_spec=None, width_mm=None, length_mm=2500, quantity=2, unit="m")
    )
    assert linear.engineering_quantity == 5
    piece = calculate_item(
        item(unfolded_spec=None, width_mm=None, length_mm=None, quantity=3, unit="件")
    )
    assert piece.engineering_quantity == 3


def test_audited_engineering_quantity_expression_supports_billing_axis_and_multiplier():
    width_axis = calculate_item(
        item(
            unfolded_spec=None,
            width_mm=3600,
            length_mm=900,
            quantity=1,
            unit="m",
            engineering_quantity_expression="width_mm * quantity / 1000",
            engineering_quantity_basis="线性构件按平面可见长边计延米",
            engineering_quantity_evidence_ids=["entity:width:3600"],
        )
    )
    assert width_axis.engineering_quantity == 3.6

    two_runs = calculate_item(
        item(
            unfolded_spec=None,
            width_mm=70,
            length_mm=3000,
            quantity=1,
            unit="m",
            engineering_quantity_expression="length_mm * 2 / 1000",
            engineering_quantity_basis="节点证实同一构件包含两条实体线",
            engineering_quantity_evidence_ids=["entity:run:left", "entity:run:right"],
        )
    )
    assert two_runs.quantity == 1
    assert two_runs.engineering_quantity == 6


@pytest.mark.parametrize(
    "expression",
    [
        "unknown_mm / 1000",
        "__import__('os').system('whoami')",
        "length_mm / 0",
        "7.25",
        "length_mm + quantity",
        "length_mm * quantity / 1000000",
        "length_mm * 1.05 / 1000",
        "length_mm * (quantity-quantity+2) / 1000",
        "1e308*1e308/1e308/1e308",
    ],
)
def test_invalid_engineering_quantity_expression_blocks(expression: str):
    calculated = calculate_item(
        item(
            unit="m",
            engineering_quantity_expression=expression,
            engineering_quantity_basis="已核对",
            engineering_quantity_evidence_ids=["entity:1"],
            engineering_quantity=999,
            amount=999,
            status=ReviewStatus.PASS,
        )
    )
    assert calculated.status == ReviewStatus.BLOCK
    assert calculated.engineering_quantity is None
    assert calculated.amount is None
    assert "表达式" in (calculated.block_reason or "")


def test_custom_engineering_quantity_expression_requires_audit_evidence():
    calculated = calculate_item(
        item(
            unit="m",
            engineering_quantity_expression="length_mm * 2 / 1000",
            status=ReviewStatus.PASS,
        )
    )
    assert calculated.status == ReviewStatus.BLOCK
    assert calculated.engineering_quantity is None
    assert "计算依据" in (calculated.block_reason or "")
    assert "证据ID" in (calculated.block_reason or "")


def test_custom_engineering_quantity_expression_requires_visible_quantity():
    calculated = calculate_item(
        item(
            unit="m",
            quantity=None,
            engineering_quantity_expression="length_mm*2/1000",
            engineering_quantity_basis="两条计价线",
            engineering_quantity_evidence_ids=["entity:1", "entity:2"],
            status=ReviewStatus.PASS,
        )
    )
    assert calculated.status == ReviewStatus.BLOCK
    assert calculated.engineering_quantity is None
    assert calculated.block_reason == "缺少构件数量"


def test_area_engineering_quantity_expression_enforces_dimension_and_scale():
    valid = calculate_item(
        item(
            unit="㎡",
            width_mm=200,
            length_mm=5000,
            quantity=4,
            engineering_quantity_expression="width_mm*length_mm*quantity/1000000",
            engineering_quantity_basis="已确认展开宽、长度和四个实例",
            engineering_quantity_evidence_ids=["entity:w", "entity:l", "entity:q"],
        )
    )
    assert valid.engineering_quantity == 4
    invalid = calculate_item(
        item(
            unit="㎡",
            engineering_quantity_expression="length_mm*quantity/1000000",
            engineering_quantity_basis="缺少面积第二轴",
            engineering_quantity_evidence_ids=["entity:l", "entity:q"],
            status=ReviewStatus.PASS,
        )
    )
    assert invalid.status == ReviewStatus.BLOCK
    assert "dimension" in (invalid.block_reason or "")


def test_engineering_quantity_expression_excel_compiler_is_fixed_grammar():
    assert (
        engineering_quantity_expression_to_excel(
            "width_mm * quantity / 1000",
            row=2,
        )
        == "=((I2*K2)/1000)"
    )
    with pytest.raises(ValueError):
        engineering_quantity_expression_to_excel("SUM(width_mm)", row=2)
    with pytest.raises(ValueError, match="unsafe numeric literal"):
        engineering_quantity_expression_to_excel("1e309", row=2)


def test_exact_approved_price_match():
    material = MaterialSpec(
        id="m1",
        mt_code="MT-01",
        name="古铜色不锈钢",
        grade="304",
        thickness_mm=1.2,
        finish="PVD",
        process="折弯",
    )
    entry = PriceEntry(
        id="p1",
        version="v1",
        approved=True,
        mt_code="MT-01",
        material="古铜色不锈钢",
        grade="304",
        thickness_mm=1.2,
        finish="PVD",
        process="折弯",
        pricing_method="按实际展开面积计算",
        unit="㎡",
        unit_price=500,
        tax_included=True,
        source="approved.xlsx",
    )
    book = PriceBook(version="v1", approved=True, source="approved.xlsx", entries=[entry])
    priced, issues = apply_price(item(), material, book)
    assert not issues
    assert priced.unit_price == 500
    assert priced.price_entry_id == "p1"
    assert calculate_item(priced).amount == 2000


def test_expired_or_incomplete_price_never_matches():
    material = MaterialSpec(
        id="m1",
        mt_code="MT-01",
        name="古铜色不锈钢",
        grade="304",
        thickness_mm=1.2,
        finish="PVD",
        process="折弯",
    )
    expired = PriceEntry(
        id="expired",
        version="v1",
        approved=True,
        mt_code="MT-01",
        material="古铜色不锈钢",
        grade="304",
        thickness_mm=1.2,
        finish="PVD",
        process="折弯",
        pricing_method="按实际展开面积计算",
        unit="㎡",
        unit_price=500,
        tax_included=True,
        valid_to="2000-01-01",
        source="expired.xlsx",
    )
    book = PriceBook(version="v1", approved=True, source="expired.xlsx", entries=[expired])
    priced, issues = apply_price(item(), material, book, quote_date="2026-08-26")
    assert priced.unit_price is None
    assert any("过期" in issue for issue in issues)

    incomplete_material = material.model_copy(update={"process": None})
    priced, issues = apply_price(item(), incomplete_material, book, quote_date="1999-01-01")
    assert priced.unit_price is None
    assert any("材料证据缺少" in issue for issue in issues)


def test_portable_quote_workbook(tmp_path: Path):
    calculated = calculate_item(item(unit_price=500, status=ReviewStatus.PASS))
    output = build_quote_workbook([calculated], tmp_path / "quote.xlsx")
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["报价表", "来源追踪", "待确认", "运行信息"]
    assert [cell.value for cell in workbook["报价表"][1]] == QUOTE_HEADERS
    assert workbook["报价表"]["L2"].value.startswith("=IF(")
    assert workbook["报价表"]["P2"].value.startswith("=IF(")
    calculated_values = load_workbook(output, data_only=True)
    assert calculated_values["报价表"]["L2"].value == 4
    assert calculated_values["报价表"]["P2"].value == 2000


def test_quote_uses_audited_engineering_expression_formula_and_cache(tmp_path: Path):
    rows = [
        calculate_item(
            item(
                sequence=1,
                unfolded_spec=None,
                width_mm=3600,
                length_mm=900,
                quantity=1,
                unit="m",
                unit_price=100,
                engineering_quantity=999,
                engineering_quantity_expression="width_mm * quantity / 1000",
                engineering_quantity_basis="按已确认的平面长边计价",
                engineering_quantity_evidence_ids=["entity:width:3600"],
                component_id="component:width-axis",
                status=ReviewStatus.PASS,
            )
        ),
        calculate_item(
            item(
                sequence=2,
                unfolded_spec=None,
                width_mm=70,
                length_mm=3000,
                quantity=1,
                unit="m",
                unit_price=100,
                engineering_quantity_expression="length_mm * 2 / 1000",
                engineering_quantity_basis="两条物理线",
                engineering_quantity_evidence_ids=["entity:left", "entity:right"],
                component_id="component:two-runs",
                status=ReviewStatus.PASS,
            )
        ),
    ]
    proof_rows = {
        "component:width-axis": rows[0],
        "component:two-runs": rows[1],
    }
    engineering_edges = [
        EvidenceEdge(
            id=f"edge:{index}",
            relation="component_to_engineering_quantity_evidence",
            source_id=component_id,
            target_id=target_id,
            basis=[
                f"expression:{proof_rows[component_id].engineering_quantity_expression}",
                f"basis:{proof_rows[component_id].engineering_quantity_basis}",
                "source_file:file:synthetic",
                "sheet:sheet:elevation",
                f"handle:{index:X}",
                "bbox:(0, 0, 1, 1)",
            ],
            confidence=1,
            status=ReviewStatus.PASS,
        )
        for index, (component_id, target_id) in enumerate(
            (
                ("component:width-axis", "entity:width:3600"),
                ("component:two-runs", "entity:left"),
                ("component:two-runs", "entity:right"),
            ),
            start=1,
        )
    ]
    output = build_quote_workbook(
        rows,
        tmp_path / "custom-formula.xlsx",
        edges=engineering_edges,
    )
    formula_book = load_workbook(output, data_only=False)
    formulas = formula_book["报价表"]
    values = load_workbook(output, data_only=True)["报价表"]
    assert formulas["L2"].value == "=((I2*K2)/1000)"
    assert formulas["L3"].value == "=((J3*2)/1000)"
    assert values["L2"].value == 3.6
    assert values["L3"].value == 6
    assert values["P2"].value == 360
    assert values["P3"].value == 600
    assert "按已确认的平面长边计价" in formulas["L2"].comment.text
    assert "entity:width:3600" in formulas["L2"].comment.text
    assert formula_book["来源追踪"]["B2"].value == (
        "component_to_engineering_quantity_evidence"
    )
    assert formula_book["来源追踪"]["E2"].value == "engineering_quantity_evidence"
    assert formula_book["来源追踪"]["F2"].value == "width_mm * quantity / 1000"
    assert formula_book["来源追踪"]["I2"].value == "file:synthetic"
    assert formula_book["来源追踪"]["J2"].value == "sheet:elevation"
    assert formula_book["来源追踪"]["K2"].value == "entity:width:3600"


def test_quote_blocks_custom_expression_without_portable_pass_evidence_edges(
    tmp_path: Path,
):
    source = item(
        unfolded_spec=None,
        width_mm=70,
        length_mm=3000,
        quantity=1,
        unit="m",
        engineering_quantity_expression="length_mm*2/1000",
        engineering_quantity_basis="两条物理线",
        engineering_quantity_evidence_ids=["entity:left", "entity:right"],
        component_id="component:two-runs",
        status=ReviewStatus.PASS,
    )
    output = build_quote_workbook([source], tmp_path / "missing-proof-edges.xlsx")
    formulas = load_workbook(output, data_only=False)
    values = load_workbook(output, data_only=True)
    assert formulas["报价表"]["Q2"].value.startswith("[BLOCK]")
    assert "缺少PASS工程量证据边" in formulas["待确认"]["D2"].value
    assert values["报价表"]["L2"].value is None


def test_quote_rejects_pass_evidence_edge_for_an_old_expression(tmp_path: Path):
    source = item(
        unfolded_spec=None,
        width_mm=70,
        length_mm=3000,
        quantity=1,
        unit="m",
        engineering_quantity_expression="length_mm*2/1000",
        engineering_quantity_basis="当前两条物理线",
        engineering_quantity_evidence_ids=["entity:run"],
        component_id="component:run",
        status=ReviewStatus.PASS,
    )
    stale_edge = EvidenceEdge(
        id="edge:stale",
        relation="component_to_engineering_quantity_evidence",
        source_id="component:run",
        target_id="entity:run",
        basis=[
            "expression:length_mm*quantity/1000",
            "basis:旧版单线计价依据",
        ],
        confidence=1,
        status=ReviewStatus.PASS,
    )
    output = build_quote_workbook(
        [source],
        tmp_path / "stale-proof-edge.xlsx",
        edges=[stale_edge],
    )
    formulas = load_workbook(output, data_only=False)
    assert formulas["报价表"]["Q2"].value.startswith("[BLOCK]")
    assert "缺少PASS工程量证据边" in formulas["待确认"]["D2"].value


def test_quote_rejects_ambiguous_multi_identity_engineering_edge(tmp_path: Path):
    source = item(
        unfolded_spec=None,
        width_mm=70,
        length_mm=3000,
        quantity=1,
        unit="m",
        engineering_quantity_expression="length_mm*2/1000",
        engineering_quantity_basis="当前两条物理线",
        engineering_quantity_evidence_ids=["entity:run"],
        component_id="component:run",
        status=ReviewStatus.PASS,
    )
    ambiguous_edge = EvidenceEdge(
        id="edge:ambiguous",
        relation="component_to_engineering_quantity_evidence",
        source_id="component:run",
        target_id="entity:run",
        basis=[
            "expression:length_mm*2/1000",
            "expression:length_mm*quantity/1000",
            "basis:当前两条物理线",
            "basis:旧版单线计价依据",
        ],
        confidence=1,
        status=ReviewStatus.PASS,
    )
    output = build_quote_workbook(
        [source],
        tmp_path / "ambiguous-proof-edge.xlsx",
        edges=[ambiguous_edge],
    )
    formulas = load_workbook(output, data_only=False)
    assert formulas["报价表"]["Q2"].value.startswith("[BLOCK]")
    assert "缺少PASS工程量证据边" in formulas["待确认"]["D2"].value


def test_quote_downgrades_invalid_pass_expression_and_surfaces_pending_row(
    tmp_path: Path,
):
    invalid = item(
        unit="m",
        engineering_quantity=999,
        amount=999,
        engineering_quantity_expression="length_mm/0",
        engineering_quantity_basis="错误表达式",
        engineering_quantity_evidence_ids=["entity:1"],
        unit_price=100,
        status=ReviewStatus.PASS,
    )
    output = build_quote_workbook([invalid], tmp_path / "invalid-custom-expression.xlsx")
    formulas = load_workbook(output, data_only=False)
    values = load_workbook(output, data_only=True)
    assert formulas["报价表"]["L2"].value == '=""'
    assert values["报价表"]["L2"].value is None
    assert values["报价表"]["P2"].value is None
    assert formulas["报价表"]["Q2"].value.startswith("[BLOCK]")
    assert formulas["待确认"]["A2"].value == "算量项"
    assert "工程量表达式无效" in formulas["待确认"]["D2"].value


def test_optional_screenshot_evidence_sheet_embeds_images_and_marks_missing(
    tmp_path: Path,
):
    image_path = tmp_path / "evidence.png"
    _write_test_png(image_path)

    class DumpableEvidence:
        def model_dump(self, *, mode: str):
            assert mode == "json"
            return {
                "mt_code": "MT-02",
                "component_id": "component:2",
                "stage": "节点",
                "status": "REVIEW",
                "render_state": "FAILED",
                "drawing_number": "DT-02",
                "source_file_id": "file:cad",
                "focus_bbox": [10, 20, 30, 40],
                "entity_ids": ["entity:2"],
                "entity_handles": ["2B"],
                "context_image": str(image_path),
                "detail_image": str(tmp_path / "missing.png"),
                "render_reason": "渲染阶段未产出放大图",
            }

    report_path = tmp_path / "verification.json"
    output = build_quote_workbook(
        [calculate_item(item(unit_price=500, status=ReviewStatus.PASS))],
        tmp_path / "with-evidence.xlsx",
        evidence_records=[
            {
                "sequence": 7,
                "mt_code": "MT-01",
                "component_id": "component:1",
                "stage": "立面",
                "status": "REVIEW",
                "drawing_number": "EL-01",
                "source_file": "sample.dxf",
                "cad_bbox": [1, 2, 3, 4],
                "entity_ids": ["entity:1", "entity:1b"],
                "dxf_handles": ["1A", "1B"],
                "location_image": image_path,
                "zoom_image": image_path,
            },
            DumpableEvidence(),
        ],
        verification_report=report_path,
    )

    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook.sheetnames == [
            "报价表",
            "来源追踪",
            "待确认",
            "运行信息",
            "截图证据",
        ]
        sheet = workbook["截图证据"]
        assert [cell.value for cell in sheet[1]] == [
            "序号",
            "MT",
            "构件ID",
            "阶段",
            "状态",
            "图号",
            "来源文件",
            "CAD bbox",
            "实体ID",
            "DXF Handle",
            "定位图",
            "放大图",
            "缺图原因",
        ]
        assert len(sheet._images) == 3
        assert all(image.width <= 600 and image.height <= 300 for image in sheet._images)
        assert all(image.width / image.height == pytest.approx(2.0) for image in sheet._images)
        assert sheet.row_dimensions[2].height >= 180
        assert sheet["L3"].value.startswith("缺图：")
        assert "渲染阶段未产出放大图" in sheet["M3"].value
        assert sheet["E3"].value == "FAILED"
        assert sheet["G3"].value == "file:cad"
        assert sheet["H3"].value == "[10, 20, 30, 40]"
        assert sheet["J3"].value == "2B"
        assert sheet["L3"].font.color.rgb.endswith("C00000")
        assert sheet["M3"].font.color.rgb.endswith("C00000")
    finally:
        workbook.close()

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["evidence"] == {
        "record_count": 2,
        "image_count": 3,
        "missing_rows": [3],
    }
    assert all(
        report["checks"][name]
        for name in (
            "sheet_order",
            "evidence_headers",
            "evidence_image_count",
            "evidence_missing_rows",
            "evidence_missing_red",
            "evidence_image_scale",
        )
    )


def test_partial_quote_preserves_full_takeoff_evidence(tmp_path: Path):
    candidate = MeasurementCandidate(
        id="measurement:1",
        component_id="component:1",
        role="length",
        raw_value="5000",
        numeric_value=5000,
        unit="mm",
        source_file_id="file:1",
        sheet_id="sheet:1",
        entity_ids=["entity:1"],
    )
    edge = EvidenceEdge(
        id="edge:1",
        relation="component_to_dimension",
        source_id="component:1",
        target_id=candidate.id,
    )
    issue = RunIssue(
        stage="takeoff",
        severity=Severity.WARNING,
        code="REVIEW",
        message="需要复核",
    )
    payload_path = tmp_path / "takeoff.json"
    payload_path.write_text(
        json.dumps(
            {
                "items": [item(component_id="component:1").model_dump(mode="json")],
                "evidence_edges": [edge.model_dump(mode="json")],
                "measurements": [candidate.model_dump(mode="json")],
                "issues": [issue.model_dump(mode="json")],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    output = tmp_path / "partial.xlsx"
    assert command_quote(Namespace(takeoff=payload_path, out=output)) == 0
    workbook = load_workbook(output)
    assert workbook["来源追踪"]["A2"].value == "edge:1"
    assert workbook["来源追踪"]["I2"].value == "file:1"
    assert workbook["待确认"]["D3"].value == "需要复核"


def test_quote_total_excludes_unresolved_rows(tmp_path: Path):
    passed = calculate_item(item(sequence=1, unit_price=10, status=ReviewStatus.PASS))
    blocked = calculate_item(item(sequence=2, unit_price=10, status=ReviewStatus.BLOCK))
    output = build_quote_workbook([passed, blocked], tmp_path / "safe-total.xlsx")
    values = load_workbook(output, data_only=True)["报价表"]
    assert values["P2"].value == 40
    assert values["P3"].value is None
    assert values["P4"].value == 40


def test_quote_escapes_untrusted_formula_text(tmp_path: Path):
    malicious = calculate_item(
        item(
            name="=WEBSERVICE(\"https://example.invalid\")",
            plan_location="+cmd|' /C calc'!A0",
            note="@SUM(1,1)",
            status=ReviewStatus.BLOCK,
        )
    )
    output = build_quote_workbook([malicious], tmp_path / "literal-text.xlsx")
    sheet = load_workbook(output, data_only=False)["报价表"]
    for coordinate in ("B2", "E2", "Q2"):
        assert sheet[coordinate].data_type != "f"


def test_quote_amount_rounding_matches_deterministic_json(tmp_path: Path):
    rows = [
        calculate_item(
            item(
                sequence=sequence,
                unfolded_spec=None,
                width_mm=1,
                length_mm=5000,
                quantity=1,
                unit_price=1,
                status=ReviewStatus.PASS,
            )
        )
        for sequence in (1, 2)
    ]
    output = build_quote_workbook(rows, tmp_path / "rounding.xlsx")
    values = load_workbook(output, data_only=True)["报价表"]
    assert [row.amount for row in rows] == [0.01, 0.01]
    assert values["P2"].value == 0.01
    assert values["P3"].value == 0.01
    assert values["P4"].value == 0.02


def test_evaluation_separates_precision_and_automation():
    gold = [calculate_item(item(status=ReviewStatus.PASS))]
    predicted = [calculate_item(item(status=ReviewStatus.PASS))]
    report = evaluate_takeoff(predicted, gold)
    assert report["pass_precision"] == 1
    assert report["automation_rate"] == 1
