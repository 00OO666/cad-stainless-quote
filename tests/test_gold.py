"""Gold-workbook import and audit regression tests."""

from __future__ import annotations

import json
from base64 import b64decode
from pathlib import Path

import pytest
import xlsxwriter
from cadquote.gold import import_gold_workbook
from openpyxl import Workbook


def _write_canonical_xlsx(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "报价表"
    sheet.append(
        [
            "序号",
            "名称",
            "MT编号",
            "材料",
            "平面图位置",
            "对应立面",
            "对应节点",
            "展开规格",
            "宽",
            "长度",
            "数量",
            "工程量",
            "单位",
            "计价方式",
            "单价",
            "金额",
            "备注",
        ]
    )
    sheet.append(
        [
            1,
            "墙面收口",
            "mt1",
            "304不锈钢",
            "门厅",
            "EL-01",
            "DT-01",
            "50+150",
            200,
            1000,
            2,
            0.4,
            "㎡",
            "按展开面积",
            100,
            "=L2*O2",
            "人工待审",
        ]
    )
    sheet.append(
        [
            2,
            "踢脚线",
            "MT-02",
            "304不锈钢",
            "走廊",
            "EL-02",
            "DT-02",
            None,
            80,
            2000,
            2,
            2,
            "m",
            "按米",
            None,
            None,
            None,
        ]
    )
    sheet["L2"].number_format = sheet["L3"].number_format = "0.000"
    # A stray style at Excel's last column must not inflate the logical table.
    sheet["XFD1"].number_format = "0.00"
    workbook.save(path)


def _write_composite_xlsx(path: Path, image_path: Path) -> None:
    image_path.write_bytes(
        b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk"
            "/x8AAusB9Y9Z4WQAAAAASUVORK5CYII="
        )
    )
    workbook = xlsxwriter.Workbook(path)
    sheet = workbook.add_worksheet("候选金标")
    sheet.write("A1", "合成报价表")
    for column, value in enumerate(
        [
            "序号",
            "材料名称",
            "位置（图号）",
            "立面示意图",
            "节点示意图",
            "材料代号",
            "材料特征描述",
            "展开尺寸",
        ]
    ):
        sheet.write(2, column, value)
    sheet.merge_range("I3:K3", "尺寸")
    sheet.write_row("I4", ["宽/展开尺寸", "高/长", "计量"])
    for coordinate, value in (
        ("L3", "数量"),
        ("M3", "单位"),
        ("N3", "计算方式"),
        ("O3", "单价（含税）"),
        ("P3", "合计"),
        ("Q3", "备注"),
    ):
        sheet.write(coordinate, value)
    sheet.write("A5", "区域甲")
    sheet.write_row(
        "A6",
        [
            1,
            "通用收边",
            "E-01",
            None,
            None,
            "GC-SS-201",
            "拉丝不锈钢",
            "10+20+10",
        ],
    )
    sheet.write_array_formula("I6:I6", "=EVALUATE(H6)", None, 40)
    sheet.write("J6", 2000)
    sheet.write("K6", 2)
    sheet.write_formula("L6", '=IF(M6="㎡",I6*J6*K6/1000000,K6)', None, 0.16)
    sheet.write("M6", "㎡")
    sheet.write("N6", "按展开面积")
    sheet.write("O6", 100)
    sheet.write_formula("P6", "=L6*O6", None, 16)
    sheet.write("Q6", "合成待审")
    sheet.insert_image("D6", str(image_path), {"x_scale": 8, "y_scale": 8})
    sheet.insert_image("E6", str(image_path), {"x_scale": 8, "y_scale": 8})
    sheet.write("A7", "小计")
    sheet.write_formula("P7", "=SUM(P6:P6)", None, 16)
    workbook.close()


def test_import_canonical_xlsx_preserves_evidence_and_never_passes(tmp_path: Path) -> None:
    source = tmp_path / "gold.xlsx"
    _write_canonical_xlsx(source)

    result = import_gold_workbook(source)

    assert result.workbook_format == "xlsx"
    assert result.schema_version == "1.1"
    assert result.summary.row_count == 2
    assert result.sheets[0].max_column == 17
    assert result.summary.mt_distribution == {"MT-01": 1, "MT-02": 1}
    assert result.summary.pass_count == 0
    assert result.summary.review_count == 1
    assert result.summary.block_count == 1
    assert result.summary.quantity_mismatch_count == 1
    assert result.rows[0].quantity_audit == "MATCH"
    assert result.rows[0].item.status.value == "REVIEW"
    assert result.rows[0].field_cells["mt_code"] == ["C2"]
    assert len(result.rows[0].raw_cells) == 17
    assert result.rows[0].raw_cells[15].coordinate == "P2"
    assert result.rows[0].raw_cells[15].formula == "=L2*O2"
    assert result.rows[1].recalculated_engineering_quantity == 4
    assert result.rows[1].reported_engineering_quantity == 2
    assert {issue.code for issue in result.issues} == {
        "AMOUNT_NOT_CALCULABLE",
        "QUANTITY_FORMULA_MISMATCH",
    }

    payload = json.loads(result.to_json())
    assert payload["source_sha256"] == result.source_sha256
    assert payload["rows"][0]["raw_cells"][2]["raw_value"] == "mt1"
    output = result.write_json(tmp_path / "gold.json")
    assert output.is_file()


def test_import_composite_headers_images_and_non_mt_material_code(tmp_path: Path) -> None:
    source = tmp_path / "composite.xlsx"
    _write_composite_xlsx(source, tmp_path / "pixel.png")

    result = import_gold_workbook(source)

    assert result.summary.row_count == 1
    assert result.sheets[0].header_rows == [3, 4]
    assert result.sheets[0].field_columns["quantity"] == 11
    assert result.sheets[0].field_columns["engineering_quantity"] == 12
    assert result.rows[0].row == 6
    assert result.rows[0].source_material_code == "GC-SS-201"
    assert result.rows[0].item.unfolded_spec == "10+20+10"
    assert result.rows[0].field_cells["unfolded_spec"] == ["H6"]
    assert result.rows[0].mt_code_source is None
    assert result.rows[0].item.mt_code == ""
    assert result.rows[0].item.status.value == "BLOCK"
    assert result.rows[0].quantity_audit == "MATCH"
    assert {issue.code for issue in result.issues} == {"MISSING_REQUIRED_FIELD"}

    assert len(result.rows[0].image_evidence) == 2
    assert {image.anchor_coordinate for image in result.rows[0].image_evidence} == {
        "D6",
        "E6",
    }
    assert {image.category for image in result.rows[0].image_evidence} == {
        "elevation",
        "detail",
    }
    assert all(image.source_type == "embedded" for image in result.rows[0].image_evidence)
    width_cell = next(cell for cell in result.rows[0].raw_cells if cell.coordinate == "I6")
    assert width_cell.formula_kind == "array"
    assert width_cell.formula == "=EVALUATE(H6)"
    assert width_cell.formula_range == "I6"
    assert width_cell.raw_value == 40
    engineering_cell = next(cell for cell in result.rows[0].raw_cells if cell.coordinate == "L6")
    assert engineering_cell.formula is not None
    assert engineering_cell.raw_value == pytest.approx(0.16)


def test_explicit_mt_material_code_is_preserved_and_can_supply_mt(tmp_path: Path) -> None:
    source = tmp_path / "material-code.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "序号",
            "项目",
            "使用位置",
            "材料编号",
            "材料表面",
            "宽度mm",
            "长度mm",
            "件数",
            "单位",
            "工程量",
        ]
    )
    sheet.append([1, "装饰线", "区域乙", "MT-7", "镜面不锈钢", 50, 3000, 2, "m", 6])
    workbook.save(source)

    result = import_gold_workbook(source)

    assert result.summary.row_count == 1
    assert result.rows[0].source_material_code == "MT-7"
    assert result.rows[0].mt_code_source == "material_code"
    assert result.rows[0].item.mt_code == "MT-07"
    assert result.rows[0].field_cells["mt_code"] == ["D2"]


def test_abnormal_unit_and_negative_engineering_quantity_are_audited(
    tmp_path: Path,
) -> None:
    source = tmp_path / "abnormal.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "序号",
            "名称",
            "MT编号",
            "宽",
            "长度",
            "数量",
            "工程量",
            "单位",
        ]
    )
    sheet.append([1, "异形板", "MT-3", 100, 1000, 1, -0.1, "公斤"])
    workbook.save(source)

    result = import_gold_workbook(source)

    assert result.summary.row_count == 1
    assert result.rows[0].reported_engineering_quantity == pytest.approx(-0.1)
    assert result.rows[0].item.engineering_quantity is None
    assert result.rows[0].item.status.value == "BLOCK"
    assert {issue.code for issue in result.issues} >= {
        "INVALID_NUMERIC_VALUE",
        "UNRECOGNIZED_UNIT",
        "QUANTITY_NOT_CALCULABLE",
    }


def test_rejects_non_excel_file(tmp_path: Path) -> None:
    source = tmp_path / "gold.csv"
    source.write_text("序号,名称", encoding="utf-8")
    with pytest.raises(ValueError, match="unsupported gold workbook format"):
        import_gold_workbook(source)
